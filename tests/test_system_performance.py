#!/usr/bin/env python3
"""
企业微信群聊分析服务 - 系统性能完整测试脚本

功能：
1. 从Java数据源接口拉取真实消息数据
2. 统计群聊数量、消息数量
3. 调用批量分析接口进行完整分析
4. 统计分析耗费时间
5. 统计Token使用情况
6. 测试所有分析类型（情感、敏感词、摘要、高频词、漏回分析）
7. 生成详细的性能测试报告

使用方式：
    python tests/test_system_performance.py

环境变量：
    JAVA_DATA_SOURCE_URL - Java数据源地址（默认: http://192.168.0.129:8081/qxChat/）
    ANALYSIS_SERVICE_URL - 分析服务地址（默认: http://localhost:8000）
"""

import json
import sys
import time
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, field
from datetime import datetime
from collections import defaultdict
import argparse

import httpx

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

sys.path.insert(0, str(__file__).rsplit("/tests", 1)[0])

from config.settings import settings
from utils.llm_client import LLMClient


@dataclass
class RoomData:
    room_id: str
    room_name: Optional[str]
    messages: List[Dict[str, Any]]
    members: Dict[str, Dict[str, Any]] = field(default_factory=dict)


@dataclass
class PerformanceMetrics:
    total_rooms: int = 0
    total_messages: int = 0
    analysis_time: float = 0.0
    total_tokens: int = 0
    prompt_tokens: int = 0
    completion_tokens: int = 0
    success_count: int = 0
    failed_count: int = 0
    analysis_types: List[str] = field(default_factory=list)
    room_details: List[Dict[str, Any]] = field(default_factory=list)
    room_members: Dict[str, Dict[str, Dict[str, Any]]] = field(default_factory=dict)


class SystemPerformanceTester:
    def __init__(
        self,
        data_source_url: Optional[str] = None,
        analysis_service_url: Optional[str] = None,
        max_concurrent: int = 5,
        max_rooms: Optional[int] = None,
    ):
        self.data_source_url = data_source_url or settings.JAVA_DATA_SOURCE_URL
        self.analysis_service_url = analysis_service_url or "http://localhost:8000"
        self.max_concurrent = max_concurrent
        self.max_rooms = max_rooms
        self.http_client = httpx.Client(timeout=1200)
        self.llm_client = LLMClient()
        self.metrics = PerformanceMetrics()

    def fetch_data_from_java(self) -> List[RoomData]:
        print(f"\n{'='*80}")
        print(f"步骤1: 从Java数据源拉取数据")
        print(f"{'='*80}")
        print(f"数据源地址: {self.data_source_url}")
        print()

        try:
            response = self.http_client.get(self.data_source_url)
            response.raise_for_status()
            raw_data = response.json()

            messages = raw_data.get("data", [])
            print(f"✓ 数据拉取成功!")
            print(f"  - 总消息数: {len(messages)}")

            if not messages:
                print("✗ 没有消息数据")
                return []

            rooms_dict = defaultdict(lambda: {"room_id": "", "room_name": "", "messages": [], "members": defaultdict(dict)})

            for msg in messages:
                room_id = msg.get("roomid", "unknown")
                if not rooms_dict[room_id]["room_id"]:
                    rooms_dict[room_id]["room_id"] = room_id
                    room_name = msg.get("re_truename") or msg.get("roomname") or f"群-{room_id[:8]}"
                    rooms_dict[room_id]["room_name"] = room_name
                rooms_dict[room_id]["messages"].append(msg)
                
                members_str = msg.get("members", "")
                if members_str:
                    try:
                        members_list = json.loads(members_str) if isinstance(members_str, str) else members_str
                        if isinstance(members_list, list):
                            for m in members_list:
                                userid = m.get("userid", "")
                                if userid:
                                    rooms_dict[room_id]["members"][userid] = m
                    except:
                        pass

            rooms = []
            for room_id, room_info in rooms_dict.items():
                rooms.append(RoomData(
                    room_id=room_info["room_id"],
                    room_name=room_info["room_name"],
                    messages=room_info["messages"],
                    members=room_info["members"]
                ))
                self.metrics.room_members[room_info["room_id"]] = room_info["members"]

            if self.max_rooms and len(rooms) > self.max_rooms:
                rooms = rooms[:self.max_rooms]
                print(f"  - 限制测试群聊数量: {self.max_rooms}")

            self.metrics.total_rooms = len(rooms)
            self.metrics.total_messages = sum(len(room.messages) for room in rooms)

            print(f"  - 群聊数量: {len(rooms)}")
            for room in rooms:
                print(f"    - {room.room_name}: {len(room.messages)}条消息")

            return rooms

        except httpx.ConnectError as e:
            print(f"✗ 连接失败: {e}")
            print(f"  请确认Java服务 ({self.data_source_url}) 已启动")
            sys.exit(1)
        except httpx.TimeoutException:
            print(f"✗ 请求超时")
            sys.exit(1)
        except Exception as e:
            print(f"✗ 请求失败: {e}")
            sys.exit(1)

    def batch_analyze(self, rooms: List[RoomData], analysis_types: List[str]) -> Dict[str, Any]:
        print(f"\n{'='*80}")
        print(f"步骤2: 调用批量分析接口")
        print(f"{'='*80}")
        print(f"分析服务地址: {self.analysis_service_url}/api/v1/chat/batch-analyze")
        print(f"群聊数量: {len(rooms)}")
        print(f"分析类型: {', '.join(analysis_types)}")
        print(f"并发数: {self.max_concurrent}")
        print()

        if not rooms:
            print("✗ 没有群聊数据可分析")
            return {}

        rooms_data = []
        for room in rooms:
            room_data = {
                "room_id": room.room_id,
                "room_name": room.room_name,
                "messages": room.messages
            }
            
            if room.members:
                all_members = list(room.members.values())
                if all_members:
                    room_data["members"] = json.dumps(all_members, ensure_ascii=False)
            
            rooms_data.append(room_data)

        request_body = {
            "rooms": rooms_data,
            "analysis_type": analysis_types,
            "max_concurrent": self.max_concurrent
        }

        try:
            self.llm_client.reset_token_stats()
            
            start_time = time.time()
            response = self.http_client.post(
                f"{self.analysis_service_url}/api/v1/chat/batch-analyze",
                json=request_body,
                timeout=1200,
            )
            elapsed = time.time() - start_time

            response.raise_for_status()
            result = response.json()

            self.metrics.analysis_time = elapsed
            token_stats = self.llm_client.token_stats
            self.metrics.total_tokens = token_stats["total_tokens"]
            self.metrics.prompt_tokens = token_stats["total_prompt_tokens"]
            self.metrics.completion_tokens = token_stats["total_completion_tokens"]
            self.metrics.analysis_types = analysis_types

            print(f"✓ 批量分析完成! (耗时: {elapsed:.2f}秒)")
            print(f"  - 返回码: {result.get('code')}")
            print(f"  - 状态: {result.get('message')}")

            if "data" in result:
                data = result["data"]
                self.metrics.success_count = data.get('success_count', 0)
                self.metrics.failed_count = data.get('failed_count', 0)
                print(f"  - 总群数: {data.get('total_rooms', 0)}")
                print(f"  - 成功: {self.metrics.success_count}")
                print(f"  - 失败: {self.metrics.failed_count}")
                print(f"  - 总耗时: {data.get('elapsed_seconds', 0)}秒")

            return result

        except httpx.ConnectError as e:
            print(f"✗ 连接失败: {e}")
            print(f"  请确认Python分析服务 ({self.analysis_service_url}) 已启动")
            print(f"  启动命令: uvicorn api.main:app --host 0.0.0.0 --port 8000")
            sys.exit(1)
        except httpx.TimeoutException:
            print(f"✗ 请求超时 (分析时间超过1200秒)")
            sys.exit(1)
        except Exception as e:
            print(f"✗ 批量分析失败: {e}")
            sys.exit(1)

    def print_performance_report(self, result: Dict[str, Any]):
        print(f"\n{'='*80}")
        print(f"步骤3: 系统性能测试报告")
        print(f"{'='*80}")

        print(f"\n【测试概览】")
        print(f"  测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"  数据源: {self.data_source_url}")
        print(f"  分析服务: {self.analysis_service_url}")

        print(f"\n【数据统计】")
        print(f"  群聊数量: {self.metrics.total_rooms}")
        print(f"  消息总数: {self.metrics.total_messages}")
        if self.metrics.total_rooms > 0:
            avg_messages = self.metrics.total_messages / self.metrics.total_rooms
            print(f"  平均每群消息数: {avg_messages:.1f}")

        print(f"\n【分析性能】")
        print(f"  分析类型: {', '.join(self.metrics.analysis_types)}")
        print(f"  分析耗时: {self.metrics.analysis_time:.2f}秒")
        if self.metrics.total_rooms > 0:
            avg_time = self.metrics.analysis_time / self.metrics.total_rooms
            print(f"  平均每群耗时: {avg_time:.2f}秒")
        if self.metrics.total_messages > 0:
            time_per_msg = self.metrics.analysis_time / self.metrics.total_messages * 1000
            print(f"  平均每条消息耗时: {time_per_msg:.2f}毫秒")

        print(f"\n【Token使用统计】")
        print(f"  总Token数: {self.metrics.total_tokens:,}")
        print(f"  提示Token数: {self.metrics.prompt_tokens:,}")
        print(f"  完成Token数: {self.metrics.completion_tokens:,}")
        if self.metrics.total_rooms > 0:
            avg_tokens = self.metrics.total_tokens / self.metrics.total_rooms
            print(f"  平均每群Token数: {avg_tokens:.1f}")
        if self.metrics.total_messages > 0:
            tokens_per_msg = self.metrics.total_tokens / self.metrics.total_messages
            print(f"  平均每条消息Token数: {tokens_per_msg:.2f}")

        print(f"\n【成功率统计】")
        print(f"  成功分析: {self.metrics.success_count}")
        print(f"  失败分析: {self.metrics.failed_count}")
        if self.metrics.total_rooms > 0:
            success_rate = (self.metrics.success_count / self.metrics.total_rooms) * 100
            print(f"  成功率: {success_rate:.1f}%")

        if "data" in result:
            data = result["data"]
            results = data.get("results", [])

            print(f"\n【各群详细分析结果】")
            for i, room_result in enumerate(results, 1):
                room_name = room_result.get("room_name", "未知")
                status = room_result.get("status", "unknown")
                room_data = room_result.get("data", {})

                status_icon = "✓" if status == "success" else "✗"
                print(f"\n  {i}. {room_name} [{status_icon}]")

                if status == "success" and room_data:
                    msg_count = room_data.get("message_count", 0)
                    print(f"     消息数: {msg_count}")

                    sentiment = room_data.get("sentiment", {})
                    if sentiment:
                        summary = sentiment.get("summary", {})
                        customer_summary = summary.get("customer", {})
                        employee_summary = summary.get("employee", {})
                        print(f"     情感分析: 客户好评={customer_summary.get('good_reviews', 0)}, 客户差评={customer_summary.get('bad_reviews', 0)}, 员工积极={employee_summary.get('positive', 0)}, 员工恶劣={employee_summary.get('bad_attitude', 0)}")

                    sensitive = room_data.get("sensitive_words", {})
                    if sensitive:
                        print(f"     敏感词: {sensitive.get('total_hits', 0)}次命中")

                    summary_text = room_data.get("summary")
                    if summary_text:
                        print(f"     摘要: {summary_text[:100]}{'...' if len(summary_text) > 100 else ''}")

                    highfreq = room_data.get("high_freq_words", {})
                    if highfreq and highfreq.get("words"):
                        words = highfreq["words"][:5]
                        word_str = ", ".join([f"{w['word']}({w['count']})" for w in words])
                        print(f"     高频词: {word_str}")

                    unanswered = room_data.get("unanswered_status", {})
                    if unanswered:
                        is_missed = unanswered.get("is_missed", False)
                        risk_level = unanswered.get("risk_level", "low")
                        missed_count = len(unanswered.get("missed_messages", []))
                        print(f"     漏回状态: {'存在漏回' if is_missed else '无漏回'} (风险等级: {risk_level}, 漏回消息数: {missed_count})")
                        if unanswered.get("suggested_action"):
                            print(f"     建议操作: {unanswered.get('suggested_action')}")
                else:
                    error = room_result.get("error_message", "未知错误")
                    print(f"     错误: {error}")

        print(f"\n{'='*80}")
        print("系统性能测试完成!")
        print(f"{'='*80}\n")

    def save_report_to_file(self, result: Dict[str, Any], output_file: str = "performance_report.json"):
        report = {
            "test_time": datetime.now().isoformat(),
            "data_source": self.data_source_url,
            "analysis_service": self.analysis_service_url,
            "metrics": {
                "total_rooms": self.metrics.total_rooms,
                "total_messages": self.metrics.total_messages,
                "analysis_time_seconds": round(self.metrics.analysis_time, 2),
                "total_tokens": self.metrics.total_tokens,
                "prompt_tokens": self.metrics.prompt_tokens,
                "completion_tokens": self.metrics.completion_tokens,
                "success_count": self.metrics.success_count,
                "failed_count": self.metrics.failed_count,
                "analysis_types": self.metrics.analysis_types,
            },
            "analysis_result": result.get("data", {})
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        print(f"✓ 性能报告已保存到: {output_file}")

    def generate_markdown_report(self, result: Dict[str, Any], output_file: str = "系统测试报告.md"):
        md_content = f"""# 企业微信群聊分析服务 - 系统测试报告

## 一、测试概览

| 项目 | 内容 |
| :--- | :--- |
| 测试时间 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} |
| 数据源地址 | {self.data_source_url} |
| 分析服务地址 | {self.analysis_service_url} |
| 并发数 | {self.max_concurrent} |

## 二、数据统计

| 指标 | 数值 |
| :--- | :--- |
| 群聊数量 | {self.metrics.total_rooms} |
| 消息总数 | {self.metrics.total_messages} |
| 平均每群消息数 | {(self.metrics.total_messages / self.metrics.total_rooms if self.metrics.total_rooms > 0 else 0):.1f} |

## 三、分析性能

| 指标 | 数值 |
| :--- | :--- |
| 分析类型 | {', '.join(self.metrics.analysis_types)} |
| 总分析耗时 | {self.metrics.analysis_time:.2f}秒 |
| 平均每群耗时 | {(self.metrics.analysis_time / self.metrics.total_rooms if self.metrics.total_rooms > 0 else 0):.2f}秒 |
| 平均每条消息耗时 | {(self.metrics.analysis_time / self.metrics.total_messages * 1000 if self.metrics.total_messages > 0 else 0):.2f}毫秒 |

## 四、Token使用统计

| 指标 | 数值 |
| :--- | :--- |
| 总Token数 | {self.metrics.total_tokens:,} |
| 提示Token数 | {self.metrics.prompt_tokens:,} |
| 完成Token数 | {self.metrics.completion_tokens:,} |
| 平均每群Token数 | {(self.metrics.total_tokens / self.metrics.total_rooms if self.metrics.total_rooms > 0 else 0):.1f} |
| 平均每条消息Token数 | {(self.metrics.total_tokens / self.metrics.total_messages if self.metrics.total_messages > 0 else 0):.2f} |

## 五、成功率统计

| 指标 | 数值 |
| :--- | :--- |
| 成功分析 | {self.metrics.success_count} |
| 失败分析 | {self.metrics.failed_count} |
| 成功率 | {(self.metrics.success_count / self.metrics.total_rooms * 100 if self.metrics.total_rooms > 0 else 0):.1f}% |

## 六、各群详细分析结果

"""
        if "data" in result:
            data = result["data"]
            results = data.get("results", [])
            
            for i, room_result in enumerate(results, 1):
                room_name = room_result.get("room_name", "未知")
                status = room_result.get("status", "unknown")
                room_data = room_result.get("data", {})
                
                status_icon = "✓" if status == "success" else "✗"
                md_content += f"### {i}. {room_name} [{status_icon}]\n\n"
                
                if status == "success" and room_data:
                    msg_count = room_data.get("message_count", 0)
                    md_content += f"**消息数**: {msg_count}\n\n"
                    
                    sentiment = room_data.get("sentiment", {})
                    if sentiment:
                        summary = sentiment.get("summary", {})
                        customer_summary = summary.get("customer", {})
                        employee_summary = summary.get("employee", {})
                        md_content += f"**情感分析**: 客户好评={customer_summary.get('good_reviews', 0)}, 客户差评={customer_summary.get('bad_reviews', 0)}, 员工积极={employee_summary.get('positive', 0)}, 员工恶劣={employee_summary.get('bad_attitude', 0)}\n\n"
                    
                    sensitive = room_data.get("sensitive_words", {})
                    if sensitive:
                        md_content += f"**敏感词**: {sensitive.get('total_hits', 0)}次命中\n\n"
                    
                    summary_text = room_data.get("summary")
                    if summary_text:
                        md_content += f"**摘要**: {summary_text}\n\n"
                    
                    highfreq = room_data.get("high_freq_words", {})
                    if highfreq and highfreq.get("words"):
                        words = highfreq["words"][:5]
                        word_str = ", ".join([f"{w['word']}({w['count']})" for w in words])
                        md_content += f"**高频词**: {word_str}\n\n"
                    
                    unanswered = room_data.get("unanswered_status", {})
                    if unanswered:
                        is_missed = unanswered.get("is_missed", False)
                        risk_level = unanswered.get("risk_level", "low")
                        missed_count = len(unanswered.get("missed_messages", []))
                        md_content += f"**漏回状态**: {'存在漏回' if is_missed else '无漏回'} (风险等级: {risk_level}, 漏回消息数: {missed_count})\n\n"
                        if unanswered.get("suggested_action"):
                            md_content += f"**建议操作**: {unanswered.get('suggested_action')}\n\n"
                else:
                    error = room_result.get("error_message", "未知错误")
                    md_content += f"**错误**: {error}\n\n"
                
                md_content += "---\n\n"

        md_content += self._generate_evaluation_report()
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(md_content)
        
        print(f"✓ 测试文档已保存到: {output_file}")

    def _generate_evaluation_report(self) -> str:
        avg_time_per_room = self.metrics.analysis_time / self.metrics.total_rooms if self.metrics.total_rooms > 0 else 0
        avg_tokens_per_room = self.metrics.total_tokens / self.metrics.total_rooms if self.metrics.total_rooms > 0 else 0
        success_rate = (self.metrics.success_count / self.metrics.total_rooms * 100) if self.metrics.total_rooms > 0 else 0
        
        performance_grade = "优秀" if avg_time_per_room < 5 else "良好" if avg_time_per_room < 10 else "一般" if avg_time_per_room < 20 else "需优化"
        reliability_grade = "优秀" if success_rate >= 95 else "良好" if success_rate >= 90 else "一般" if success_rate >= 80 else "需改进"
        efficiency_grade = "优秀" if avg_tokens_per_room < 1000 else "良好" if avg_tokens_per_room < 2000 else "一般" if avg_tokens_per_room < 3000 else "需优化"
        
        overall_score = 0
        if performance_grade == "优秀": overall_score += 30
        elif performance_grade == "良好": overall_score += 25
        elif performance_grade == "一般": overall_score += 20
        else: overall_score += 10
        
        if reliability_grade == "优秀": overall_score += 40
        elif reliability_grade == "良好": overall_score += 35
        elif reliability_grade == "一般": overall_score += 25
        else: overall_score += 15
        
        if efficiency_grade == "优秀": overall_score += 30
        elif efficiency_grade == "良好": overall_score += 25
        elif efficiency_grade == "一般": overall_score += 20
        else: overall_score += 10
        
        overall_grade = "优秀" if overall_score >= 90 else "良好" if overall_score >= 75 else "一般" if overall_score >= 60 else "需改进"
        
        evaluation = f"""## 七、系统评估报告

### 7.1 性能评估

| 评估项 | 指标 | 评级 |
| :--- | :--- | :--- |
| 平均分析速度 | {avg_time_per_room:.2f}秒/群 | {performance_grade} |
| 并发处理能力 | {self.max_concurrent}个并发 | {'优秀' if self.max_concurrent >= 10 else '良好' if self.max_concurrent >= 5 else '一般'} |
| 吞吐量 | {self.metrics.total_rooms / self.metrics.analysis_time * 60:.1f}群/分钟 | {'优秀' if self.metrics.total_rooms / self.metrics.analysis_time * 60 >= 10 else '良好' if self.metrics.total_rooms / self.metrics.analysis_time * 60 >= 5 else '一般'} |

**性能分析**:
- 系统在{self.metrics.analysis_time:.2f}秒内完成了{self.metrics.total_rooms}个群聊的分析
- 平均每个群聊耗时{avg_time_per_room:.2f}秒，{'表现优秀' if avg_time_per_room < 5 else '表现良好' if avg_time_per_room < 10 else '有优化空间'}
- 系统吞吐量为{self.metrics.total_rooms / self.metrics.analysis_time * 60:.1f}群/分钟

### 7.2 可靠性评估

| 评估项 | 指标 | 评级 |
| :--- | :--- | :--- |
| 成功率 | {success_rate:.1f}% | {reliability_grade} |
| 失败数量 | {self.metrics.failed_count}个 | {'优秀' if self.metrics.failed_count == 0 else '良好' if self.metrics.failed_count < 3 else '需改进'} |
| 稳定性 | {'稳定' if success_rate >= 95 else '较稳定' if success_rate >= 90 else '不稳定'} | {reliability_grade} |

**可靠性分析**:
- 系统成功分析了{self.metrics.success_count}个群聊，失败{self.metrics.failed_count}个
- 成功率达到{success_rate:.1f}%，{'表现优秀' if success_rate >= 95 else '表现良好' if success_rate >= 90 else '需要改进'}
- {'系统运行稳定，无失败案例' if self.metrics.failed_count == 0 else f'有{self.metrics.failed_count}个群聊分析失败，建议检查失败原因'}

### 7.3 成本评估

| 评估项 | 指标 | 评级 |
| :--- | :--- | :--- |
| Token使用效率 | {avg_tokens_per_room:.1f} tokens/群 | {efficiency_grade} |
| 提示Token占比 | {(self.metrics.prompt_tokens / self.metrics.total_tokens * 100 if self.metrics.total_tokens > 0 else 0):.1f}% | {'优秀' if (self.metrics.prompt_tokens / self.metrics.total_tokens * 100 if self.metrics.total_tokens > 0 else 0) < 70 else '良好'} |
| 完成Token占比 | {(self.metrics.completion_tokens / self.metrics.total_tokens * 100 if self.metrics.total_tokens > 0 else 0):.1f}% | {'优秀' if (self.metrics.completion_tokens / self.metrics.total_tokens * 100 if self.metrics.total_tokens > 0 else 0) < 30 else '良好'} |

**成本分析**:
- 本次测试共消耗{self.metrics.total_tokens:,}个Token
- 平均每个群聊消耗{avg_tokens_per_room:.1f}个Token
- 提示Token占比{(self.metrics.prompt_tokens / self.metrics.total_tokens * 100 if self.metrics.total_tokens > 0 else 0):.1f}%，完成Token占比{(self.metrics.completion_tokens / self.metrics.total_tokens * 100 if self.metrics.total_tokens > 0 else 0):.1f}%
- {'Token使用效率优秀，成本控制良好' if avg_tokens_per_room < 2000 else 'Token使用效率良好，建议优化Prompt以降低成本'}

### 7.4 综合评分

| 维度 | 权重 | 得分 | 加权得分 |
| :--- | :--- | :--- | :--- |
| 性能 | 30% | {30 if performance_grade == '优秀' else 25 if performance_grade == '良好' else 20 if performance_grade == '一般' else 10} | {(30 if performance_grade == '优秀' else 25 if performance_grade == '良好' else 20 if performance_grade == '一般' else 10) * 0.3:.1f} |
| 可靠性 | 40% | {40 if reliability_grade == '优秀' else 35 if reliability_grade == '良好' else 25 if reliability_grade == '一般' else 15} | {(40 if reliability_grade == '优秀' else 35 if reliability_grade == '良好' else 25 if reliability_grade == '一般' else 15) * 0.4:.1f} |
| 成本 | 30% | {30 if efficiency_grade == '优秀' else 25 if efficiency_grade == '良好' else 20 if efficiency_grade == '一般' else 10} | {(30 if efficiency_grade == '优秀' else 25 if efficiency_grade == '良好' else 20 if efficiency_grade == '一般' else 10) * 0.3:.1f} |
| **总分** | **100%** | **{overall_score}** | **{overall_score}** |

**综合评级**: **{overall_grade}**

### 7.5 改进建议

"""
        suggestions = []
        
        if performance_grade in ["一般", "需优化"]:
            suggestions.append("1. **性能优化建议**:\n   - 考虑增加并发数以提高处理速度\n   - 优化LLM调用策略，减少不必要的请求\n   - 对大消息量的群聊进行消息采样或分批处理")
        
        if reliability_grade in ["一般", "需改进"]:
            suggestions.append("2. **可靠性改进建议**:\n   - 检查失败案例的具体错误原因\n   - 增加重试机制以提高成功率\n   - 完善异常处理和错误恢复机制")
        
        if efficiency_grade in ["一般", "需优化"]:
            suggestions.append("3. **成本优化建议**:\n   - 优化Prompt长度，减少提示Token消耗\n   - 对消息进行预处理，过滤无效内容\n   - 考虑使用更经济的模型处理简单任务")
        
        if not suggestions:
            suggestions.append("系统表现优秀，建议保持当前配置并持续监控运行状态。")
        
        evaluation += "\n".join(suggestions)
        
        evaluation += f"""

### 7.6 测试结论

本次测试对系统进行了全面的性能评估，测试覆盖了{self.metrics.total_rooms}个群聊、{self.metrics.total_messages}条消息，执行了{len(self.metrics.analysis_types)}种分析类型。

**主要发现**:
- 系统整体性能{performance_grade}，平均每群分析耗时{avg_time_per_room:.2f}秒
- 系统可靠性{reliability_grade}，成功率达到{success_rate:.1f}%
- Token使用效率{efficiency_grade}，平均每群消耗{avg_tokens_per_room:.1f}个Token

**总体评价**: 系统综合评级为**{overall_grade}**，{'表现优秀，可以投入生产使用。' if overall_grade == '优秀' else '表现良好，建议根据改进建议进行优化后投入使用。' if overall_grade == '良好' else '表现一般，建议根据改进建议进行优化后再投入使用。' if overall_grade == '一般' else '需要改进，建议全面优化后再进行测试。'}

---

*报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*
"""
        
        return evaluation

    def generate_visualization_table(self, result: Dict[str, Any], output_file: str = "可视化测试结果表格.md"):
        if "data" not in result:
            print("✗ 没有分析结果数据")
            return
        
        data = result["data"]
        results = data.get("results", [])
        
        table_header = """# 企业微信群聊智能分析测试结果可视化表格

## 测试概览

| 项目 | 内容 |
| :--- | :--- |
| 测试时间 | {test_time} |
| 群聊总数 | {total_rooms} |
| 成功分析 | {success_count} |
| 失败分析 | {failed_count} |
| 成功率 | {success_rate:.1f}% |

## 详细分析结果表格

| 序号 | 权限群聊名称 | 客户 | 售后人员 | 当日消息总量 | 售后回复量 | 核心信息摘要（通知/决策/待办） | 漏报消息分析（风险预警） | 漏报消息证据 | 高频词统计 | 客户情感分析 | 客户差评证据 | 客户负面内容 | 售后情感分析 | 售后负面内容 | 敏感词触发 | 风险等级 | 备注 |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
""".format(
            test_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_rooms=data.get('total_rooms', 0),
            success_count=data.get('success_count', 0),
            failed_count=data.get('failed_count', 0),
            success_rate=(data.get('success_count', 0) / data.get('total_rooms', 1) * 100) if data.get('total_rooms', 0) > 0 else 0
        )
        
        table_rows = []
        
        for i, room_result in enumerate(results, 1):
            room_id = room_result.get("room_id", "")
            room_name = room_result.get("room_name", "未知")
            status = room_result.get("status", "unknown")
            room_data = room_result.get("data", {})
            
            if status == "success" and room_data:
                msg_count = room_data.get("message_count", 0)
                msg_total = f"{msg_count}条"
                
                employee_reply_count = self._extract_employee_reply_count(room_data)
                employee_reply = f"{employee_reply_count}条"
                
                core_summary = self._extract_core_summary(room_data.get("summary", ""))
                
                unanswered_data = room_data.get("unanswered_status", {})
                unanswered_analysis = self._extract_unanswered_analysis(unanswered_data)
                missed_evidence = self._extract_missed_message_evidence(unanswered_data)
                
                highfreq_words = self._extract_highfreq_words(room_data.get("high_freq_words", {}))
                
                sentiment_data = room_data.get("sentiment", {})
                customer_sentiment = self._extract_customer_sentiment(sentiment_data)
                customer_bad_evidence = self._extract_customer_bad_review_evidence(sentiment_data)
                customer_negative = self._extract_customer_negative(sentiment_data)
                
                employee_sentiment = self._extract_employee_sentiment(sentiment_data)
                employee_negative = self._extract_employee_negative(sentiment_data)
                
                sensitive_words = self._extract_sensitive_words(room_data.get("sensitive_words", {}))
                
                risk_level = self._determine_risk_level(room_data, unanswered_data)
                
                remarks = self._extract_remarks(room_data, unanswered_data)
                
                customer_names, employee_names = self._extract_participants(room_data, room_id)
                
            else:
                msg_total = "-"
                employee_reply = "-"
                core_summary = f"分析失败: {room_result.get('error_message', '未知错误')}"
                unanswered_analysis = "✅无漏回"
                missed_evidence = "-"
                highfreq_words = "-"
                customer_sentiment = "暂无分析"
                customer_bad_evidence = "-"
                customer_negative = "无"
                employee_sentiment = "暂无分析"
                employee_negative = "无"
                sensitive_words = "无"
                risk_level = "🔴高"
                remarks = "需人工核查"
                customer_names = "-"
                employee_names = "-"
            
            row = f"| {i} | {room_name} | {customer_names} | {employee_names} | {msg_total} | {employee_reply} | {core_summary} | {unanswered_analysis} | {missed_evidence} | {highfreq_words} | {customer_sentiment} | {customer_bad_evidence} | {customer_negative} | {employee_sentiment} | {employee_negative} | {sensitive_words} | {risk_level} | {remarks} |"
            table_rows.append(row)
        
        table_content = table_header + "\n".join(table_rows)
        
        table_content += """

## 字段说明

1. **权限群聊名称**: 企业微信群聊的完整名称（优先显示room_name，如为空则显示room_id）
2. **当日消息总量**: 群聊中的消息总数，反映群活跃程度
3. **售后回复量**: 售后人员回复的消息数量
4. **核心信息摘要**: 从群聊JSON摘要中提取的核心概述、客户诉求、处理进展（LLM生成JSON格式）
5. **漏报消息分析（风险预警）**: 检测是否存在消息漏回情况及建议行动
6. **漏报消息证据**: 存在漏回时，展示漏回消息的原始内容（发送者、时间、消息内容）
7. **高频词统计**: 群聊中出现频率最高的业务相关词汇
8. **客户情感分析**: 客户消息的情感倾向统计（好评/差评数量）
9. **客户差评证据**: 客户表达不满或抱怨的原始消息内容（发送者、时间、消息内容）
10. **客户负面内容**: 客户表达不满或抱怨的具体消息内容
11. **售后情感分析**: 售后人员消息的情感倾向统计（积极/恶劣数量）
12. **售后负面内容**: 售后人员表达恶劣态度的具体消息内容
13. **敏感词触发**: 群聊中触发的敏感词详情
14. **风险等级**: 综合评估的风险等级（🔴高/🟡中/🟢低/无）
15. **备注**: 从摘要JSON中提取的待办事项(todos)，以及漏回建议

---

*表格生成时间: {time}*
""".format(time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(table_content)
        
        print(f"✓ 可视化表格已保存到: {output_file}")

    def _extract_employee_reply_count(self, room_data: Dict[str, Any]) -> int:
        return room_data.get("employee_reply_count", 0)

    def _extract_core_summary(self, summary_text: str, for_html: bool = False) -> str:
        if not summary_text:
            return "群内互动较少，暂无核心议题。"
        
        try:
            import json
            cleaned = summary_text.strip()
            if cleaned.startswith("```json"):
                cleaned = cleaned[7:]
            elif cleaned.startswith("```"):
                cleaned = cleaned[3:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()
            
            summary_json = json.loads(cleaned)
            
            parts = []
            
            overview = summary_json.get("overview", "")
            if overview and overview != "无":
                overview = self._clean_markdown_text(overview)
                if for_html:
                    parts.append(f'<div class="summary-item"><div class="summary-label">【核心概述】</div><div class="summary-content">{overview}</div></div>')
                else:
                    parts.append(f"【核心概述】{overview}")
            
            demands = summary_json.get("demands", [])
            if isinstance(demands, list) and demands:
                demands_text = self._format_list_items(demands, for_html)
                if for_html:
                    parts.append(f'<div class="summary-item"><div class="summary-label">【客户诉求】</div><div class="summary-content">{demands_text}</div></div>')
                else:
                    parts.append(f"【客户诉求】{demands_text}")
            
            actions = summary_json.get("actions", [])
            if isinstance(actions, list) and actions:
                actions_text = self._format_list_items(actions, for_html)
                if for_html:
                    parts.append(f'<div class="summary-item"><div class="summary-label">【处理进展】</div><div class="summary-content">{actions_text}</div></div>')
                else:
                    parts.append(f"【处理进展】{actions_text}")
            
            todos = summary_json.get("todos", [])
            if isinstance(todos, list) and todos:
                todos_text = self._format_list_items(todos, for_html)
                if for_html:
                    parts.append(f'<div class="summary-item"><div class="summary-label">【待办跟进】</div><div class="summary-content">{todos_text}</div></div>')
                else:
                    parts.append(f"【待办跟进】{todos_text}")
            
            if for_html:
                return "".join(parts) if parts else "群内互动较少，暂无核心议题。"
            else:
                return "<br>".join(parts) if parts else "群内互动较少，暂无核心议题。"
            
        except (json.JSONDecodeError, TypeError):
            formatted = self._clean_markdown_text(summary_text)
            return formatted
    
    def _format_list_items(self, items: list, for_html: bool = False) -> str:
        if not items:
            return ""
        
        formatted_items = []
        for item in items:
            if item and item != "无":
                item = self._clean_markdown_text(item)
                if for_html:
                    formatted_items.append(f"• {item}")
                else:
                    formatted_items.append(f"• {item}")
        
        if for_html:
            return "<br>".join(formatted_items)
        else:
            return "<br>".join(formatted_items)
    
    def _clean_markdown_text(self, text: str) -> str:
        import re
        if not text:
            return text
        
        text = text.replace('\n', '<br>')
        text = re.sub(r'(?<!<br>)^- ', '• ', text, flags=re.MULTILINE)
        text = re.sub(r'^\* ', '• ', text, flags=re.MULTILINE)
        text = re.sub(r'-\s*-\s*-\s*-\s*-\s*-\s*-\s*-\s*-\s*-\s*-\s*-\s*-\s*-', '——', text)
        text = re.sub(r'\s*<br>\s*', '<br>', text)
        text = re.sub(r'(<br>){3,}', '<br><br>', text)
        text = re.sub(r'^\s+', '', text, flags=re.MULTILINE)
        text = re.sub(r'\s+$', '', text, flags=re.MULTILINE)
        
        return text

    def _extract_highfreq_words(self, highfreq_data: Dict[str, Any]) -> str:
        if not highfreq_data or not highfreq_data.get("words"):
            return "-"
        
        words = highfreq_data.get("words", [])[:5]
        if not words:
            return "-"
        
        word_list = []
        for w in words:
            word = w.get("word", "")
            count = w.get("count", 0)
            if word:
                word_list.append(f"{word}（{count}次）")
        
        return "<br>".join(word_list) if word_list else "-"
    
    def _extract_unanswered_analysis(self, unanswered_data: Dict[str, Any]) -> str:
        if not unanswered_data:
            return "✅无漏回"
        
        is_missed = unanswered_data.get("is_missed", False)
        
        if is_missed:
            suggested_action = unanswered_data.get("suggested_action", "请及时跟进")
            return f"❗存在漏回<br>建议：{suggested_action}"
        else:
            return "✅无漏回"

    def _extract_missed_message_evidence(self, unanswered_data: Dict[str, Any]) -> str:
        if not unanswered_data:
            return "-"
        
        is_missed = unanswered_data.get("is_missed", False)
        if not is_missed:
            return "-"
        
        missed_messages = unanswered_data.get("missed_messages", [])
        if not missed_messages:
            return "-"
        
        evidence_list = []
        for msg in missed_messages[:3]:
            sender = msg.get("sender_name", "未知")
            time = msg.get("msgtime", "")
            content = msg.get("content", "")
            if content:
                evidence_list.append(f"【{sender}】{time}<br>{content[:80]}{'...' if len(content) > 80 else ''}")
        
        return "<br><br>".join(evidence_list) if evidence_list else "-"

    def _extract_customer_bad_review_evidence(self, sentiment_data: Dict[str, Any]) -> str:
        if not sentiment_data:
            return "-"
        
        details = sentiment_data.get("details", {})
        if not isinstance(details, dict):
            return "-"
        
        customer_bad = details.get("customer_bad", [])
        if not isinstance(customer_bad, list) or not customer_bad:
            return "-"
        
        bad_reviews = []
        for msg in customer_bad[:3]:
            if isinstance(msg, dict):
                sender = msg.get("sender_name", "未知")
                time = msg.get("msgtime", "")
                content = msg.get("content", "")
                if content:
                    bad_reviews.append(f"【{sender}】{time}<br>{content[:80]}{'...' if len(content) > 80 else ''}")
        
        if not bad_reviews:
            return "-"
        
        return "<br><br>".join(bad_reviews)

    def _extract_participants(self, room_data: Dict[str, Any], room_id: str = None) -> tuple:
        customer_set = set()
        employee_set = set()
        
        members = {}
        if room_id and room_id in self.metrics.room_members:
            members = self.metrics.room_members[room_id]
        
        for userid, member in members.items():
            member_type = member.get("type", 1)
            name = member.get("name", "")
            group_nickname = member.get("group_nickname", "")
            display_name = name or group_nickname
            
            if not display_name:
                continue
            
            if member_type == 2:
                customer_set.add(display_name)
            else:
                employee_set.add(display_name)
        
        sentiment = room_data.get("sentiment", {})
        if sentiment:
            details = sentiment.get("details", {})
            if isinstance(details, dict):
                for key in ["customer_good", "customer_bad", "customer_neutral"]:
                    for msg in details.get(key, []):
                        if isinstance(msg, dict):
                            name = msg.get("sender_name", "")
                            if name:
                                customer_set.add(name)
                
                for key in ["employee_positive", "employee_bad"]:
                    for msg in details.get(key, []):
                        if isinstance(msg, dict):
                            name = msg.get("sender_name", "")
                            if name:
                                employee_set.add(name)
        
        unanswered = room_data.get("unanswered_status", {})
        if unanswered:
            for msg in unanswered.get("missed_messages", []):
                if isinstance(msg, dict):
                    name = msg.get("sender_name", "")
                    role = msg.get("sender_role", "")
                    if name:
                        if role == "客户":
                            customer_set.add(name)
                        elif role in ["售后", "员工", "销售"]:
                            employee_set.add(name)
        
        customer_names = "、".join(sorted(customer_set)) if customer_set else "-"
        employee_names = "、".join(sorted(employee_set)) if employee_set else "-"
        
        return customer_names, employee_names

    def _extract_customer_sentiment(self, sentiment_data: Dict[str, Any]) -> str:
        if not sentiment_data:
            return "暂无分析"
        
        summary = sentiment_data.get("summary", {})
        customer = summary.get("customer", {})
        
        good = customer.get("good_reviews", 0)
        bad = customer.get("bad_reviews", 0)
        
        if good == 0 and bad == 0:
            return "😐中性（无明确情感）"
        
        return f"😊好评：{good}<br>😞差评：{bad}"

    def _extract_customer_negative(self, sentiment_data: Dict[str, Any]) -> str:
        if not sentiment_data:
            return "无"
        
        details = sentiment_data.get("details", {})
        if not isinstance(details, dict):
            return "无"
        
        customer_bad = details.get("customer_bad", [])
        if not isinstance(customer_bad, list) or not customer_bad:
            return "无"
        
        negative_msgs = []
        
        for msg in customer_bad[:3]:
            if isinstance(msg, dict):
                content = msg.get("content", "")
                if content:
                    negative_msgs.append(content[:50])
        
        if not negative_msgs:
            return "无"
        
        return "<br>".join(negative_msgs)

    def _extract_employee_sentiment(self, sentiment_data: Dict[str, Any]) -> str:
        if not sentiment_data:
            return "暂无分析"
        
        summary = sentiment_data.get("summary", {})
        employee = summary.get("employee", {})
        
        positive = employee.get("positive", 0)
        negative = employee.get("bad_attitude", 0)
        
        if positive == 0 and negative == 0:
            return "😐中性（无明确情感）"
        
        return f"🌟积极：{positive}<br>😞消极：{negative}"

    def _extract_employee_negative(self, sentiment_data: Dict[str, Any]) -> str:
        if not sentiment_data:
            return "无"
        
        details = sentiment_data.get("details", {})
        if not isinstance(details, dict):
            return "无"
        
        employee_bad = details.get("employee_bad_attitude", [])
        if not isinstance(employee_bad, list) or not employee_bad:
            return "无"
        
        negative_msgs = []
        
        for msg in employee_bad[:3]:
            if isinstance(msg, dict):
                content = msg.get("content", "")
                if content:
                    negative_msgs.append(content[:50])
        
        if not negative_msgs:
            return "无"
        
        return "<br>".join(negative_msgs)

    def _extract_sensitive_words(self, sensitive_data: Dict[str, Any]) -> str:
        if not sensitive_data:
            return "无"
        
        total_hits = sensitive_data.get("total_hits", 0)
        if total_hits == 0:
            return "无"
        
        words = sensitive_data.get("words", [])
        if not words:
            return f"触发{total_hits}次"
        
        word_list = []
        for w in words[:5]:
            word = w.get("word", "")
            count = w.get("count", 0)
            if word:
                word_list.append(f"{word}（{count}次）")
        
        return "、".join(word_list) if word_list else f"触发{total_hits}次"

    def _determine_risk_level(self, room_data: Dict[str, Any], unanswered: Dict[str, Any]) -> str:
        risk_score = 0
        
        sentiment = room_data.get("sentiment", {})
        if sentiment:
            summary = sentiment.get("summary", {})
            customer = summary.get("customer", {})
            if customer.get("bad_reviews", 0) > 0:
                risk_score += 2
            
            employee = summary.get("employee", {})
            if employee.get("bad_attitude", 0) > 0:
                risk_score += 3
        
        sensitive = room_data.get("sensitive_words", {})
        if sensitive and sensitive.get("total_hits", 0) > 0:
            risk_score += 2
        
        if unanswered:
            if unanswered.get("is_missed", False):
                risk_score += 2
            if unanswered.get("risk_level") == "high":
                risk_score += 2
        
        if risk_score >= 6:
            return "🔴高"
        elif risk_score >= 3:
            return "🟡中"
        elif risk_score >= 1:
            return "🟢低"
        else:
            return "无"

    def _extract_remarks(self, room_data: Dict[str, Any], unanswered: Dict[str, Any], for_html: bool = False) -> str:
        notes = []
        
        if unanswered and unanswered.get("is_missed") and unanswered.get("suggested_action"):
            suggested = self._clean_markdown_text(unanswered.get("suggested_action", ""))
            if for_html:
                notes.append(f'<div class="remark-item remark-warning"><span class="remark-icon">🔔</span> 漏报建议：{suggested}</div>')
            else:
                notes.append(f"🔔 漏报建议：{suggested}")
        
        if for_html:
            return "".join(notes) if notes else '-'
        else:
            return '<br>'.join(notes) if notes else '-'

    def generate_html_table(self, result: Dict[str, Any], output_file: str = "可视化测试结果表格.html"):
        if "data" not in result:
            print("✗ 没有分析结果数据")
            return
        
        data = result["data"]
        results = data.get("results", [])
        
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>企业微信群聊智能分析测试结果</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        
        .container {{
            max-width: 1800px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}
        
        .header p {{
            opacity: 0.9;
            font-size: 14px;
        }}
        
        .overview {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
        }}
        
        .overview-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            text-align: center;
        }}
        
        .overview-card .label {{
            color: #6c757d;
            font-size: 14px;
            margin-bottom: 8px;
        }}
        
        .overview-card .value {{
            font-size: 24px;
            font-weight: bold;
            color: #667eea;
        }}
        
        .table-container {{
            padding: 30px;
            overflow-x: auto;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            background: white;
        }}
        
        thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}
        
        th {{
            padding: 15px 10px;
            text-align: left;
            font-weight: 600;
            white-space: nowrap;
            border: none;
        }}
        
        td {{
            padding: 12px 10px;
            border-bottom: 1px solid #e9ecef;
            vertical-align: top;
        }}
        
        tr:hover {{
            background: #f8f9fa;
        }}
        
        .risk-high {{
            color: #dc3545;
            font-weight: bold;
        }}
        
        .risk-medium {{
            color: #ffc107;
            font-weight: bold;
        }}
        
        .risk-low {{
            color: #28a745;
            font-weight: bold;
        }}
        
        .summary-item {{
            margin-bottom: 10px;
            padding: 8px 12px;
            background: #f8f9fa;
            border-radius: 6px;
            border-left: 3px solid #667eea;
        }}
        
        .summary-label {{
            font-weight: 600;
            color: #495057;
            margin-bottom: 4px;
            font-size: 13px;
        }}
        
        .summary-content {{
            line-height: 1.7;
            color: #333;
            font-size: 13px;
            word-break: break-word;
        }}
        
        .remark-item {{
            margin-bottom: 6px;
            padding: 4px 8px;
            background: #e7f5ff;
            border-radius: 4px;
        }}
        
        .remark-warning {{
            background: #fff3cd;
        }}
        
        .remark-icon {{
            font-size: 14px;
        }}
        
        .tag {{
            display: inline-block;
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 11px;
            margin: 2px;
        }}
        
        .tag-success {{
            background: #d4edda;
            color: #155724;
        }}
        
        .tag-warning {{
            background: #fff3cd;
            color: #856404;
        }}
        
        .tag-danger {{
            background: #f8d7da;
            color: #721c24;
        }}
        
        .emoji {{
            font-size: 16px;
        }}
        
        .footer {{
            text-align: center;
            padding: 20px;
            background: #f8f9fa;
            color: #6c757d;
            font-size: 12px;
        }}
        
        @media (max-width: 768px) {{
            .overview {{
                grid-template-columns: repeat(2, 1fr);
            }}
            
            table {{
                font-size: 11px;
            }}
            
            th, td {{
                padding: 8px 5px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>企业微信群聊智能分析测试结果</h1>
            <p>测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <div class="overview">
            <div class="overview-card">
                <div class="label">群聊总数</div>
                <div class="value">{data.get('total_rooms', 0)}</div>
            </div>
            <div class="overview-card">
                <div class="label">成功分析</div>
                <div class="value">{data.get('success_count', 0)}</div>
            </div>
            <div class="overview-card">
                <div class="label">失败分析</div>
                <div class="value">{data.get('failed_count', 0)}</div>
            </div>
            <div class="overview-card">
                <div class="label">成功率</div>
                <div class="value">{(data.get('success_count', 0) / data.get('total_rooms', 1) * 100) if data.get('total_rooms', 0) > 0 else 0:.1f}%</div>
            </div>
        </div>
        
        <div class="table-container">
            <table>
                <thead>
                    <tr>
                        <th>序号</th>
                        <th>权限群聊名称</th>
                        <th>客户</th>
                        <th>售后人员</th>
                        <th>当日消息总量</th>
                        <th>售后回复量</th>
                        <th>核心信息摘要</th>
                        <th>漏报消息分析</th>
                        <th>漏报消息证据</th>
                        <th>高频词统计</th>
                        <th>客户情感分析</th>
                        <th>客户差评证据</th>
                        <th>客户负面内容</th>
                        <th>售后情感分析</th>
                        <th>售后负面内容</th>
                        <th>敏感词触发</th>
                        <th>风险等级</th>
                        <th>备注</th>
                    </tr>
                </thead>
                <tbody>
"""
        
        for i, room_result in enumerate(results, 1):
            room_id = room_result.get("room_id", "")
            room_name = room_result.get("room_name", "未知")
            status = room_result.get("status", "unknown")
            room_data = room_result.get("data", {})
            
            if status == "success" and room_data:
                msg_count = room_data.get("message_count", 0)
                msg_total = f"{msg_count}条"
                
                employee_reply_count = self._extract_employee_reply_count(room_data)
                employee_reply = f"{employee_reply_count}条"
                
                core_summary = self._extract_core_summary(room_data.get("summary", ""))
                
                unanswered_data = room_data.get("unanswered_status", {})
                unanswered_analysis = self._extract_unanswered_analysis(unanswered_data)
                missed_evidence = self._extract_missed_message_evidence(unanswered_data)
                
                highfreq_words = self._extract_highfreq_words(room_data.get("high_freq_words", {}))
                
                sentiment_data = room_data.get("sentiment", {})
                customer_sentiment = self._extract_customer_sentiment(sentiment_data)
                customer_bad_evidence = self._extract_customer_bad_review_evidence(sentiment_data)
                customer_negative = self._extract_customer_negative(sentiment_data)
                
                employee_sentiment = self._extract_employee_sentiment(sentiment_data)
                employee_negative = self._extract_employee_negative(sentiment_data)
                
                sensitive_words = self._extract_sensitive_words(room_data.get("sensitive_words", {}))
                
                risk_level = self._determine_risk_level(room_data, unanswered_data)
                
                remarks = self._extract_remarks(room_data, unanswered_data)
                
                customer_names, employee_names = self._extract_participants(room_data, room_id)
                
            else:
                msg_total = "-"
                employee_reply = "-"
                core_summary = f"分析失败: {room_result.get('error_message', '未知错误')}"
                unanswered_analysis = "✅无漏回"
                missed_evidence = "-"
                highfreq_words = "-"
                customer_sentiment = "无数据"
                customer_bad_evidence = "-"
                customer_negative = "无"
                employee_sentiment = "无数据"
                employee_negative = "无"
                sensitive_words = "无"
                risk_level = "🔴高"
                remarks = "需人工核查"
                customer_names = "-"
                employee_names = "-"
            
            risk_class = ""
            if "🔴" in risk_level:
                risk_class = "risk-high"
            elif "🟡" in risk_level:
                risk_class = "risk-medium"
            elif "🟢" in risk_level:
                risk_class = "risk-low"
            
            html_content += f"""
                    <tr>
                        <td>{i}</td>
                        <td><strong>{room_name}</strong></td>
                        <td>{customer_names}</td>
                        <td>{employee_names}</td>
                        <td>{msg_total}</td>
                        <td>{employee_reply}</td>
                        <td class="summary-content">{core_summary}</td>
                        <td>{unanswered_analysis}</td>
                        <td class="summary-content">{missed_evidence}</td>
                        <td>{highfreq_words}</td>
                        <td>{customer_sentiment}</td>
                        <td class="summary-content">{customer_bad_evidence}</td>
                        <td>{customer_negative}</td>
                        <td>{employee_sentiment}</td>
                        <td>{employee_negative}</td>
                        <td>{sensitive_words}</td>
                        <td class="{risk_class}">{risk_level}</td>
                        <td>{remarks}</td>
                    </tr>
"""
        
        html_content += """
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>企业微信群聊智能分析系统 | 自动生成报告</p>
            <p>如有疑问请联系技术支持</p>
        </div>
    </div>
</body>
</html>
"""
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

        print(f"✓ HTML可视化表格已保存到: {output_file}")

    def generate_excel_table(self, result: Dict[str, Any], output_file: str = "可视化测试结果表格.xlsx"):
        if not OPENPYXL_AVAILABLE:
            print("⚠️ openpyxl未安装，无法生成Excel表格")
            print("  请运行: pip install openpyxl")
            return

        if "data" not in result:
            print("✗ 没有分析结果数据")
            return

        data = result["data"]
        results = data.get("results", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "分析结果"

        headers = [
            "序号", "权限群聊名称", "客户", "售后人员", "当日消息总量",
            "售后回复量", "核心信息摘要", "漏报消息分析", "漏报消息证据",
            "高频词统计", "客户情感分析", "客户差评证据", "客户负面内容",
            "售后情感分析", "售后负面内容", "敏感词触发", "风险等级", "备注"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for i, room_result in enumerate(results, 1):
            room_id = room_result.get("room_id", "")
            room_name = room_result.get("room_name", "未知")
            status = room_result.get("status", "unknown")
            room_data = room_result.get("data", {})

            if status == "success" and room_data:
                msg_count = room_data.get("message_count", 0)
                msg_total = f"{msg_count}条"

                employee_reply_count = self._extract_employee_reply_count(room_data)
                employee_reply = f"{employee_reply_count}条"

                core_summary = self._extract_core_summary(room_data.get("summary", ""))
                core_summary = core_summary.replace('<br>', '\n')

                unanswered_data = room_data.get("unanswered_status", {})
                unanswered_analysis = self._extract_unanswered_analysis(unanswered_data)
                unanswered_analysis = unanswered_analysis.replace('<br>', '\n')

                missed_evidence = self._extract_missed_message_evidence(unanswered_data)
                missed_evidence = missed_evidence.replace('<br>', '\n')

                highfreq_words = self._extract_highfreq_words(room_data.get("high_freq_words", {}))
                highfreq_words = highfreq_words.replace('<br>', '\n')

                sentiment_data = room_data.get("sentiment", {})
                customer_sentiment = self._extract_customer_sentiment(sentiment_data)
                customer_sentiment = customer_sentiment.replace('<br>', '\n')

                customer_bad_evidence = self._extract_customer_bad_review_evidence(sentiment_data)
                customer_bad_evidence = customer_bad_evidence.replace('<br>', '\n')

                customer_negative = self._extract_customer_negative(sentiment_data)
                customer_negative = customer_negative.replace('<br>', '\n')

                employee_sentiment = self._extract_employee_sentiment(sentiment_data)
                employee_sentiment = employee_sentiment.replace('<br>', '\n')

                employee_negative = self._extract_employee_negative(sentiment_data)
                employee_negative = employee_negative.replace('<br>', '\n')

                sensitive_words = self._extract_sensitive_words(room_data.get("sensitive_words", {}))
                sensitive_words = sensitive_words.replace('<br>', '\n')

                risk_level = self._determine_risk_level(room_data, unanswered_data)

                remarks = self._extract_remarks(room_data, unanswered_data)
                remarks = remarks.replace('<br>', '\n')

                customer_names, employee_names = self._extract_participants(room_data, room_id)

            else:
                msg_total = "-"
                employee_reply = "-"
                core_summary = f"分析失败: {room_result.get('error_message', '未知错误')}"
                unanswered_analysis = "✅无漏回"
                missed_evidence = "-"
                highfreq_words = "-"
                customer_sentiment = "暂无分析"
                customer_bad_evidence = "-"
                customer_negative = "无"
                employee_sentiment = "暂无分析"
                employee_negative = "无"
                sensitive_words = "无"
                risk_level = "🔴高"
                remarks = "需人工核查"
                customer_names = "-"
                employee_names = "-"

            row_data = [
                i, room_name, customer_names, employee_names, msg_total,
                employee_reply, core_summary, unanswered_analysis, missed_evidence,
                highfreq_words, customer_sentiment, customer_bad_evidence, customer_negative,
                employee_sentiment, employee_negative, sensitive_words, risk_level, remarks
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i + 1, column=col, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 35

        try:
            wb.save(output_file)
            print(f"✓ Excel可视化表格已保存到: {output_file}")
        except Exception as e:
            print(f"✗ 保存Excel文件失败: {e}")

    def run_test(self, analysis_types: Optional[List[str]] = None, save_report: bool = False, generate_doc: bool = False, generate_table: bool = False, generate_html: bool = False, generate_excel: bool = False):
        print("\n" + "="*80)
        print("企业微信群聊分析服务 - 系统性能完整测试")
        print("="*80)
        print(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Python服务: {self.analysis_service_url}")
        print(f"Java数据源: {self.data_source_url}")

        rooms = self.fetch_data_from_java()

        if not rooms:
            print("\n⚠️ Java接口返回数据为空")
            print("  可能原因:")
            print("  1. Java服务未启动或接口地址不正确")
            print("  2. 当前没有群聊消息数据")
            print("  3. 接口认证失败")
            return

        if analysis_types is None:
            analysis_types = ["sentiment", "sensitive", "summary", "highfreq", "unanswered"]

        result = self.batch_analyze(rooms, analysis_types)
        self.print_performance_report(result)

        if save_report:
            self.save_report_to_file(result)

        if generate_doc:
            self.generate_markdown_report(result)

        if generate_table:
            self.generate_visualization_table(result)

        if generate_html:
            self.generate_html_table(result)

        if generate_excel:
            self.generate_excel_table(result)

    def close(self):
        self.http_client.close()


def main():
    parser = argparse.ArgumentParser(description="企业微信群聊分析服务系统性能测试")
    parser.add_argument(
        "--data-source",
        type=str,
        default=settings.JAVA_DATA_SOURCE_URL,
        help=f"Java数据源地址 (默认: {settings.JAVA_DATA_SOURCE_URL})",
    )
    parser.add_argument(
        "--service-url",
        type=str,
        default="http://localhost:8000",
        help="Python分析服务地址 (默认: http://localhost:8000)",
    )
    parser.add_argument(
        "--max-concurrent",
        type=int,
        default=5,
        help="最大并发数 (默认: 5)",
    )
    parser.add_argument(
        "--max-rooms",
        type=int,
        default=None,
        help="最大测试群聊数量 (默认: 全部测试)",
    )
    parser.add_argument(
        "--types",
        type=str,
        nargs="+",
        default=["sentiment", "sensitive", "summary", "highfreq", "unanswered"],
        choices=["sentiment", "sensitive", "summary", "highfreq", "unanswered"],
        help="分析类型 (默认全部)",
    )
    parser.add_argument(
        "--save-report",
        action="store_true",
        help="保存性能报告到JSON文件",
    )
    parser.add_argument(
        "--generate-doc",
        action="store_true",
        help="生成Markdown格式的测试文档",
    )
    parser.add_argument(
        "--generate-table",
        action="store_true",
        help="生成Markdown可视化表格",
    )
    parser.add_argument(
        "--generate-html",
        action="store_true",
        help="生成HTML可视化表格（推荐）",
    )
    parser.add_argument(
        "--generate-excel",
        action="store_true",
        help="生成Excel可视化表格",
    )

    args = parser.parse_args()

    tester = SystemPerformanceTester(
        data_source_url=args.data_source,
        analysis_service_url=args.service_url,
        max_concurrent=args.max_concurrent,
        max_rooms=args.max_rooms,
    )

    try:
        tester.run_test(
            analysis_types=args.types,
            save_report=args.save_report,
            generate_doc=args.generate_doc,
            generate_table=args.generate_table,
            generate_html=args.generate_html,
            generate_excel=args.generate_excel,
        )
    finally:
        tester.close()


if __name__ == "__main__":
    main()
