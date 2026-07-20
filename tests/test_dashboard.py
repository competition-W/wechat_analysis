import unittest
import json
import tempfile
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from contextlib import contextmanager
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from services import db_dashboard


class DashboardParsingTests(unittest.TestCase):
    def test_lims_outage_keeps_unfiltered_project_groups_visible(self):
        dimension = {
            "codes": ["LC-P2026001"], "projects": [], "regions": [],
            "aftersalers": [], "dimension_source": "lims_unavailable",
        }
        self.assertTrue(db_dashboard._dimension_matches_or_degrades(dimension))
        self.assertFalse(db_dashboard._dimension_matches_or_degrades(dimension, region="华东"))
        self.assertEqual(db_dashboard._scope_dimension(dimension)["codes"], ["LC-P2026001"])

    def test_extracts_multiple_project_codes_and_removes_duplicates(self):
        value = "客户群 LC-P2026001 / LC-X99 / LC-SP2026002 / LC-P2026001"
        self.assertEqual(
            db_dashboard.extract_project_codes(value),
            ["LC-P2026001", "LC-SP2026002"],
        )

    def test_extracts_supported_project_code_followed_by_chinese_text(self):
        self.assertEqual(
            db_dashboard.extract_project_codes("LC-SP202604130083项目综合群"),
            ["LC-SP202604130083"],
        )

    def test_ignores_non_focus_group_codes(self):
        self.assertEqual(db_dashboard.extract_project_codes("LC-C202604130083项目综合群"), [])
        self.assertFalse(db_dashboard.is_focus_group_name("普通交流群"))
        self.assertTrue(db_dashboard.is_focus_group_name("客户 LC-P202604130083 项目群"))

    def test_parses_count_fields(self):
        self.assertEqual(
            db_dashboard.parse_count_map("好评: 3, 差评：2"),
            {"好评": 3, "差评": 2},
        )

    def test_normalizes_key_account_flags(self):
        self.assertEqual(db_dashboard.normalize_key_account("0", "客户甲"), "")
        self.assertEqual(db_dashboard.normalize_key_account("1", "客户甲"), "客户甲")
        self.assertEqual(db_dashboard.normalize_key_account("KA-001", "客户甲"), "KA-001")
        self.assertEqual(db_dashboard.normalize_key_account("0", "客户甲", "KA-002"), "KA-002")

    def test_parse_active_day_preserves_float_precision(self):
        self.assertIsNone(db_dashboard.parse_active_day(None))
        self.assertIsNone(db_dashboard.parse_active_day(""))
        self.assertIsNone(db_dashboard.parse_active_day("未知"))
        self.assertEqual(db_dashboard.parse_active_day(0.08), 0.08)
        self.assertEqual(db_dashboard.parse_active_day(12), 12.0)
        self.assertEqual(db_dashboard.parse_active_day(12.5), 12.5)
        self.assertEqual(db_dashboard.parse_active_day("0.08"), 0.08)
        self.assertEqual(db_dashboard.parse_active_day("12.5天"), 12.5)
        self.assertEqual(db_dashboard.parse_active_day("-3.2"), -3.2)
        self.assertFalse(db_dashboard.parse_active_day(True) is not None)

    def test_extract_lims_active_day_preserves_float(self):
        self.assertEqual(db_dashboard.extract_lims_active_day({"activeDay": 0.08}), 0.08)
        self.assertEqual(db_dashboard.extract_lims_active_day({"activeDay": 12.5}), 12.5)
        self.assertEqual(db_dashboard.extract_lims_active_day({"activellay": 7.4}), 7.4)
        self.assertIsNone(db_dashboard.extract_lims_active_day({}))

    def test_active_durations_from_lims_preserves_float_max(self):
        dimensions = {
            "群A": {"projects": [{"active_day": 0.08}, {"active_day": 3.5}]},
            "群B": {"projects": [{"active_day": "12.5"}]},
            "群C": {"projects": [{"active_day": None}]},
        }
        durations = db_dashboard._active_durations_from_lims(dimensions)
        self.assertEqual(durations["群A"], 3.5)
        self.assertEqual(durations["群B"], 12.5)
        self.assertNotIn("群C", durations)

    def test_period_boundaries(self):
        today = date(2026, 7, 6)
        self.assertEqual(
            db_dashboard.resolve_period("week", today=today),
            (date(2026, 7, 6), date(2026, 7, 6), "week"),
        )
        self.assertEqual(
            db_dashboard.resolve_period("month", today=today),
            (date(2026, 7, 1), date(2026, 7, 6), "month"),
        )
        self.assertEqual(
            db_dashboard.resolve_period("quarter", today=today),
            (date(2026, 7, 1), date(2026, 7, 6), "quarter"),
        )

    def test_custom_period_rejects_reversed_dates(self):
        with self.assertRaises(ValueError):
            db_dashboard.resolve_period("custom", "2026-07-06", "2026-07-01")

    def test_dimension_filters_are_combined(self):
        dimension = {
            "regions": ["华东"],
            "aftersalers": ["张三"],
            "projects": [{"region": "华东", "raw_aftersaler": "张三", "category_l1": "环境", "category_l2": "常规转录组", "key_account": "客户甲"}],
        }
        self.assertTrue(
            db_dashboard._dimension_matches(
                dimension, "华东", "张三", "常规转录组", "客户甲"
            )
        )
        self.assertFalse(db_dashboard._dimension_matches(dimension, region="华南"))
        self.assertFalse(db_dashboard._dimension_matches(dimension, category="食品"))

    def test_all_products_excludes_projects_outside_the_three_allowed_categories(self):
        dimension = {
            "regions": ["华东"],
            "aftersalers": ["张三"],
            "projects": [
                {"project_code": "LC-P1", "region": "华东", "category_l2": "常规转录组"},
                {"project_code": "LC-P2", "region": "华东", "category_l2": "蛋白质组"},
            ],
        }

        scoped = db_dashboard._scope_dimension(dimension)

        self.assertTrue(db_dashboard._dimension_matches(dimension))
        self.assertEqual([item["project_code"] for item in scoped["projects"]], ["LC-P1"])
        self.assertEqual(scoped["codes"], ["LC-P1"])
        self.assertFalse(db_dashboard._dimension_matches({
            **dimension,
            "projects": [{"category_l2": "蛋白质组"}],
        }))

    def test_project_filters_must_match_the_same_lims_record(self):
        dimension = {
            "aftersalers": ["张三"],
            "projects": [
                {"region": "华东", "raw_aftersaler": "张三", "category_l2": "常规转录组", "key_account": "客户甲"},
                {"region": "华南", "raw_aftersaler": "李四", "category_l2": "微生物", "key_account": "客户乙"},
            ],
        }

        self.assertFalse(db_dashboard._dimension_matches(
            dimension, region="华东", category="微生物", key_account="客户乙"
        ))
        self.assertTrue(db_dashboard._dimension_matches(
            dimension, region="华南", category="微生物", key_account="客户乙"
        ))
        self.assertFalse(db_dashboard._dimension_matches(
            dimension, region="华南", aftersaler="张三", category="微生物", key_account="客户乙"
        ))

    def test_top_message_groups_ranks_all_groups_without_requiring_lims_customer_data(self):
        def group(name, messages, days, words=None, dimension=None):
            return {
                "group_name": name, "messages": messages, "dates": set(days),
                "high_freq": Counter(words or {}), "rows": [],
                "dimension": dimension or {},
            }

        groups = {
            "无客户字段群": group("无客户字段群", 50, ["2026-07-01", "2026-07-02"], {"进度": 3}),
            "客户群一": group("客户群一", 40, ["2026-07-01"], {"交付": 2}, {
                "codes": ["LC-P1"], "aftersalers": ["张三"],
                "projects": [{
                    "project_code": "LC-P1", "work_unit": "单位甲",
                    "customer_name": "客户甲", "category_l2": "常规转录组",
                }],
            }),
            "客户群二": group("客户群二", 30, ["2026-07-01"]),
            "客户群三": group("客户群三", 20, ["2026-07-01"]),
            "客户群四": group("客户群四", 10, ["2026-07-01"]),
            "未进入前五群": group("未进入前五群", 5, ["2026-07-01"]),
        }

        result = db_dashboard._build_top_message_groups(groups, 155)

        self.assertEqual(result["actual_count"], 5)
        self.assertEqual(result["total_groups"], 6)
        self.assertEqual(result["top5_messages"], 150)
        self.assertEqual(result["coverage_percentage"], 96.8)
        self.assertEqual(result["items"][0]["group_name"], "无客户字段群")
        self.assertEqual(result["items"][0]["active_days"], 2)
        self.assertEqual(result["items"][0]["high_frequency_top5"][0], {"word": "进度", "count": 3})
        self.assertEqual(result["items"][1]["customer_units"], ["单位甲"])
        self.assertNotIn("summary", result["items"][1])

    def test_top_message_groups_tie_breaks_by_active_days_then_name(self):
        groups = {
            "B群": {"messages": 10, "dates": {"2026-07-01"}, "high_freq": Counter(), "dimension": {}},
            "C群": {"messages": 10, "dates": {"2026-07-01", "2026-07-02"}, "high_freq": Counter(), "dimension": {}},
            "A群": {"messages": 10, "dates": {"2026-07-01"}, "high_freq": Counter(), "dimension": {}},
        }
        result = db_dashboard._build_top_message_groups(groups, 30)
        self.assertEqual([item["group_name"] for item in result["items"]], ["C群", "A群", "B群"])

    def test_top_message_groups_returns_stable_empty_shape(self):
        self.assertEqual(db_dashboard._build_top_message_groups({}, 0), {
            "limit": 5, "actual_count": 0, "total_groups": 0,
            "total_messages": 0, "top5_messages": 0,
            "coverage_percentage": 0.0, "items": [],
        })

    def test_project_year_parsing_prefers_code_and_normalizes_lims_start_time(self):
        self.assertEqual(db_dashboard.extract_project_code_year("LC-P2026001"), 2026)
        self.assertEqual(db_dashboard.extract_project_code_year("lc-sp20250102003"), 2025)
        self.assertIsNone(db_dashboard.extract_project_code_year("LC-P12345"))
        self.assertIsNone(db_dashboard.extract_project_code_year("VIP2026001"))
        self.assertEqual(
            db_dashboard.parse_lims_start_date("2024-03-02 10:30:00"),
            date(2024, 3, 2),
        )
        self.assertEqual(
            db_dashboard.parse_lims_start_date("开始于2023年7月9日"),
            date(2023, 7, 9),
        )
        self.assertIsNone(db_dashboard.parse_lims_start_date("未知"))

    def test_project_year_distribution_deduplicates_and_falls_back_to_earliest_start(self):
        groups = {
            "群A": {"dimension": {
                "codes": ["LC-P2026001", "LC-P12345", "LC-P42"],
                "aftersalers": ["张三"],
                "projects": [
                    {
                        "project_code": "LC-P2026001", "start_time": "2024-01-01",
                        "product_name": "项目甲", "customer_name": "客户甲",
                        "work_unit": "单位甲", "category_l2": "常规转录组",
                        "category_l3": "mRNA", "region": "华东",
                        "final_aftersaler": "张三",
                    },
                    {"project_code": "LC-P12345", "start_time": "2025-06-01"},
                ],
            }},
            "群B": {"dimension": {
                "codes": ["LC-P12345", "LC-SP2025001"],
                "projects": [
                    {"project_code": "LC-P12345", "start_time": "2024-03-02"},
                    {"project_code": "LC-SP2025001", "start_time": "2023-01-01"},
                ],
            }},
        }

        result = db_dashboard._build_project_year_distribution(groups, current_year=2026)

        self.assertEqual(result["total_projects"], 4)
        self.assertEqual(result["recognized_projects"], 3)
        self.assertEqual(result["unknown_projects"], 1)
        self.assertEqual(result["coverage_percentage"], 75.0)
        self.assertEqual(result["current_year_projects"], 1)
        self.assertEqual(result["previous_year_projects"], 1)
        self.assertEqual(result["older_projects"], 1)
        self.assertEqual([item["year"] for item in result["items"]], [2026, 2025, 2024, None])
        projects = {
            project["project_code"]: project
            for item in result["items"]
            for project in item["projects"]
        }
        self.assertEqual(projects["LC-P2026001"]["year_source"], "project_code")
        self.assertEqual(projects["LC-P12345"]["year"], 2024)
        self.assertEqual(projects["LC-P12345"]["year_source"], "lims_start_time")
        self.assertEqual(projects["LC-P12345"]["start_time"], "2024-03-02")
        self.assertEqual(projects["LC-P12345"]["group_count"], 2)
        self.assertEqual(projects["LC-P42"]["year_source"], "unknown")

    def test_project_year_distribution_returns_stable_empty_shape(self):
        self.assertEqual(db_dashboard._build_project_year_distribution({}, 2026), {
            "current_year": 2026,
            "total_projects": 0,
            "recognized_projects": 0,
            "unknown_projects": 0,
            "coverage_percentage": 0.0,
            "current_year_projects": 0,
            "previous_year_projects": 0,
            "older_projects": 0,
            "source_priority": ["project_code", "lims_start_time"],
            "items": [],
        })

    def test_overview_applies_product_scope_to_topics_accounts_and_cross_analysis(self):
        rows = [
            {
                "groupName": "允许产品群", "CREATEDTIME": "2026-07-08 10:00:00",
                "messageToDayCount": 10, "isMissedMessage": "0",
                "customerEmotionAnalysis": "好评:1", "saleEmotionAnalysis": "积极:1",
                "highFrequencyWords": "允许主题:3",
            },
            {
                "groupName": "其它产品群", "CREATEDTIME": "2026-07-08 10:00:00",
                "messageToDayCount": 99, "isMissedMessage": "1",
                "customerEmotionAnalysis": "差评:9", "saleEmotionAnalysis": "负向:9",
                "highFrequencyWords": "其它主题:20",
            },
        ]
        dimensions = {
            "允许产品群": {
                "codes": ["LC-P1", "LC-P3", "LC-P4"], "regions": ["华东"], "aftersalers": ["张三"],
                "projects": [
                    {
                        "project_code": "LC-P1", "region": "华东", "sales_person": "销售甲",
                        "raw_aftersaler": "张三", "analysis_simple_remark": "问题项目",
                        "category_l2": "常规转录组", "category_l3": "mRNA",
                        "key_account": "重点客户甲", "customer_name": "客户甲",
                        "work_unit": "单位甲", "active_day": 12,
                    },
                    {
                        "project_code": "LC-P3", "region": "华东", "sales_person": "销售甲",
                        "raw_aftersaler": "张三", "analysis_simple_remark": "正常交付",
                        "category_l2": "表观组学", "category_l3": "甲基化",
                        "active_day": 8,
                    },
                    {
                        "project_code": "LC-P4", "region": "华东", "sales_person": "销售甲",
                        "raw_aftersaler": "张三", "analysis_simple_remark": "暂不交付",
                        "category_l2": "微生物", "category_l3": "扩增子",
                        "active_day": 5,
                    },
                ],
            },
            "其它产品群": {
                "codes": ["LC-P2"], "regions": ["华南"], "aftersalers": ["李四"],
                "projects": [{
                    "project_code": "LC-P2", "region": "华南", "sales_person": "销售乙",
                    "raw_aftersaler": "李四", "analysis_simple_remark": "问题项目",
                    "category_l2": "蛋白质组", "category_l3": "蛋白",
                    "key_account": "其它重点客户", "customer_name": "客户乙",
                    "work_unit": "单位乙", "active_day": 20,
                }],
            },
        }

        @contextmanager
        def fake_database(_operation):
            yield object()

        with (
            patch.object(db_dashboard, "database", fake_database),
            patch.object(db_dashboard, "_latest_rows", return_value=(rows, 2)),
            patch.object(db_dashboard, "_load_dimensions", return_value=(dimensions, {})),
            patch.object(db_dashboard, "_raw_analysis_count", return_value=1),
            patch.object(db_dashboard, "_query_group_rows", return_value=[]),
            patch.object(db_dashboard, "_query_chat_rows", return_value=[]),
        ):
            result = db_dashboard._build_overview(
                date(2026, 7, 1), date(2026, 7, 8), "custom"
            )
            microbe_result = db_dashboard._build_overview(
                date(2026, 7, 1), date(2026, 7, 8), "custom", category="微生物"
            )

        self.assertEqual(result["summary"]["total_groups"], 1)
        self.assertEqual(result["summary"]["total_messages"], 10)
        self.assertEqual(result["communication"]["high_frequency"], [
            {"word": "允许主题", "count": 3}
        ])
        self.assertEqual(
            [item["key_account"] for item in result["business"]["key_accounts"]],
            ["重点客户甲"],
        )
        self.assertEqual(
            result["communication"]["top_message_groups"]["items"][0]["group_name"],
            "允许产品群",
        )
        self.assertEqual(result["communication"]["top_message_groups"]["coverage_percentage"], 100.0)
        self.assertEqual(
            microbe_result["communication"]["top_message_groups"]["items"][0]["group_name"],
            "允许产品群",
        )
        self.assertEqual(
            [item["region"] for item in result["cross_analysis"]["region_sales"]],
            ["华东"],
        )
        self.assertEqual(result["service_quality"]["unanswered"]["missed_groups"], 0)
        self.assertEqual(result["project_attention"]["total_projects"], 2)
        self.assertEqual(
            [item["status"] for item in result["project_attention"]["summary"]],
            ["问题项目", "暂不交付"],
        )
        self.assertEqual(
            [item["project_code"] for item in result["project_attention"]["items"]],
            ["LC-P1", "LC-P4"],
        )
        self.assertEqual(
            [item["project_code"] for item in microbe_result["project_attention"]["items"]],
            ["LC-P4"],
        )

    def test_excel_export_splits_dashboard_modules_and_raw_sources_into_sheets(self):
        from openpyxl import load_workbook

        overview = {
            "summary": {
                "total_groups": 1, "total_messages": 10, "project_groups": 1,
                "regions": 1, "aftersaler_count": 1, "product_categories": 1,
                "key_accounts": 1, "short_active_ratio": 100,
            },
            "communication": {
                "trend": [{"date": "2026-07-08", "messages": 10, "groups": 1, "missed": 0}],
                "high_frequency": [{"word": "转录", "count": 3}],
                "active_duration": [{"range": "8-30天", "label": "短期服务", "count": 1, "percentage": 100}],
                "top_message_groups": {
                    "items": [{
                        "rank": 1, "group_name": "允许产品群", "message_count": 10,
                        "percentage_of_all": 100, "active_days": 1,
                        "high_frequency_top5": [{"word": "转录", "count": 3}],
                        "project_codes": ["LC-P1"], "customer_units": ["单位甲"],
                        "customer_names": ["客户甲"], "product_categories": ["常规转录组"],
                        "aftersalers": ["张三"],
                    }],
                },
            },
            "business": {
                "regions": [{"region": "华东", "group_count": 1}],
                "aftersalers": [{"name": "张三", "group_count": 1}],
                "product_categories": [{"category": "常规转录组", "project_count": 1}],
                "key_accounts": [{"key_account": "重点客户甲", "project_count": 1}],
                "project_year_distribution": {
                    "items": [{
                        "year": 2026, "label": "今年（2026）", "project_count": 1,
                        "percentage": 100, "group_count": 1,
                        "projects": [{
                            "project_code": "LC-P2026001", "year": 2026,
                            "year_source": "project_code", "start_time": "2026-01-02",
                            "project_names": ["RNA"], "customer_names": ["客户甲"],
                            "work_units": ["单位甲"], "product_categories": ["常规转录组"],
                            "regions": ["华东"], "aftersalers": ["张三"],
                            "group_count": 1, "group_names": ["允许产品群"],
                        }],
                    }],
                },
            },
            "project_attention": {
                "target_statuses": ["问题项目", "暂不交付"],
                "total_projects": 1,
                "summary": [{"status": "问题项目", "project_count": 1}],
                "items": [{"status": "问题项目", "project_code": "LC-P1"}],
            },
            "cross_analysis": {
                "region_sales": [], "region_after": [], "region_product": [],
            },
            "service_quality": {
                "unanswered": {"total_groups": 1, "missed_groups": 0},
                "sentiment": {"customer_good": 1, "customer_bad": 0},
            },
            "data_quality": {"project_codes": 1, "matched_project_codes": 1},
        }
        scope = {
            "project_codes": ["LC-P1"], "group_names": ["允许产品群"],
            "raw_rows": [{"groupName": "允许产品群"}],
            "group_rows": [{"name": "允许产品群"}],
            "chat_rows": [{"roomid": "room-1", "content": "测试"}],
        }

        @contextmanager
        def fake_database(_operation):
            yield object()

        with (
            patch.object(db_dashboard, "get_overview", return_value=overview),
            patch.object(db_dashboard, "database", fake_database),
            patch.object(db_dashboard, "_query_qx_raw_scope", return_value=scope),
            patch.object(db_dashboard, "fetch_lims_base_data", return_value=({
                "LC-P1": [{
                    "projectCode": "LC-P1", "productBigSortTwo": "常规转录组",
                    "productName": "RNA", "customerName": "客户甲",
                }]
            }, {"requests": 1, "records": 1, "errors": 0})),
        ):
            content = db_dashboard.get_export_excel(
                "custom", "2026-07-01", "2026-07-08"
            )

        workbook = load_workbook(BytesIO(content), read_only=True)
        self.assertTrue({
            "导出说明", "经营摘要", "高频关注主题", "重点客户", "消息Top5群聊",
            "项目年份分布", "项目年份明细", "项目状态关注", "服务质量",
            "analysis原始数据", "group原始数据", "chat原始数据", "LIMS原始数据",
        }.issubset(set(workbook.sheetnames)))
        self.assertEqual(workbook["高频关注主题"]["A2"].value, "转录")
        self.assertEqual(workbook["消息Top5群聊"]["B2"].value, "允许产品群")
        self.assertEqual(workbook["消息Top5群聊"]["G2"].value, "LC-P1")
        self.assertEqual(workbook["项目年份分布"]["A2"].value, 2026)
        self.assertEqual(workbook["项目年份明细"]["C2"].value, "LC-P2026001")
        self.assertEqual(workbook["项目状态关注"]["A2"].value, "问题项目")
        self.assertEqual(workbook["LIMS原始数据"]["B2"].value, "LC-P1")

    def test_lims_record_keeps_work_unit_and_l2_category(self):
        item = db_dashboard.normalize_lims_api_record(
            {
                "projectCode": "LC-P2026001",
                "workUnit": "客户单位A",
                "productBigSortTwo": "常规转录组",
                "productBigSortThree": "mRNA",
                "analysisSimpleRemark": "  问题项目  ",
            },
            "LC-P2026001",
        )

        self.assertEqual(item["work_unit"], "客户单位A")
        self.assertEqual(item["category_l2"], "常规转录组")
        self.assertEqual(item["analysis_simple_remark"], "问题项目")

    def test_time_period_breakdown_uses_real_msgtime_and_separates_weekend(self):
        groups = {
            "客户 LC-P2026001 项目群": {
                "dimension": {"aftersalers": ["张三"]},
            }
        }
        chat_rows_by_group = {
            "客户 LC-P2026001 项目群": [
                {"raw_json": '{"msgtime":"2026-07-06 09:00:00"}'},
                {"raw_json": '{"msgtime":"2026-07-06 13:00:00"}'},
                {"raw_json": '{"msgtime":"2026-07-06 18:00:00"}'},
                {"raw_json": '{"msgtime":"2026-07-11 10:00:00"}'},
            ]
        }

        result = db_dashboard._time_period_breakdown(groups, {}, chat_rows_by_group)
        item = result["items"][0]

        self.assertEqual(item["morning"]["count"], 1)
        self.assertEqual(item["afternoon"]["count"], 1)
        self.assertEqual(item["after_hours"]["count"], 1)
        self.assertEqual(item["weekend"]["count"], 1)
        self.assertIn("不包含周末", result["after_hours"])

    def test_multi_aftersaler_group_counts_only_each_owners_messages(self):
        messages = [
            {"raw_json": json.dumps({"from": "WuZhihao", "truename": "吴志浩", "msgtime": "2026-07-06 09:00:00"}, ensure_ascii=False)},
            {"raw_json": json.dumps({"from": "YangJiajun", "truename": "杨嘉俊", "msgtime": "2026-07-06 13:00:00"}, ensure_ascii=False)},
            {"raw_json": json.dumps({"from": "wmCustomer001", "truename": "吴志浩", "msgtime": "2026-07-06 14:00:00"}, ensure_ascii=False)},
            {"raw_json": json.dumps({"from": "OtherEmployee", "truename": "其他员工", "msgtime": "2026-07-06 15:00:00"}, ensure_ascii=False)},
        ]

        counts = db_dashboard._aftersaler_message_counts(
            ["吴志浩", "杨嘉俊"], messages, group_message_count=99,
        )

        self.assertEqual(counts, {"吴志浩": 1, "杨嘉俊": 1})
        self.assertEqual(
            db_dashboard._aftersaler_message_counts(["吴志浩"], messages, group_message_count=99),
            {"吴志浩": 4},
        )

    def test_multi_aftersaler_period_breakdown_uses_personal_speech(self):
        group_name = "客户多项目群"
        groups = {
            group_name: {"dimension": {"aftersalers": ["吴志浩", "杨嘉俊"]}},
        }
        chat_rows = {
            group_name: [
                {"raw_json": json.dumps({"from": "WuZhihao", "truename": "吴志浩", "msgtime": "2026-07-06 09:00:00"}, ensure_ascii=False)},
                {"raw_json": json.dumps({"from": "YangJiajun", "truename": "杨嘉俊", "msgtime": "2026-07-06 13:00:00"}, ensure_ascii=False)},
                {"raw_json": json.dumps({"from": "wmCustomer001", "truename": "客户甲", "msgtime": "2026-07-06 18:00:00"}, ensure_ascii=False)},
            ],
        }

        result = db_dashboard._time_period_breakdown(groups, {}, chat_rows)
        items = {item["aftersaler"]: item for item in result["items"]}

        self.assertEqual(items["吴志浩"]["total"], 1)
        self.assertEqual(items["吴志浩"]["morning"]["count"], 1)
        self.assertEqual(items["杨嘉俊"]["total"], 1)
        self.assertEqual(items["杨嘉俊"]["afternoon"]["count"], 1)

    def test_chat_msgtime_prefers_raw_json_msgtime(self):
        row = {
            "msgtime": "2026-07-06 09:00:00",
            "raw_json": '{"msgtime":"2026-07-06 10:30:00"}',
        }
        self.assertEqual(
            db_dashboard.chat_msgtime(row).strftime("%Y-%m-%d %H:%M:%S"),
            "2026-07-06 10:30:00",
        )

    def test_evidence_message_uses_raw_chat_msgtime(self):
        missed = [{"msgid": "m1", "content": "请问进度", "msgtime": "2026-07-06 09:00:00"}]
        raw_messages = [{
            "msgid": "m1",
            "content": "请问进度",
            "sender_name": "客户",
            "msgtime": "2026-07-06 18:30:00",
        }]

        result = db_dashboard._evidence_messages("unanswered", "", "", missed, raw_messages)

        self.assertEqual(result[0]["msgtime"], "2026-07-06 18:30:00")
        self.assertEqual(result[0]["sender_name"], "客户")

    def test_extract_missed_messages_parses_pure_text_with_multiple_messages(self):
        """用户实际场景：missedMessageList 存的是 '<sender>:<content>' 纯文本。"""
        text = '南枝:这个组Ss_Tcorolla只有2、3吗 \n 1呢,南枝:@李祖杰 这会能电话沟通下吗'
        result = db_dashboard._extract_missed_messages(text)
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["sender_name"], "南枝")
        self.assertIn("这个组Ss_Tcorolla", result[0]["content"])
        self.assertEqual(result[1]["sender_name"], "南枝")
        self.assertIn("@李祖杰", result[1]["content"])

    def test_extract_missed_messages_handles_json_with_msgTime_alias(self):
        """Java 端写入时使用 msgTime 字段名也应能解析。"""
        text = '[{"id":"m1","msgTime":"2026-07-11 09:30:00","text":"怎么弄","sender":"A"}]'
        result = db_dashboard._extract_missed_messages(text)
        self.assertEqual(result[0]["msgtime"], "2026-07-11 09:30:00")
        self.assertEqual(result[0]["content"], "怎么弄")

    def test_extract_missed_messages_fallback_for_plain_text(self):
        """不是 sender:content 格式也不是 JSON 时，整段当作一条 content。"""
        text = "这是一段普通文本，没有冒号分隔"
        result = db_dashboard._extract_missed_messages(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["content"], text)
        self.assertEqual(result[0]["sender_name"], "")

    def test_pure_text_missed_messages_get_msgtime_via_raw_chat_match(self):
        """纯文本格式解析出的漏回消息，能通过 content 匹配拿到 qx_chat 的 msgtime。"""
        text = '南枝:这个组Ss_Tcorolla只有2、3吗 \n 1呢,南枝:@李祖杰 这会能电话沟通下吗'
        missed = db_dashboard._extract_missed_messages(text)
        raw_messages = [{
            "msgid": "r1",
            "sender_name": "南枝",
            "content": "这个组Ss_Tcorolla只有2、3吗\n 1呢",
            "msgtime": "2026-07-11 13:42:15",
        }, {
            "msgid": "r2",
            "sender_name": "南枝",
            "content": "@李祖杰 这会能电话沟通下吗",
            "msgtime": "2026-07-11 13:45:30",
        }]

        display = db_dashboard._evidence_messages("unanswered", "", "", missed, raw_messages)

        self.assertEqual(len(display), 2)
        self.assertEqual(display[0]["msgtime"], "2026-07-11 13:42:15")
        self.assertEqual(display[1]["msgtime"], "2026-07-11 13:45:30")

    def test_extract_missed_messages_parses_sender_with_spaces_and_phone(self):
        """用户实际数据：发送人姓名带空格+手机号也能正确解析。"""
        text = (
            "@张兴瑞 @李祖杰 这些分别代表什么,"
            "张敏  13103738626:@李祖杰 每个通路具体哪些分子？,"
            "张敏  13103738626:「张敏  13103738626：@李祖杰 每个通路具体哪些分子？」"
        )
        result = db_dashboard._extract_missed_messages(text)
        self.assertEqual(len(result), 3)
        # 首条自由文本（无 sender）
        self.assertEqual(result[0]["sender_name"], "")
        self.assertIn("张兴瑞", result[0]["content"])
        # 第二条：sender 含空格+手机号
        self.assertEqual(result[1]["sender_name"], "张敏  13103738626")
        self.assertIn("每个通路具体哪些分子", result[1]["content"])
        # 第三条：含全角引号 + 内部全角冒号
        self.assertEqual(result[2]["sender_name"], "张敏  13103738626")
        self.assertTrue(result[2]["content"].startswith("「"))

    def test_display_chat_message_never_uses_roomid_as_sender(self):
        roomid = "wrROOM_123456"
        message = db_dashboard._display_chat_message({
            "roomid": roomid,
            "from": roomid,
            "truename": roomid,
            "content": "请问什么时候有结果？",
            "msgtime": "2026-07-11 10:00:00",
        })
        self.assertEqual(message["sender_name"], "未知发送人")
        self.assertEqual(message["sender_role"], "未知")

    def test_raw_message_role_uses_qx_group_member_type(self):
        grouped = db_dashboard._raw_messages_by_group(
            [{
                "chat_id": "room-1", "name": "测试群",
                "member_list_json": '[{"userid":"wmInternal001","name":"员工甲","type":1}]',
            }],
            [{
                "roomid": "room-1", "from": "wmInternal001", "truename": "员工甲",
                "content": "我来处理", "msgtime": "2026-07-11 10:00:00",
            }],
        )
        self.assertEqual(grouped["测试群"][0]["sender_role"], "员工")

    def test_unanswered_reconciliation_excludes_message_with_employee_reply(self):
        raw = [
            {
                "msgid": "c1", "sender_name": "客户甲", "sender_userid": "wmCustomer001",
                "sender_role": "客户", "roomid": "r1", "content": "什么时候出结果？",
                "msgtime": "2026-07-11 10:00:00", "time_source": "original_message",
            },
            {
                "msgid": "e1", "sender_name": "徐工", "sender_userid": "XuJun",
                "sender_role": "员工", "roomid": "r1", "content": "预计今天下午给您。",
                "msgtime": "2026-07-11 11:00:00", "time_source": "original_message",
            },
        ]
        result = db_dashboard._evaluate_unanswered_messages(
            [{"msgid": "c1", "content": "什么时候出结果？"}], raw,
            datetime(2026, 7, 11, 12, 0, 0),
        )
        self.assertEqual(result[0]["verification_status"], "answered_before_analysis")
        self.assertEqual(result[0]["reply_sender_name"], "徐工")

    def test_unanswered_reconciliation_keeps_unique_unanswered_customer_message(self):
        raw = [{
            "msgid": "c1", "sender_name": "客户甲", "sender_userid": "wmCustomer001",
            "sender_role": "客户", "roomid": "r1", "content": "什么时候出结果？",
            "msgtime": "2026-07-11 10:00:00", "time_source": "original_message",
        }]
        result = db_dashboard._evaluate_unanswered_messages(
            [
                {"msgid": "c1", "content": "什么时候出结果？"},
                {"msgid": "c1", "content": "什么时候出结果？"},
            ], raw, datetime(2026, 7, 11, 12, 0, 0),
        )
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["verification_status"], "unanswered")
        self.assertEqual(result[0]["msgtime"], "2026-07-11 10:00:00")

    def test_unanswered_reconciliation_rejects_phone_number(self):
        raw = [{
            "msgid": "c1", "sender_name": "客户甲", "sender_userid": "wmCustomer001",
            "sender_role": "客户", "roomid": "r1", "content": "15642622729",
            "msgtime": "2026-07-11 10:00:00", "time_source": "original_message",
        }]
        result = db_dashboard._evaluate_unanswered_messages(
            [{"msgid": "c1", "content": "15642622729"}], raw,
            datetime(2026, 7, 11, 12, 0, 0),
        )
        self.assertEqual(result[0]["verification_status"], "no_action_needed")

    def test_unanswered_reconciliation_rejects_declarative_context(self):
        raw = [{
            "msgid": "c1", "sender_name": "客户甲", "sender_userid": "wmCustomer001",
            "sender_role": "客户", "roomid": "r1", "content": "刚刚那个截图就是KEGG里面的结果",
            "msgtime": "2026-07-11 10:00:00", "time_source": "original_message",
        }]
        result = db_dashboard._evaluate_unanswered_messages(
            [{"msgid": "c1", "content": "刚刚那个截图就是KEGG里面的结果"}], raw,
            datetime(2026, 7, 11, 12, 0, 0),
        )
        self.assertEqual(result[0]["verification_status"], "no_action_needed")

    def test_unanswered_reconciliation_checks_reply_body_not_quoted_question(self):
        content = "「客户：能打个折吗？」\n- - - - - - - -\n收到老师，我跟领导申请一下"
        raw = [{
            "msgid": "c1", "sender_name": "外部联系人", "sender_userid": "wmCustomer001",
            "sender_role": "客户", "roomid": "r1", "content": content,
            "msgtime": "2026-07-11 10:00:00", "time_source": "original_message",
        }]
        result = db_dashboard._evaluate_unanswered_messages(
            [{"msgid": "c1", "content": content}], raw,
            datetime(2026, 7, 11, 12, 0, 0),
        )
        self.assertEqual(result[0]["verification_status"], "no_action_needed")

    def test_get_evidence_does_not_treat_analysis_time_as_message_time(self):
        """没有原始消息证据时不能用 CREATEDTIME 冒充消息发送时间。"""
        from contextlib import contextmanager
        db_dashboard._evidence_cache.clear()

        mock_row = {
            "id": 1,
            "groupName": "客户 LC-P2026001 项目群",
            "isMissedMessage": "1",
            "missedMessageList": "客户A:催一下结果,客户A:还有进度吗",
            "CREATEDTIME": "2026-07-11 14:00:00",
            "messageToDayCount": 0,
            "highFrequencyWords": "",
            "customerEmotionAnalysis": "",
            "saleEmotionAnalysis": "",
            "afterSalesTopicAnalysis": "",
        }
        test_dimension = {
            "客户 LC-P2026001 项目群": {
                "codes": ["LC-P2026001"], "regions": [], "aftersalers": [],
                "projects": [{
                    "project_code": "LC-P2026001", "category_l2": "常规转录组",
                    "region": "", "raw_aftersaler": "",
                }],
            },
        }

        @contextmanager
        def fake_database(_operation):
            yield object()

        with (
            patch.object(db_dashboard, "database", fake_database),
            patch.object(db_dashboard, "_latest_rows", return_value=([mock_row], 1)),
            patch.object(db_dashboard, "_load_dimensions", return_value=(test_dimension, {"matched_project_codes": 1})),
            patch.object(db_dashboard, "_query_group_rows", return_value=[]),
            patch.object(db_dashboard, "_query_chat_rows", return_value=[]),
        ):
            response = db_dashboard.get_evidence(
                "unanswered", "custom", "2026-07-11", "2026-07-11",
            )

        matched = [
            item for item in response.get("items", [])
            if item.get("group_name") == "客户 LC-P2026001 项目群"
        ]
        self.assertFalse(matched)
        self.assertEqual(response["verification"].get("unverified"), 2)

    def test_get_evidence_excludes_unverified_analysis_text(self):
        """模型文本匹配不到原始消息时进入待核查，不进入确认漏回。"""
        from contextlib import contextmanager
        db_dashboard._evidence_cache.clear()

        mock_row = {
            "id": 1,
            "groupName": "客户 LC-P2026001 项目群",
            "isMissedMessage": "1",
            "missedMessageList": "客户A:催一下结果",
            "CREATEDTIME": "2026-07-11 14:00:00",
            "messageToDayCount": 0,
            "highFrequencyWords": "",
            "customerEmotionAnalysis": "",
            "saleEmotionAnalysis": "",
            "afterSalesTopicAnalysis": "",
        }
        test_dimension = {
            "客户 LC-P2026001 项目群": {
                "codes": ["LC-P2026001"], "regions": [], "aftersalers": [],
                "projects": [{
                    "project_code": "LC-P2026001", "category_l2": "常规转录组",
                    "region": "", "raw_aftersaler": "",
                }],
            },
        }

        @contextmanager
        def fake_database(_operation):
            yield object()

        with (
            patch.object(db_dashboard, "database", fake_database),
            patch.object(db_dashboard, "_latest_rows", return_value=([mock_row], 1)),
            patch.object(db_dashboard, "_load_dimensions", return_value=(test_dimension, {"matched_project_codes": 1})),
            patch.object(db_dashboard, "_query_group_rows", return_value=[]),
            patch.object(db_dashboard, "_query_chat_rows", return_value=[]),
        ):
            response = db_dashboard.get_evidence(
                "unanswered", "custom", "2026-07-11", "2026-07-11",
            )

        matched = [
            item for item in response.get("items", [])
            if item.get("group_name") == "客户 LC-P2026001 项目群"
        ]
        self.assertFalse(matched)
        self.assertEqual(response["verification"].get("unverified"), 1)

    def test_get_evidence_excludes_unmatched_text_even_when_other_raw_rows_have_time(self):
        """群内其他消息的时间不能被宽松匹配到漏回候选上。"""
        from contextlib import contextmanager
        db_dashboard._evidence_cache.clear()

        mock_row = {
            "id": 1,
            "groupName": "客户 LC-P2026001 项目群",
            "isMissedMessage": "1",
            "missedMessageList": (
                "@张兴瑞 @李祖杰 这些分别代表什么,"
                "张敏  13103738626:@李祖杰 每个通路具体哪些分子？,"
                "张敏  13103738626:「张敏  13103738626：@李祖杰 每个通路具体哪些分子？」"
            ),
            "CREATEDTIME": "2026-07-11 14:00:00",
            "messageToDayCount": 0,
            "highFrequencyWords": "",
            "customerEmotionAnalysis": "",
            "saleEmotionAnalysis": "",
            "afterSalesTopicAnalysis": "",
        }
        test_dimension = {
            "客户 LC-P2026001 项目群": {
                "codes": ["LC-P2026001"], "regions": [], "aftersalers": [],
                "projects": [{
                    "project_code": "LC-P2026001", "category_l2": "常规转录组",
                    "region": "", "raw_aftersaler": "",
                }],
            },
        }

        @contextmanager
        def fake_database(_operation):
            yield object()

        with (
            patch.object(db_dashboard, "database", fake_database),
            patch.object(db_dashboard, "_latest_rows", return_value=([mock_row], 1)),
            patch.object(db_dashboard, "_load_dimensions", return_value=(test_dimension, {"matched_project_codes": 1})),
            patch.object(db_dashboard, "_query_group_rows", return_value=[]),
            patch.object(db_dashboard, "_query_chat_rows", return_value=[]),
        ):
            response = db_dashboard.get_evidence(
                "unanswered", "custom", "2026-07-11", "2026-07-11",
            )

        matched = [
            item for item in response.get("items", [])
            if item.get("group_name") == "客户 LC-P2026001 项目群"
        ]
        self.assertFalse(matched)
        self.assertGreaterEqual(response["verification"].get("unverified", 0), 1)

    def test_lims_api_dimensions_use_computed_final_after_saler(self):
        group_name = "Customer LC-P2026001 support"
        records = {
            "LC-P2026001": [{
                "projectCode": "LC-P2026001",
                "afterSaler": "RawAfter",
                "finalAfterSaler": "LegacyFinal",
                "members": "SomeoneElse",
                "orgName": "RegionA",
                "productName": "ProductA",
                "productBigSortOne": "CategoryA",
                "productBigSortThree": "CategoryA",
            }]
        }
        mapping_snapshot = {
            "available": True, "version_id": 1, "effective_month": "2026-07",
            "revision": 1, "reason": "ok", "rules": [{
                "id": 1, "version_id": 1, "product_name": "ProductA",
                "product_keywords": ["CategoryA"], "region_name": "RegionA",
                "lims_aftersaler": "RawAfter", "actual_aftersaler": "RealAfter",
            }],
        }
        dimensions, quality = db_dashboard._dimensions_from_lims_api(
            {group_name: ["LC-P2026001"]},
            records,
            {"requests": 1, "records": 1, "errors": 0},
            mapping_snapshot,
        )

        self.assertEqual(dimensions[group_name]["aftersalers"], ["RealAfter"])
        self.assertEqual(dimensions[group_name]["raw_aftersalers"], ["RawAfter"])
        self.assertEqual(dimensions[group_name]["projects"][0]["aftersaler_source"], "mapping")
        self.assertEqual(dimensions[group_name]["tentative_aftersalers"], [])
        self.assertEqual(quality["groups_with_aftersaler"], 1)

    def test_lims_unavailable_does_not_use_business_table_fallback(self):
        rows = [{"groupName": "Customer LC-P2026001 support", "member": ""}]
        with patch.object(
            db_dashboard,
            "fetch_lims_base_data",
            return_value=({}, {"available": False, "requests": 1, "records": 0, "errors": 1}),
        ):
            dimensions, quality = db_dashboard._load_dimensions(None, rows)

        self.assertEqual(dimensions["Customer LC-P2026001 support"]["projects"], [])
        self.assertEqual(dimensions["Customer LC-P2026001 support"]["dimension_source"], "lims_unavailable")
        self.assertEqual(quality["matched_project_codes"], 0)
        self.assertEqual(quality["lims_source"], "base_data_api_unavailable")

    def test_latest_rows_filters_non_focus_groups_from_rows_and_raw_count(self):
        latest_rows = [
            {"groupName": "客户 LC-P2026001 项目群"},
            {"groupName": "普通交流群"},
        ]
        raw_rows = [
            {"groupName": "客户 LC-P2026001 项目群", "count": 3},
            {"groupName": "普通交流群", "count": 20},
            {"groupName": "客户 LC-SP2026002 特殊项目群", "count": 2},
        ]
        with patch.object(db_dashboard, "_query", side_effect=[latest_rows, raw_rows]):
            rows, raw_count = db_dashboard._latest_rows(None, date(2026, 7, 1), date(2026, 7, 8))

        self.assertEqual([row["groupName"] for row in rows], ["客户 LC-P2026001 项目群"])
        self.assertEqual(raw_count, 5)

    def test_query_chat_rows_pushes_period_filter_into_sql(self):
        with patch.object(db_dashboard, "_query_by_chunks", return_value=[]) as query:
            rows = db_dashboard._query_chat_rows(
                object(),
                [{"chat_id": "room-1"}],
                date(2026, 7, 1),
                date(2026, 7, 16),
            )

        self.assertEqual(rows, [])
        sql = query.call_args.args[2]
        self.assertIn("msgtime >= %s", sql)
        self.assertIn("msgtime < %s", sql)
        self.assertEqual(
            query.call_args.kwargs["suffix_params"],
            ["2026-07-01", "2026-07-17"],
        )

    def test_lims_fresh_cache_avoids_remote_request(self):
        with db_dashboard._lims_cache_lock:
            db_dashboard._lims_cache["LC-P1"] = (
                time.time(), [{"projectCode": "LC-P1"}],
            )

        records, stats = db_dashboard.fetch_lims_base_data(["LC-P1"])

        self.assertEqual(records["LC-P1"][0]["projectCode"], "LC-P1")
        self.assertEqual(stats["requests"], 0)
        self.assertEqual(stats["cache_hits"], 1)

    def test_lims_cache_survives_process_restart(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            cache_file = Path(temp_dir) / "dashboard_lims_cache.json"
            cache_file.write_text(json.dumps({
                "saved_at": time.time() - 3600,
                "records": {"LC-P1": [{"projectCode": "LC-P1"}]},
            }), encoding="utf-8")
            with patch.object(db_dashboard, "LIMS_CACHE_FILE", cache_file):
                with db_dashboard._lims_cache_lock:
                    db_dashboard._lims_cache.clear()
                    db_dashboard._lims_stale.clear()
                    db_dashboard._lims_cache_loaded = False
                db_dashboard._load_persisted_lims_cache()
                with db_dashboard._lims_cache_lock:
                    self.assertEqual(
                        db_dashboard._lims_stale["LC-P1"][0]["projectCode"],
                        "LC-P1",
                    )

    def test_project_code_diagnostics_lists_missing_groups_and_unmatched_codes(self):
        diagnostics = db_dashboard._project_code_diagnostics(
            ["No project group", "Alpha LC-P2026001", "Beta LC-P404"],
            {
                "No project group": {"codes": []},
                "Alpha LC-P2026001": {"codes": ["LC-P2026001"]},
                "Beta LC-P404": {"codes": ["LC-P404"]},
            },
            {"LC-P2026001": [{"projectCode": "LC-P2026001"}]},
        )

        self.assertEqual(diagnostics["groups_without_project_code"], ["No project group"])
        self.assertEqual(diagnostics["unmatched_codes"], ["LC-P404"])
        self.assertEqual(diagnostics["code_to_groups"]["LC-P404"], ["Beta LC-P404"])
        self.assertEqual(diagnostics["groups_without_lims_link"], ["Beta LC-P404"])


class DashboardDrilldownTests(unittest.TestCase):
    @staticmethod
    def _shared_business_snapshot():
        group_name = "客户 LC-P100 LC-P101 项目群"
        chat_rows = [
            {
                "id": index, "roomid": "room-1", "msgtime": f"2026-07-06 09:0{index}:00",
                "raw_json": json.dumps({
                    "from": f"wmCustomer{index}", "truename": f"客户{index}",
                    "content": f"消息{index}", "msgtime": f"2026-07-06 09:0{index}:00",
                }, ensure_ascii=False),
            }
            for index in range(1, 4)
        ]
        projects = [
            {
                "project_code": code, "product_name": code, "region": "华东",
                "raw_aftersaler": "张三", "final_aftersaler": "张三",
                "aftersaler_source": "mapping", "category_l2": "常规转录组",
                "category_l3": "mRNA", "key_account": "重点客户甲", "active_day": 20,
            }
            for code in ("LC-P100", "LC-P101")
        ]
        latest_rows = [{
            "id": 1, "groupName": group_name, "CREATEDTIME": "2026-07-06 23:00:00",
            "messageToDayCount": 999, "isMissedMessage": 0,
            "customerEmotionAnalysis": "", "saleEmotionAnalysis": "", "highFrequencyWords": "",
        }]
        group_rows = [{"name": group_name, "chat_id": "room-1", "member_list_json": "[]"}]
        dimensions = {
            group_name: {
                "codes": ["LC-P100", "LC-P101"], "projects": projects,
                "regions": ["华东"], "aftersalers": ["张三"], "raw_aftersalers": ["张三"],
                "tentative_aftersalers": [],
            }
        }
        quality = {
            "project_codes": 2, "matched_project_codes": 2,
            "product_projects": 2, "matched_products": 2,
            "groups_with_aftersaler": 1, "groups_with_confirmed_aftersaler": 1,
        }
        normalized = db_dashboard._raw_messages_by_group(group_rows, chat_rows)
        return {
            "latest_rows": latest_rows, "raw_rows": latest_rows,
            "raw_count_before_filter": 1, "dimensions": dimensions, "quality": quality,
            "filter_options": db_dashboard._dashboard_filter_options(dimensions),
            "group_names": [group_name], "project_codes": ["LC-P100", "LC-P101"],
            "group_rows": group_rows, "chat_rows": chat_rows,
            "chat_rows_by_group": normalized, "audit_chat_rows": chat_rows,
            "audit_raw_messages": normalized,
        }

    def test_business_message_metrics_share_qx_chat_and_deduplicate_group_category(self):
        snapshot = self._shared_business_snapshot()
        overview = db_dashboard._build_overview(
            date(2026, 7, 1), date(2026, 7, 17), "custom", snapshot=snapshot,
        )

        aftersaler = next(item for item in overview["business"]["aftersalers"] if item["name"] == "张三")
        product = next(item for item in overview["business"]["product_categories"] if item["category"] == "常规转录组")
        product_leaf = overview["business"]["product_hierarchy"][0]["children"][0]

        self.assertEqual(aftersaler["message_count"], 3)
        self.assertEqual(product["project_count"], 2)
        self.assertEqual(product["message_count"], 3)
        self.assertEqual(product_leaf["message_count"], 3)

    def test_drilldown_dashboard_value_uses_same_snapshot_as_detail_rows(self):
        snapshot = self._shared_business_snapshot()
        scope_key = db_dashboard._dashboard_snapshot_scope_key(
            date(2026, 7, 1), date(2026, 7, 17), "custom",
        )
        snapshot_id = db_dashboard._store_dashboard_snapshot(scope_key, snapshot)

        @contextmanager
        def fake_database(_operation):
            yield object()

        with (
            patch.object(db_dashboard, "database", fake_database),
            patch.object(db_dashboard, "_query_qx_raw_scope", side_effect=AssertionError("snapshot must be reused")),
            patch.object(db_dashboard, "get_overview", side_effect=AssertionError("cached overview must not be used")),
        ):
            aftersaler = db_dashboard._build_drilldown(
                "business.aftersaler", "message_count", period="custom",
                start_date="2026-07-01", end_date="2026-07-17", dimension_value="张三",
                snapshot_id=snapshot_id,
            )
            product = db_dashboard._build_drilldown(
                "business.product", "message_count", period="custom",
                start_date="2026-07-01", end_date="2026-07-17", dimension_value="常规转录组",
                snapshot_id=snapshot_id,
            )

        self.assertEqual(aftersaler["reconciliation"]["dashboard_value"], 3)
        self.assertEqual(aftersaler["reconciliation"]["detail_value"], 3)
        self.assertTrue(aftersaler["reconciliation"]["consistent"])
        self.assertEqual(product["reconciliation"]["dashboard_value"], 3)
        self.assertEqual(product["reconciliation"]["detail_value"], 3)
        self.assertTrue(product["reconciliation"]["consistent"])

    def test_drilldown_target_and_measure_are_allowlisted(self):
        with self.assertRaisesRegex(ValueError, "unsupported drilldown target"):
            db_dashboard._build_drilldown("raw.sql", "message_count")
        with self.assertRaisesRegex(ValueError, "unsupported drilldown measure"):
            db_dashboard._build_drilldown("summary.total_groups", "message_count")

    def test_drilldown_pagination_keeps_reconciliation_total(self):
        built = {
            "target": "summary.total_groups", "measure": "group_count",
            "title": "服务群明细", "level": "group", "definition": "test",
            "period": {"name": "month", "start": "2026-07-01", "end": "2026-07-31"},
            "filters": {}, "reconciliation": {"dashboard_value": 45, "detail_value": 45},
            "columns": [{"key": "group_name", "label": "群聊名称"}],
            "rows": [{"group_name": f"群{i}"} for i in range(45)],
        }
        with patch.object(db_dashboard, "_build_drilldown", return_value=built):
            result = db_dashboard.get_drilldown(
                "summary.total_groups", "group_count", page=2, page_size=20,
            )
        self.assertEqual(result["total"], 45)
        self.assertEqual(result["total_pages"], 3)
        self.assertEqual(len(result["items"]), 20)
        self.assertEqual(result["items"][0]["group_name"], "群20")
        self.assertEqual(result["reconciliation"]["detail_value"], 45)

    def test_drilldown_masks_phone_and_wechat_identifiers(self):
        masked = db_dashboard._mask_drilldown_text("联系 13812345678，账号 wxid_abcdefghijk")
        self.assertEqual(masked, "联系 138****5678，账号 wxid_abcd***")

    def test_drilldown_excel_contains_scope_and_formula_safe_detail(self):
        from openpyxl import load_workbook

        built = {
            "target": "summary.total_groups", "measure": "group_count",
            "title": "服务群明细", "level": "group", "definition": "一条一群",
            "period": {"name": "custom", "start": "2026-07-01", "end": "2026-07-02"},
            "filters": {"region": "华东", "aftersaler": "", "category": "", "key_account": "", "dimension_value": "", "secondary_value": "", "search": ""},
            "reconciliation": {"dashboard_value": 1, "detail_value": 1, "consistent": True},
            "columns": [{"key": "group_name", "label": "群聊名称"}],
            "rows": [{"group_name": "=2+2"}],
        }
        with patch.object(db_dashboard, "_build_drilldown", return_value=built):
            content = db_dashboard.get_drilldown_export_excel(
                target="summary.total_groups", measure="group_count",
            )
        workbook = load_workbook(BytesIO(content), read_only=True)
        self.assertEqual(workbook.sheetnames, ["导出说明", "明细数据"])
        self.assertEqual(workbook["明细数据"]["A2"].value, "'=2+2")
        self.assertEqual(workbook["导出说明"]["B2"].value, "服务群明细")

    def test_drilldown_export_rejects_oversized_result(self):
        built = {
            "target": "summary.total_groups", "measure": "group_count",
            "title": "服务群明细", "level": "group", "definition": "test",
            "period": {"name": "month", "start": "2026-07-01", "end": "2026-07-31"},
            "filters": {}, "reconciliation": {}, "columns": [],
            "rows": [{"group_name": "A"}, {"group_name": "B"}],
        }
        with (
            patch.object(db_dashboard, "_build_drilldown", return_value=built),
            patch.object(db_dashboard, "DASHBOARD_DRILLDOWN_EXPORT_MAX_ROWS", 1),
        ):
            with self.assertRaisesRegex(ValueError, "超过单次导出上限"):
                db_dashboard.get_drilldown_export_excel(
                    target="summary.total_groups", measure="group_count",
                )

    def test_drilldown_api_rejects_unknown_target(self):
        from fastapi.testclient import TestClient
        from api.main import app

        response = TestClient(app).get(
            "/api/v1/dashboard/drilldown",
            params={"target": "unknown.target", "measure": "group_count"},
        )
        self.assertEqual(response.status_code, 422)
        self.assertIn("unsupported drilldown target", response.text)

    def test_drilldown_api_returns_preview_and_excel(self):
        from fastapi.testclient import TestClient
        from api.main import app

        preview = {
            "title": "服务群明细", "items": [{"group_name": "群A"}],
            "total": 1, "page": 1, "page_size": 20, "total_pages": 1,
        }
        client = TestClient(app)
        with patch.object(db_dashboard, "get_drilldown", return_value=preview):
            response = client.get(
                "/api/v1/dashboard/drilldown",
                params={"target": "summary.total_groups", "measure": "group_count"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["data"]["items"][0]["group_name"], "群A")

        with patch.object(db_dashboard, "get_drilldown_export_excel", return_value=b"xlsx"):
            response = client.get(
                "/api/v1/dashboard/drilldown/export",
                params={"target": "summary.total_groups", "measure": "group_count"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, b"xlsx")
        self.assertIn("attachment", response.headers["content-disposition"])


class DashboardCacheTests(unittest.TestCase):
    def setUp(self):
        db_dashboard.clear_cache()

    def test_cache_key_isolated_by_business_filter(self):
        base = {
            "meta": {}, "summary": {}, "service_quality": {}, "communication": {},
            "business": {}, "cross_analysis": {}, "data_quality": {},
        }
        with patch.object(db_dashboard, "_build_overview", return_value=base) as build:
            db_dashboard.get_overview("month", region="华东")
            db_dashboard.get_overview("month", region="华南")
            db_dashboard.get_overview("month", region="华东")
        self.assertEqual(build.call_count, 2)

    def test_evidence_result_is_cached_for_retries(self):
        result = {"items": [], "total": 0, "page": 1, "total_pages": 1}
        with patch.object(db_dashboard, "_build_evidence", return_value=result) as build:
            first = db_dashboard.get_evidence("unanswered")
            second = db_dashboard.get_evidence("unanswered")

        self.assertEqual(first, second)
        self.assertEqual(build.call_count, 1)

    def test_simultaneous_overview_requests_share_one_build(self):
        base = {
            "meta": {}, "summary": {}, "service_quality": {}, "communication": {},
            "business": {}, "cross_analysis": {}, "data_quality": {},
        }
        barrier = threading.Barrier(2)

        def build(*_args, **_kwargs):
            time.sleep(0.05)
            return base

        def request():
            barrier.wait()
            return db_dashboard.get_overview("month")

        with patch.object(db_dashboard, "_build_overview", side_effect=build) as mocked:
            with ThreadPoolExecutor(max_workers=2) as pool:
                results = list(pool.map(lambda _value: request(), range(2)))

        self.assertEqual(mocked.call_count, 1)
        self.assertEqual({item["meta"]["cache"] for item in results}, {"miss", "coalesced"})


if __name__ == "__main__":
    unittest.main()
