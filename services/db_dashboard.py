"""Read-only dashboard analytics built from the existing MySQL analysis tables."""

from __future__ import annotations

import copy
import json
import os
import re
import threading
import time
import uuid
from collections import Counter, defaultdict
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from loguru import logger
from services import aftersaler_mapping
from services.preprocessor import infer_sender_role, parse_members_payload, safe_sender_name


CACHE_TTL_SECONDS = 300
EVIDENCE_CACHE_TTL_SECONDS = 120
LIMS_CACHE_TTL_SECONDS = 1800
LIMS_REQUEST_BUDGET_SECONDS = 10.0
UNANSWERED_RESULT_TABLE = "qx_analysis_result_sale"
DASHBOARD_DRILLDOWN_EXPORT_MAX_ROWS = max(
    1, int(os.getenv("DASHBOARD_DRILLDOWN_EXPORT_MAX_ROWS", "100000"))
)
LIMS_CACHE_FILE = Path(os.getenv("DASHBOARD_LIMS_CACHE_FILE", "logs/dashboard_lims_cache.json"))
PROJECT_CODE_RE = re.compile(r"LC-(?:SP|P)\d+(?![A-Z0-9])", re.IGNORECASE)
ALLOWED_PRODUCT_L2 = {"常规转录组", "表观组学", "微生物"}
PROJECT_ATTENTION_STATUSES = ("问题项目", "暂不交付")
_cache: Dict[str, Tuple[float, dict]] = {}
_last_success: Dict[str, dict] = {}
_evidence_cache: Dict[str, Tuple[float, dict]] = {}
_dashboard_snapshot_cache: Dict[str, Tuple[float, str, dict]] = {}
_overview_inflight: Dict[str, threading.Event] = {}
_evidence_inflight: Dict[str, threading.Event] = {}
_cache_lock = threading.Lock()
_lims_cache: Dict[str, Tuple[float, List[dict]]] = {}
_lims_stale: Dict[str, List[dict]] = {}
_lims_cache_lock = threading.Lock()
_lims_refresh_lock = threading.Lock()
_lims_cache_loaded = False


DRILLDOWN_TARGETS = {
    "summary.total_groups": {"title": "服务群明细", "level": "group", "measures": {"group_count"}},
    "summary.total_messages": {"title": "沟通消息明细", "level": "message", "measures": {"message_count"}},
    "summary.project_groups": {"title": "项目群明细", "level": "group", "measures": {"group_count"}},
    "summary.regions": {"title": "销售区域明细", "level": "dimension", "measures": {"dimension_count"}},
    "summary.aftersalers": {"title": "售后人员明细", "level": "dimension", "measures": {"dimension_count"}},
    "summary.product_categories": {"title": "产品类别明细", "level": "dimension", "measures": {"dimension_count"}},
    "summary.key_accounts": {"title": "重点客户明细", "level": "dimension", "measures": {"dimension_count"}},
    "communication.daily": {"title": "每日趋势明细", "level": "auto", "measures": {"message_count", "group_count", "missed_count"}},
    "communication.period": {"title": "时段消息明细", "level": "message", "measures": {"morning", "afternoon", "after_hours", "weekend", "total"}},
    "communication.top_group": {"title": "高消息量群聊明细", "level": "message", "measures": {"message_count"}},
    "communication.active_duration": {"title": "群活跃周期明细", "level": "group", "measures": {"group_count"}},
    "communication.highfreq": {"title": "高频主题证据", "level": "message", "measures": {"occurrence_count"}},
    "business.region": {"title": "销售区域下钻", "level": "auto", "measures": {"group_count", "message_count"}},
    "business.aftersaler": {"title": "售后人员下钻", "level": "auto", "measures": {"project_count", "group_count", "message_count"}},
    "business.product": {"title": "产品分类下钻", "level": "auto", "measures": {"project_count", "group_count", "message_count"}},
    "business.key_account": {"title": "重点客户下钻", "level": "auto", "measures": {"project_count", "group_count", "message_count"}},
    "business.project_year": {"title": "项目年份明细", "level": "project", "measures": {"project_count"}},
    "business.project_attention": {"title": "异常交付项目明细", "level": "auto", "measures": {"project_count", "group_count", "message_count"}},
    "cross.region_sales": {"title": "区域销售结构明细", "level": "auto", "measures": {"group_count", "message_count"}},
    "cross.region_after": {"title": "区域售后结构明细", "level": "auto", "measures": {"group_count", "message_count"}},
    "cross.region_product": {"title": "区域产品结构明细", "level": "auto", "measures": {"group_count", "message_count"}},
    "service.unanswered": {"title": "待回复消息明细", "level": "message", "measures": {"missed_count"}},
    "service.answered": {"title": "已响应或无待回复群聊", "level": "group", "measures": {"group_count"}},
    "service.customer_positive": {"title": "客户好评分析明细", "level": "analysis", "measures": {"analysis_count"}},
    "service.customer_negative": {"title": "客户负向证据", "level": "message", "measures": {"analysis_count"}},
    "service.employee_positive": {"title": "积极服务分析明细", "level": "analysis", "measures": {"analysis_count"}},
    "service.employee_negative": {"title": "服务负向证据", "level": "message", "measures": {"analysis_count"}},
    "quality.groups_without_project_code": {"title": "未提取项目号群", "level": "group", "measures": {"group_count"}},
    "quality.unmatched_project_codes": {"title": "LIMS 未返回项目号", "level": "project_code", "measures": {"project_count"}},
    "quality.groups_without_lims_link": {"title": "完全无 LIMS 关联群", "level": "group", "measures": {"group_count"}},
    "quality.mapping_fallback_records": {"title": "售后映射回退明细", "level": "project", "measures": {"project_count"}},
    "quality.mapping_conflict_records": {"title": "售后映射冲突明细", "level": "project", "measures": {"project_count"}},
    "quality.duplicate_rows_removed": {"title": "被去重分析记录", "level": "analysis", "measures": {"analysis_count"}},
}


@contextmanager
def database(operation: str):
    """Open an isolated connection so concurrent requests never share protocol state."""
    import pymysql
    from config.settings import settings

    started = time.perf_counter()
    logger.info("dashboard.db.connect.start operation={}", operation)
    conn = pymysql.connect(
        host=settings.MYSQL_HOST,
        port=settings.MYSQL_PORT,
        user=settings.MYSQL_USER,
        password=settings.MYSQL_PASSWORD,
        database=settings.MYSQL_DATABASE,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=8,
        read_timeout=30,
        write_timeout=30,
        autocommit=True,
    )
    try:
        yield conn
    finally:
        conn.close()
        logger.info(
            "dashboard.db.connect.end operation={} elapsed_ms={:.1f}",
            operation,
            (time.perf_counter() - started) * 1000,
        )


def _query(conn, operation: str, sql: str, params: Iterable[Any] = ()) -> List[dict]:
    started = time.perf_counter()
    logger.info("dashboard.db.query.start operation={}", operation)
    with conn.cursor() as cursor:
        cursor.execute(sql, tuple(params))
        rows = list(cursor.fetchall())
    logger.info(
        "dashboard.db.query.end operation={} rows={} elapsed_ms={:.1f}",
        operation,
        len(rows),
        (time.perf_counter() - started) * 1000,
    )
    return rows


def _table_columns(conn, table_name: str) -> set[str]:
    rows = _query(conn, f"schema.{table_name}", f"SHOW COLUMNS FROM {table_name}")
    return {str(row.get("Field") or "").lower() for row in rows}


def _first_existing_column(columns: set[str], *candidates: str) -> Optional[str]:
    for candidate in candidates:
        if candidate.lower() in columns:
            return candidate
    return None


def _field_value_case_insensitive(item: dict, *names: str) -> Any:
    if not item:
        return None
    lowered = {str(key).lower(): value for key, value in item.items()}
    for name in names:
        if name in item and item[name] not in (None, ""):
            return item[name]
        value = lowered.get(name.lower())
        if value not in (None, ""):
            return value
    return None


def parse_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"-?\d+", str(value))
    return int(match.group()) if match else None


def parse_active_day(value: Any) -> Optional[float]:
    """原样保留 LIMS activeDay 的精度，不取整。

    接受 12 / 12.5 / 0.08 / "0.08天" / None。
    返回 float；不命中时返回 None。
    """
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"-?\d+(\.\d+)?", text)
    return float(match.group()) if match else None


def extract_lims_active_day(item: dict) -> Optional[float]:
    return parse_active_day(_field_value_case_insensitive(
        item,
        "activeDay", "active_day", "activeDays", "activeLay", "activeLlay",
        "activelay", "activellay", "ACTIVE_DAY", "ACTIVEDAY",
    ))


def clear_cache() -> None:
    global _lims_cache_loaded
    with _cache_lock:
        _cache.clear()
        _last_success.clear()
        _evidence_cache.clear()
        _dashboard_snapshot_cache.clear()
    with _lims_cache_lock:
        _lims_cache.clear()
        _lims_stale.clear()
        _lims_cache_loaded = False
    try:
        LIMS_CACHE_FILE.unlink(missing_ok=True)
    except OSError as exc:
        logger.warning("lims.cache.clear_failed path={} error={}", LIMS_CACHE_FILE, exc)


def clear_analytics_cache() -> None:
    """Invalidate derived dashboard results without discarding reusable LIMS data."""
    with _cache_lock:
        _cache.clear()
        _last_success.clear()
        _evidence_cache.clear()
        _dashboard_snapshot_cache.clear()


def _dashboard_snapshot_scope_key(
    start: date,
    end: date,
    period: str,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> str:
    return json.dumps(
        [start.isoformat(), end.isoformat(), period, region, aftersaler, category, key_account],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _store_dashboard_snapshot(
    scope_key: str,
    snapshot: dict,
    token: Optional[str] = None,
) -> str:
    """Keep or rehydrate the exact inputs behind a rendered dashboard."""
    now = time.time()
    token = token or uuid.uuid4().hex
    snapshot["_snapshot_id"] = token
    with _cache_lock:
        expired = [
            key for key, (created_at, _scope_key, _snapshot) in _dashboard_snapshot_cache.items()
            if now - created_at >= CACHE_TTL_SECONDS
        ]
        for key in expired:
            _dashboard_snapshot_cache.pop(key, None)
        _dashboard_snapshot_cache[token] = (now, scope_key, snapshot)
    return token


def _get_dashboard_snapshot(token: str, scope_key: str) -> Tuple[Optional[dict], str]:
    if not token:
        return None, "not_requested"
    now = time.time()
    with _cache_lock:
        cached = _dashboard_snapshot_cache.get(token)
        if not cached:
            logger.warning("dashboard.snapshot.miss token_prefix={}", token[:8])
            return None, "missing"
        created_at, cached_scope_key, snapshot = cached
        if now - created_at >= CACHE_TTL_SECONDS:
            _dashboard_snapshot_cache.pop(token, None)
            logger.warning(
                "dashboard.snapshot.expired token_prefix={} age_seconds={:.1f}",
                token[:8],
                now - created_at,
            )
            return None, "expired"
        if cached_scope_key != scope_key:
            logger.warning("dashboard.snapshot.scope_mismatch token_prefix={}", token[:8])
            return None, "scope_mismatch"
        return snapshot, "hit"


def extract_project_codes(group_name: str) -> List[str]:
    if not group_name:
        return []
    codes = []
    for match in PROJECT_CODE_RE.findall(group_name.upper()):
        code = match.rstrip("-")
        if code not in codes:
            codes.append(code)
    return codes


def is_focus_group_name(group_name: str) -> bool:
    return bool(extract_project_codes(group_name))


def get_project_codes(group_name: str) -> List[str]:
    return extract_project_codes(group_name)


def parse_count_map(value: Any) -> Dict[str, int]:
    """Parse the database's comma-separated `label: count` fields."""
    if value is None:
        return {}
    text = str(value).strip()
    if not text:
        return {}
    result: Dict[str, int] = {}
    for part in re.split(r"[,，]", text):
        part = part.strip()
        if not part or ":" not in part and "：" not in part:
            continue
        key, raw = re.split(r"[:：]", part, maxsplit=1)
        match = re.search(r"-?\d+", raw)
        if match:
            result[key.strip().strip("\"'")] = int(match.group())
    return result


def parse_emotion_field(value: Any) -> Dict[str, int]:
    return parse_count_map(value)


def parse_send_detail(value: Any) -> Dict[str, int]:
    return parse_count_map(value)


def parse_high_freq(value: Any) -> List[dict]:
    return [
        {"word": word, "count": count}
        for word, count in parse_count_map(value).items()
    ]


def parse_members(value: Any) -> List[str]:
    if not value:
        return []
    text = str(value).strip()
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            names = []
            for item in parsed:
                if isinstance(item, dict):
                    name = item.get("name") or item.get("group_nickname")
                else:
                    name = str(item)
                if name:
                    names.append(str(name).strip())
            return names
    except (json.JSONDecodeError, TypeError):
        pass
    return [item.strip() for item in re.split(r"[,，、;；\n]", text) if item.strip()]


def parse_json_object(value: Any) -> dict:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
    except (TypeError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def parse_msg_datetime(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d{10,13}", text):
        timestamp = int(text)
        if len(text) >= 13:
            timestamp = timestamp / 1000
        try:
            return datetime.fromtimestamp(timestamp)
        except (OSError, OverflowError, ValueError):
            return None
    normalized = text.replace("T", " ").replace("/", "-").replace("Z", "")
    if "." in normalized:
        normalized = normalized.split(".", 1)[0]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(normalized[:len(fmt.replace("%Y", "0000").replace("%m", "00").replace("%d", "00").replace("%H", "00").replace("%M", "00").replace("%S", "00"))], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(normalized).replace(tzinfo=None)
    except ValueError:
        return None


def chat_msgtime(row: dict) -> Optional[datetime]:
    raw = parse_json_object(row_get(row, "raw_json", "rawJson", default=None))
    value = _field_value_case_insensitive(raw, "msgtime", "msgTime", "msg_time", "createTime", "createtime")
    return parse_msg_datetime(value or row_get(row, "msgtime", "msg_time", default=None))


def chat_text(row: dict) -> str:
    raw = parse_json_object(row_get(row, "raw_json", "rawJson", default=None))
    value = first_nonempty(
        _field_value_case_insensitive(raw, "content", "text", "msg", "message", "msgContent"),
        row_get(row, "content", "text", "message", default=""),
    )
    # qx_chat.content is inconsistent in production: some rows store plain text,
    # while others store a JSON string such as {"content": "..."}.
    for _ in range(2):
        try:
            parsed = json.loads(value)
        except (json.JSONDecodeError, TypeError):
            break
        if isinstance(parsed, dict):
            value = first_nonempty(
                _field_value_case_insensitive(parsed, "content", "text", "msg", "message"),
                value,
            )
        elif isinstance(parsed, str):
            value = parsed
        else:
            break
    return value


def chat_sender(row: dict) -> str:
    raw = parse_json_object(row_get(row, "raw_json", "rawJson", default=None))
    userid = chat_sender_userid(row)
    roomid = str(row_get(row, "roomid", default="") or "")
    value = first_nonempty(
        _field_value_case_insensitive(raw, "sender_name", "truename", "fromName", "name"),
        row_get(row, "sender_name", "truename", default=""),
    )
    return safe_sender_name(value, userid, roomid)


def chat_sender_userid(row: dict) -> str:
    raw = parse_json_object(row_get(row, "raw_json", "rawJson", default=None))
    return first_nonempty(
        _field_value_case_insensitive(raw, "from_userid", "fromUserId", "from"),
        row_get(row, "from_userid", "from", default=""),
    )


def chat_sender_role(row: dict) -> str:
    raw = parse_json_object(row_get(row, "raw_json", "rawJson", default=None))
    userid = chat_sender_userid(row)
    members_value = row_get(row, "members", default=None) or raw.get("members")
    members = parse_members_payload(members_value)
    return infer_sender_role(
        from_userid=userid,
        roomid=str(row_get(row, "roomid", default="") or ""),
        sender_job=first_nonempty(raw.get("job"), row_get(row, "job", default="")),
        sender_position=first_nonempty(raw.get("position"), row_get(row, "position", default="")),
        member=members.get(userid),
    )


def chat_msgid(row: dict) -> str:
    raw = parse_json_object(row_get(row, "raw_json", "rawJson", default=None))
    return first_nonempty(
        _field_value_case_insensitive(raw, "msgid", "msgId", "id"),
        row_get(row, "msgid", "msgId", "id", default=""),
    )


def row_get(row: dict, *names: str, default: Any = "") -> Any:
    """Case-insensitive row lookup for tables whose column casing is inconsistent."""
    if not row:
        return default
    for name in names:
        if name in row and row[name] is not None:
            return row[name]
    lowered = {str(key).lower(): value for key, value in row.items()}
    for name in names:
        value = lowered.get(name.lower())
        if value is not None:
            return value
    return default


def first_nonempty(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def customer_lookup(customers: Dict[Any, dict], *keys: Any) -> dict:
    for key in keys:
        if key is None:
            continue
        for normalized in (key, str(key), str(key).strip()):
            if normalized in customers:
                return customers[normalized]
    return {}


def lims_base_data_url() -> str:
    from config.settings import settings

    base = settings.LIMS_API_URL.rstrip("/")
    path = settings.LIMS_BASE_DATA_PATH.strip("/")
    return f"{base}/{path}/"


def _chunked(values: List[str], size: int = 50) -> Iterable[List[str]]:
    for index in range(0, len(values), size):
        yield values[index:index + size]


def _load_persisted_lims_cache() -> None:
    """Load the last successful LIMS snapshot from the persistent logs volume."""
    global _lims_cache_loaded
    with _lims_cache_lock:
        if _lims_cache_loaded:
            return
        _lims_cache_loaded = True
        try:
            payload = json.loads(LIMS_CACHE_FILE.read_text(encoding="utf-8"))
            saved_at = float(payload.get("saved_at") or 0)
            records = payload.get("records") or {}
            if not isinstance(records, dict):
                return
            for raw_code, raw_rows in records.items():
                code = str(raw_code).upper()
                rows = raw_rows if isinstance(raw_rows, list) else []
                _lims_stale[code] = rows
                if saved_at:
                    _lims_cache[code] = (saved_at, copy.deepcopy(rows))
            logger.info(
                "lims.cache.loaded path={} codes={} age_seconds={:.0f}",
                LIMS_CACHE_FILE, len(records), max(0, time.time() - saved_at),
            )
        except FileNotFoundError:
            return
        except Exception as exc:
            logger.warning("lims.cache.load_failed path={} error={}", LIMS_CACHE_FILE, exc)


def _persist_lims_cache() -> None:
    """Atomically persist successful LIMS records for restart/outage fallback."""
    with _lims_cache_lock:
        snapshot = copy.deepcopy(_lims_stale)
    if not snapshot:
        return
    try:
        LIMS_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        temp_path = LIMS_CACHE_FILE.with_suffix(LIMS_CACHE_FILE.suffix + ".tmp")
        temp_path.write_text(
            json.dumps({"saved_at": time.time(), "records": snapshot}, ensure_ascii=False),
            encoding="utf-8",
        )
        os.replace(temp_path, LIMS_CACHE_FILE)
    except Exception as exc:
        logger.warning("lims.cache.persist_failed path={} error={}", LIMS_CACHE_FILE, exc)


def fetch_lims_base_data(project_codes: List[str]) -> Tuple[Dict[str, List[dict]], dict]:
    """Fetch LIMS project data using POST /unionLims/base_data/ with projectCode body."""
    requested_codes = sorted({str(code).upper() for code in project_codes if code})
    if not requested_codes:
        return {}, {"available": True, "requests": 0, "records": 0, "errors": 0}

    import httpx
    from config.settings import settings

    _load_persisted_lims_cache()
    now = time.time()
    records_by_code: Dict[str, List[dict]] = {}
    with _lims_cache_lock:
        for code in requested_codes:
            cached = _lims_cache.get(code)
            if cached and now - cached[0] < LIMS_CACHE_TTL_SECONDS:
                records_by_code[code] = copy.deepcopy(cached[1])

    missing_codes = [code for code in requested_codes if code not in records_by_code]
    stats = {
        "available": not missing_codes,
        "requests": 0,
        "records": sum(len(rows) for rows in records_by_code.values()),
        "errors": 0,
        "cache_hits": len(records_by_code),
        "stale_hits": 0,
    }
    if not missing_codes:
        return records_by_code, stats

    # Serialize refreshes inside a worker. A second request rechecks the cache
    # after the first one completes instead of sending the same LIMS batches.
    with _lims_refresh_lock:
        now = time.time()
        with _lims_cache_lock:
            for code in list(missing_codes):
                cached = _lims_cache.get(code)
                if cached and now - cached[0] < LIMS_CACHE_TTL_SECONDS:
                    records_by_code[code] = copy.deepcopy(cached[1])
                    stats["cache_hits"] += 1
        missing_codes = [code for code in requested_codes if code not in records_by_code]
        if not missing_codes:
            stats["available"] = True
            stats["records"] = sum(len(rows) for rows in records_by_code.values())
            return records_by_code, stats

        url = lims_base_data_url()
        configured_timeout = max(0.5, float(settings.LIMS_API_TIMEOUT))
        deadline = time.monotonic() + min(configured_timeout, LIMS_REQUEST_BUDGET_SECONDS)
        try:
            with httpx.Client() as client:
                for batch in _chunked(missing_codes):
                    remaining = deadline - time.monotonic()
                    if remaining <= 0.25:
                        stats["errors"] += 1
                        logger.warning(
                            "lims.base_data.budget_exhausted pending_codes={} budget_seconds={}",
                            len(missing_codes), LIMS_REQUEST_BUDGET_SECONDS,
                        )
                        break
                    stats["requests"] += 1
                    try:
                        response = client.post(
                            url,
                            json=[{"projectCode": code} for code in batch],
                            headers={"Accept": "application/json"},
                            timeout=max(0.5, min(configured_timeout, remaining)),
                        )
                        response.raise_for_status()
                        payload = response.json()
                        if payload.get("status") is False:
                            stats["errors"] += 1
                            logger.warning("lims.base_data.status_false message={}", payload.get("message"))
                            continue
                        batch_records: Dict[str, List[dict]] = {code: [] for code in batch}
                        for item in payload.get("data") or []:
                            code = str(item.get("projectCode") or "").upper()
                            if code in batch_records:
                                batch_records[code].append(item)
                        cached_at = time.time()
                        with _lims_cache_lock:
                            for code, rows in batch_records.items():
                                _lims_cache[code] = (cached_at, copy.deepcopy(rows))
                                _lims_stale[code] = copy.deepcopy(rows)
                        _persist_lims_cache()
                        records_by_code.update(batch_records)
                        stats["available"] = True
                    except Exception as exc:
                        stats["errors"] += 1
                        logger.warning("lims.base_data.batch_failed size={} error={}", len(batch), exc)
        except Exception as exc:
            stats["errors"] += 1
            logger.warning("lims.base_data.unavailable url={} error={}", url, exc)

        # A transient LIMS failure must not make an otherwise cached dashboard
        # unavailable. Use the last successful record for only the failed codes.
        with _lims_cache_lock:
            for code in missing_codes:
                if code not in records_by_code and code in _lims_stale:
                    records_by_code[code] = copy.deepcopy(_lims_stale[code])
                    stats["stale_hits"] += 1

    stats["records"] = sum(len(rows) for rows in records_by_code.values())
    stats["available"] = stats["available"] or bool(records_by_code)
    return records_by_code, stats


def normalize_lims_api_record(item: dict, code: str, mapping_snapshot: Optional[dict] = None) -> dict:
    members = parse_members(item.get("members"))
    owner = aftersaler_mapping.resolve_final_aftersaler(
        item.get("productBigSortThree"), item.get("orgName"), item.get("afterSaler"),
        mapping_snapshot,
    )
    return {
        "project_code": item.get("projectCode") or code,
        "customer_name": item.get("customerName") or "",
        "key_account": normalize_key_account(item.get("keyAccount"), item.get("customerName") or ""),
        "region": item.get("orgName") or "",
        "sales_person": item.get("saleName") or item.get("assignmentUser") or "",
        "product_name": item.get("productName") or "",
        "category_l1": item.get("productBigSortOne") or "未分类",
        "category_l2": item.get("productBigSortTwo") or "",
        "category_l3": item.get("productBigSortThree") or "",
        "work_unit": item.get("workUnit") or "",
        "raw_aftersaler": owner["raw_aftersaler"],
        "final_aftersaler": owner["final_aftersaler"],
        "aftersaler_source": owner["aftersaler_source"],
        "mapping_rule_id": owner["mapping_rule_id"],
        "mapping_conflict": owner["mapping_conflict"],
        "lims_members": members,
        "group_id": item.get("groupId") or "",
        "active_day": extract_lims_active_day(item),
        "start_time": item.get("startTime") or "",
        "end_time": item.get("endTime") or "",
        "analysis_simple_remark": str(item.get("analysisSimpleRemark") or "").strip(),
        "dimension_source": "lims_base_data_api",
    }


def _dimensions_from_lims_api(
    group_codes: Dict[str, List[str]],
    records_by_code: Dict[str, List[dict]],
    stats: dict,
    mapping_snapshot: Optional[dict] = None,
) -> Tuple[Dict[str, dict], dict]:
    dimensions: Dict[str, dict] = {}
    aftersaler_group_count = product_project_count = matched_product_count = 0
    groups_with_lims_link = groups_with_region = groups_with_key_account = 0
    groups_with_raw_aftersaler = groups_with_lims_members = 0
    mapping_matches = mapping_fallbacks = mapping_conflicts = 0

    for group_name, codes_for_group in group_codes.items():
        projects = []
        regions = set()
        aftersalers = set()
        raw_aftersalers = set()
        has_lims_link = has_region = has_key_account = False
        has_raw_aftersaler = has_lims_members = False

        for code in codes_for_group:
            for raw in records_by_code.get(code.upper(), []):
                item = normalize_lims_api_record(raw, code, mapping_snapshot)
                has_lims_link = True
                if item["product_name"]:
                    product_project_count += 1
                if item["category_l1"] != "未分类":
                    matched_product_count += 1
                if item["region"]:
                    regions.add(item["region"])
                    has_region = True
                if item["key_account"]:
                    has_key_account = True
                raw_after = item["raw_aftersaler"]
                if raw_after:
                    raw_aftersalers.add(raw_after)
                    aftersalers.add(item["final_aftersaler"] or raw_after)
                    has_raw_aftersaler = True
                    if item["mapping_conflict"]:
                        mapping_conflicts += 1
                    elif item["aftersaler_source"] == "mapping":
                        mapping_matches += 1
                    else:
                        mapping_fallbacks += 1
                lims_members = set(item["lims_members"])
                if lims_members:
                    has_lims_members = True
                projects.append(item)

        if aftersalers:
            aftersaler_group_count += 1
        groups_with_lims_link += 1 if has_lims_link else 0
        groups_with_region += 1 if has_region else 0
        groups_with_key_account += 1 if has_key_account else 0
        groups_with_raw_aftersaler += 1 if has_raw_aftersaler else 0
        groups_with_lims_members += 1 if has_lims_members else 0
        dimensions[group_name] = {
            "codes": codes_for_group,
            "projects": projects,
            "regions": sorted(regions),
            "aftersalers": sorted(aftersalers),
            "tentative_aftersalers": [],
            "raw_aftersalers": sorted(raw_aftersalers),
            "chat_members": [],
        }

    requested_codes = sorted({code for codes in group_codes.values() for code in codes})
    matched_codes = sum(1 for code in requested_codes if records_by_code.get(code.upper()))
    groups_without_project_code = sum(1 for codes in group_codes.values() if not codes)
    matched_code_set = {code.upper() for code in requested_codes if records_by_code.get(code.upper())}
    quality = {
        "project_codes": len(requested_codes),
        "matched_project_codes": matched_codes,
        "unmatched_project_codes": len(requested_codes) - matched_codes,
        "groups_with_project_code": len(group_codes) - groups_without_project_code,
        "groups_without_project_code": groups_without_project_code,
        "groups_without_lims_link": sum(
            1 for codes in group_codes.values()
            if codes and not any(code.upper() in matched_code_set for code in codes)
        ),
        "product_projects": product_project_count,
        "matched_products": matched_product_count,
        "groups_with_aftersaler": aftersaler_group_count,
        "groups_with_confirmed_aftersaler": aftersaler_group_count,
        "groups_with_lims_link": groups_with_lims_link,
        "groups_with_region": groups_with_region,
        "groups_with_key_account": groups_with_key_account,
        "groups_with_raw_aftersaler": groups_with_raw_aftersaler,
        "groups_with_lims_members": groups_with_lims_members,
        "lims_source": "base_data_api",
        "lims_api_requests": stats.get("requests", 0),
        "lims_api_records": stats.get("records", 0),
        "lims_api_errors": stats.get("errors", 0),
        "lims_api_cache_hits": stats.get("cache_hits", 0),
        "lims_api_stale_hits": stats.get("stale_hits", 0),
        "mapping_available": bool(mapping_snapshot and mapping_snapshot.get("available")),
        "mapping_version_id": (mapping_snapshot or {}).get("version_id"),
        "mapping_effective_month": (mapping_snapshot or {}).get("effective_month"),
        "mapping_revision": (mapping_snapshot or {}).get("revision", 0),
        "mapping_reason": (mapping_snapshot or {}).get("reason", "not_loaded"),
        "mapping_matched_records": mapping_matches,
        "mapping_fallback_records": mapping_fallbacks,
        "mapping_conflict_records": mapping_conflicts,
    }
    return dimensions, quality


def _quality_from_dimensions(dimensions: Dict[str, dict], base_quality: dict) -> dict:
    project_codes = {
        str(code).upper()
        for dim in dimensions.values()
        for code in dim.get("codes", [])
        if code
    }
    matched_codes = {
        str(project.get("project_code") or "").upper()
        for dim in dimensions.values()
        for project in dim.get("projects", [])
        if project.get("project_code")
    }
    return {
        **base_quality,
        "project_codes": len(project_codes),
        "matched_project_codes": len(project_codes & matched_codes),
        "unmatched_project_codes": len(project_codes - matched_codes),
        "groups_with_project_code": sum(1 for dim in dimensions.values() if dim.get("codes")),
        "groups_without_project_code": sum(1 for dim in dimensions.values() if not dim.get("codes")),
        "groups_without_lims_link": sum(1 for dim in dimensions.values() if dim.get("codes") and not dim.get("projects")),
        "product_projects": sum(
            1 for dim in dimensions.values() for project in dim.get("projects", [])
            if project.get("product_name")
        ),
        "matched_products": sum(
            1 for dim in dimensions.values() for project in dim.get("projects", [])
            if project.get("category_l2") in ALLOWED_PRODUCT_L2
        ),
        "groups_with_aftersaler": sum(1 for dim in dimensions.values() if dim.get("aftersalers")),
        "groups_with_confirmed_aftersaler": sum(1 for dim in dimensions.values() if dim.get("aftersalers")),
        "groups_with_lims_link": sum(1 for dim in dimensions.values() if dim.get("projects")),
        "groups_with_region": sum(1 for dim in dimensions.values() if dim.get("regions")),
        "groups_with_key_account": sum(
            1 for dim in dimensions.values()
            if any(project.get("key_account") for project in dim.get("projects", []))
        ),
        "groups_with_raw_aftersaler": sum(1 for dim in dimensions.values() if dim.get("raw_aftersalers")),
        "groups_with_lims_members": sum(1 for dim in dimensions.values() if dim.get("chat_members")),
        "mapping_matched_records": sum(
            1 for dim in dimensions.values() for project in dim.get("projects", [])
            if project.get("aftersaler_source") == "mapping"
        ),
        "mapping_fallback_records": sum(
            1 for dim in dimensions.values() for project in dim.get("projects", [])
            if project.get("raw_aftersaler") and project.get("aftersaler_source") != "mapping"
            and not project.get("mapping_conflict")
        ),
        "mapping_conflict_records": sum(
            1 for dim in dimensions.values() for project in dim.get("projects", [])
            if project.get("mapping_conflict")
        ),
    }


def _emotion_total(mapping: Dict[str, int], labels: Iterable[str]) -> int:
    total = 0
    for key, count in mapping.items():
        if any(label in key for label in labels):
            total += count
    return total


def normalize_key_account(value: Any, customer_name: str = "", fallback: str = "") -> str:
    raw = str(value or "").strip()
    lowered = raw.lower()
    if lowered in ("", "0", "false", "no", "null", "none", "否", "无"):
        return str(fallback or "").strip()
    if lowered in ("1", "true", "yes", "是", "有"):
        return str(fallback or customer_name or "").strip()
    return raw


def resolve_period(
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    today: Optional[date] = None,
) -> Tuple[date, date, str]:
    current = today or date.today()
    period = (period or "month").lower()
    if period == "custom":
        if not start_date or not end_date:
            raise ValueError("custom period requires start_date and end_date")
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    elif period in ("today", "daily"):
        start = end = current
        period = "today"
    elif period in ("week", "weekly"):
        start = current - timedelta(days=current.weekday())
        end = current
        period = "week"
    elif period in ("quarter", "quarterly"):
        month = ((current.month - 1) // 3) * 3 + 1
        start = date(current.year, month, 1)
        end = current
        period = "quarter"
    elif period in ("year", "yearly"):
        start = date(current.year, 1, 1)
        end = current
        period = "year"
    else:
        start = date(current.year, current.month, 1)
        end = current
        period = "month"
    if start > end:
        raise ValueError("start_date must not be later than end_date")
    if (end - start).days > 730:
        raise ValueError("date range cannot exceed 731 days")
    return start, end, period


def _latest_rows(conn, start: date, end: date) -> Tuple[List[dict], int]:
    params = (start.isoformat(), (end + timedelta(days=1)).isoformat())
    sql = """
        SELECT a.*
        FROM qx_analysis_result a
        JOIN (
            SELECT groupName, DATE(CREATEDTIME) analysis_date, MAX(id) latest_id
            FROM qx_analysis_result
            WHERE CREATEDTIME >= %s AND CREATEDTIME < %s
            GROUP BY groupName, DATE(CREATEDTIME)
        ) latest ON latest.latest_id = a.id
        ORDER BY a.CREATEDTIME
    """
    rows = [
        row for row in _query(conn, "analysis.latest_daily", sql, params)
        if is_focus_group_name(row.get("groupName", ""))
    ]
    raw_count_rows = _query(
        conn,
        "analysis.raw_count",
        """SELECT groupName, COUNT(*) AS count
           FROM qx_analysis_result
           WHERE CREATEDTIME >= %s AND CREATEDTIME < %s
           GROUP BY groupName""",
        params,
    )
    raw_count = sum(
        int(row.get("count") or 1)
        for row in raw_count_rows
        if is_focus_group_name(row.get("groupName", ""))
    )
    return rows, raw_count


def _latest_unanswered_rows(
    conn,
    start: date,
    end: date,
    group_names: Optional[List[str]] = None,
) -> List[dict]:
    """Load the latest nightly unanswered-analysis row for each group and day."""
    params: List[Any] = [start.isoformat(), (end + timedelta(days=1)).isoformat()]
    group_filter = ""
    if group_names is not None:
        if not group_names:
            return []
        placeholders = ",".join(["%s"] * len(group_names))
        group_filter = f" AND groupName IN ({placeholders})"
        params.extend(group_names)
    sql = f"""
        SELECT a.*
        FROM {UNANSWERED_RESULT_TABLE} a
        JOIN (
            SELECT groupName, DATE(CREATEDTIME) analysis_date, MAX(id) latest_id
            FROM {UNANSWERED_RESULT_TABLE}
            WHERE CREATEDTIME >= %s AND CREATEDTIME < %s{group_filter}
            GROUP BY groupName, DATE(CREATEDTIME)
        ) latest ON latest.latest_id = a.id
        ORDER BY a.CREATEDTIME
    """
    return [
        row for row in _query(conn, "unanswered.latest_nightly", sql, params)
        if is_focus_group_name(row.get("groupName", ""))
    ]


def _raw_unanswered_rows(
    conn,
    start: date,
    end: date,
    group_names: List[str],
) -> List[dict]:
    if not group_names:
        return []
    return _query_by_chunks(
        conn,
        "unanswered.scoped_raw",
        f"""SELECT * FROM {UNANSWERED_RESULT_TABLE}
           WHERE CREATEDTIME >= %s AND CREATEDTIME < %s
             AND groupName IN ({{placeholders}})
           ORDER BY CREATEDTIME""",
        group_names,
        prefix_params=[start.isoformat(), (end + timedelta(days=1)).isoformat()],
    )


def _raw_analysis_count(
    conn,
    start: date,
    end: date,
    group_names: List[str],
) -> int:
    """Count raw analysis records inside the already-filtered dashboard group scope."""
    if not group_names:
        return 0
    rows = _query_by_chunks(
        conn,
        "analysis.scoped_raw_count",
        """SELECT COUNT(*) AS total FROM qx_analysis_result
           WHERE CREATEDTIME >= %s AND CREATEDTIME < %s
             AND groupName IN ({placeholders})""",
        group_names,
        prefix_params=[start.isoformat(), (end + timedelta(days=1)).isoformat()],
    )
    return sum(int(row.get("total") or 0) for row in rows)


def _load_dimensions(
    conn, rows: List[dict], mapping_snapshot: Optional[dict] = None,
) -> Tuple[Dict[str, dict], dict]:
    group_codes = {row["groupName"]: extract_project_codes(row.get("groupName", "")) for row in rows}
    codes = sorted({code for values in group_codes.values() for code in values})
    if not codes:
        groups_without_project_code = len(group_codes)
        return {name: {"codes": [], "projects": [], "regions": [], "aftersalers": []}
                for name in group_codes}, {
                    "project_codes": 0, "matched_project_codes": 0,
                    "unmatched_project_codes": 0,
                    "groups_with_project_code": 0,
                    "groups_without_project_code": groups_without_project_code,
                    "groups_without_lims_link": 0,
                    "product_projects": 0, "matched_products": 0,
                    "groups_with_aftersaler": 0,
                    "groups_with_confirmed_aftersaler": 0,
                    "groups_with_lims_link": 0,
                    "groups_with_region": 0,
                    "groups_with_key_account": 0,
                    "groups_with_raw_aftersaler": 0,
                    "groups_with_lims_members": 0,
                    "lims_source": "none",
                    "mapping_available": bool(mapping_snapshot and mapping_snapshot.get("available")),
                    "mapping_version_id": (mapping_snapshot or {}).get("version_id"),
                    "mapping_effective_month": (mapping_snapshot or {}).get("effective_month"),
                    "mapping_revision": (mapping_snapshot or {}).get("revision", 0),
                    "mapping_reason": (mapping_snapshot or {}).get("reason", "not_loaded"),
                }

    lims_records_by_code, lims_stats = fetch_lims_base_data(codes)
    if lims_stats.get("available") and lims_records_by_code:
        return _dimensions_from_lims_api(
            group_codes, lims_records_by_code, lims_stats, mapping_snapshot,
        )

    logger.warning(
        "dashboard.dimension.lims_unavailable_no_business_fallback requested_codes={} lims_available={} lims_records={} lims_errors={}",
        len(codes),
        lims_stats.get("available"),
        lims_stats.get("records"),
        lims_stats.get("errors"),
    )
    dimensions: Dict[str, dict] = {
        group_name: {
            "codes": codes_for_group,
            "projects": [],
            "regions": [],
            "aftersalers": [],
            "tentative_aftersalers": [],
            "raw_aftersalers": [],
            "chat_members": parse_members(row.get("member")) if (row := next((r for r in rows if r.get("groupName") == group_name), None)) else [],
            "dimension_source": "lims_unavailable",
        }
        for group_name, codes_for_group in group_codes.items()
    }
    quality = {
        "project_codes": len(codes),
        "matched_project_codes": 0,
        "unmatched_project_codes": len(codes),
        "groups_with_project_code": sum(1 for codes_for_group in group_codes.values() if codes_for_group),
        "groups_without_project_code": sum(1 for codes_for_group in group_codes.values() if not codes_for_group),
        "groups_without_lims_link": sum(1 for codes_for_group in group_codes.values() if codes_for_group),
        "product_projects": 0,
        "matched_products": 0,
        "groups_with_aftersaler": 0,
        "groups_with_confirmed_aftersaler": 0,
        "groups_with_lims_link": 0,
        "groups_with_region": 0,
        "groups_with_key_account": 0,
        "groups_with_raw_aftersaler": 0,
        "groups_with_lims_members": 0,
        "lims_source": "base_data_api_unavailable",
        "lims_api_requests": lims_stats.get("requests", 0),
        "lims_api_records": lims_stats.get("records", 0),
        "lims_api_errors": lims_stats.get("errors", 0),
        "lims_api_cache_hits": lims_stats.get("cache_hits", 0),
        "lims_api_stale_hits": lims_stats.get("stale_hits", 0),
        "mapping_available": bool(mapping_snapshot and mapping_snapshot.get("available")),
        "mapping_version_id": (mapping_snapshot or {}).get("version_id"),
        "mapping_effective_month": (mapping_snapshot or {}).get("effective_month"),
        "mapping_revision": (mapping_snapshot or {}).get("revision", 0),
        "mapping_reason": (mapping_snapshot or {}).get("reason", "not_loaded"),
        "mapping_matched_records": 0,
        "mapping_fallback_records": 0,
        "mapping_conflict_records": 0,
    }
    return dimensions, quality


def _active_durations_from_lims(dimensions: Dict[str, dict]) -> Dict[str, float]:
    durations: Dict[str, float] = {}
    for group_name, dim in dimensions.items():
        values = [
            parse_active_day(project.get("active_day"))
            for project in dim.get("projects", [])
            if parse_active_day(project.get("active_day")) is not None
        ]
        if values:
            durations[group_name] = max(values)
    return durations


def _group_aggregates(
    rows: List[dict],
    dimensions: Dict[str, dict],
    unanswered_rows: Optional[List[dict]] = None,
) -> Dict[str, dict]:
    missed_dates_by_group: Dict[str, set] = defaultdict(set)
    missed_source_rows = rows if unanswered_rows is None else unanswered_rows
    for row in missed_source_rows:
        if str(row.get("isMissedMessage")) == "1" and row.get("groupName"):
            missed_dates_by_group[str(row["groupName"])].add(str(row.get("CREATEDTIME"))[:10])
    groups: Dict[str, dict] = {}
    for row in rows:
        name = row["groupName"]
        item = groups.setdefault(name, {
            "group_name": name, "messages": 0, "dates": set(), "missed_days": 0,
            "customer_good": 0, "customer_bad": 0, "employee_positive": 0,
            "employee_negative": 0, "high_freq": Counter(), "rows": [],
            "dimension": dimensions.get(name, {}),
        })
        item["messages"] += int(row.get("messageToDayCount") or 0)
        item["dates"].add(str(row.get("CREATEDTIME"))[:10])
        customer = parse_emotion_field(row.get("customerEmotionAnalysis"))
        employee = parse_emotion_field(row.get("saleEmotionAnalysis"))
        item["customer_good"] += _emotion_total(customer, ("好评", "正向", "满意"))
        item["customer_bad"] += _emotion_total(customer, ("差评", "负向", "不满"))
        item["employee_positive"] += _emotion_total(employee, ("积极", "正向"))
        item["employee_negative"] += _emotion_total(employee, ("恶劣", "负向", "消极"))
        for word in parse_high_freq(row.get("highFrequencyWords")):
            item["high_freq"][word["word"]] += word["count"]
        item["rows"].append(row)
    for group_name, item in groups.items():
        item["missed_days"] = len(missed_dates_by_group.get(group_name, set()))
    return groups


def _build_top_message_groups(
    groups: Dict[str, dict], total_messages: int, limit: int = 5,
) -> dict:
    """Rank scoped group chats by period message volume without requiring LIMS data."""
    items: List[dict] = []
    for group_name, group in groups.items():
        dimension = group.get("dimension") or {}
        projects = dimension.get("projects") or []
        project_codes = {
            str(value).strip().upper()
            for value in dimension.get("codes", [])
            if str(value).strip()
        }
        project_codes.update({
            str(project.get("project_code") or "").strip().upper()
            for project in projects
            if str(project.get("project_code") or "").strip()
        })
        if not project_codes:
            project_codes.update(extract_project_codes(group_name))

        items.append({
            "group_name": group_name,
            "message_count": max(0, int(group.get("messages") or 0)),
            "percentage_of_all": 0.0,
            "active_days": len(group.get("dates") or set()),
            "high_frequency_top5": [
                {"word": word, "count": count}
                for word, count in (group.get("high_freq") or Counter()).most_common(5)
            ],
            "project_codes": sorted(project_codes),
            "customer_units": sorted({
                str(project.get("work_unit") or "").strip()
                for project in projects
                if str(project.get("work_unit") or "").strip()
            }),
            "customer_names": sorted({
                str(project.get("customer_name") or "").strip()
                for project in projects
                if str(project.get("customer_name") or "").strip()
            }),
            "product_categories": sorted({
                str(project.get("category_l2") or "").strip()
                for project in projects
                if str(project.get("category_l2") or "").strip()
            }),
            "aftersalers": sorted({
                str(value).strip()
                for value in dimension.get("aftersalers", [])
                if str(value).strip()
            }),
        })

    ranked = sorted(
        items,
        key=lambda item: (
            -item["message_count"], -item["active_days"], item["group_name"]
        ),
    )[:max(0, limit)]
    for rank, item in enumerate(ranked, 1):
        item["rank"] = rank
        item["percentage_of_all"] = _ratio(item["message_count"], total_messages)

    top5_messages = sum(item["message_count"] for item in ranked)
    return {
        "limit": max(0, limit),
        "actual_count": len(ranked),
        "total_groups": len(groups),
        "total_messages": total_messages,
        "top5_messages": top5_messages,
        "coverage_percentage": _ratio(top5_messages, total_messages),
        "items": ranked,
    }


def extract_project_code_year(project_code: Any) -> Optional[int]:
    """Extract the canonical project year from LC-P/LC-SP project codes."""
    match = re.match(
        r"^LC-(?:SP|P)((?:19|20)\d{2})",
        str(project_code or "").strip().upper(),
    )
    return int(match.group(1)) if match else None


def parse_lims_start_date(value: Any) -> Optional[date]:
    """Normalize a LIMS startTime value for deterministic year fallback."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    match = re.search(
        r"((?:19|20)\d{2})\D{0,2}(\d{1,2})?\D{0,2}(\d{1,2})?",
        text,
    )
    if not match:
        return None
    try:
        return date(
            int(match.group(1)), int(match.group(2) or 1), int(match.group(3) or 1),
        )
    except ValueError:
        return None


def _build_project_year_distribution(
    groups: Dict[str, dict], current_year: Optional[int] = None,
) -> dict:
    """Aggregate unique scoped projects by canonical project year."""
    as_of_year = int(current_year or date.today().year)
    projects_by_code: Dict[str, dict] = {}

    for group_name, group in groups.items():
        dimension = group.get("dimension") or {}
        records = dimension.get("projects") or []
        records_by_code: Dict[str, List[dict]] = defaultdict(list)
        for project in records:
            code = str(project.get("project_code") or "").strip().upper()
            if code:
                records_by_code[code].append(project)

        project_codes = {
            str(code).strip().upper()
            for code in dimension.get("codes") or []
            if str(code).strip()
        }
        project_codes.update(records_by_code)
        if not project_codes:
            project_codes.update(extract_project_codes(group_name))

        for code in project_codes:
            aggregate = projects_by_code.setdefault(code, {
                "project_code": code,
                "project_names": set(),
                "customer_names": set(),
                "work_units": set(),
                "product_categories": set(),
                "regions": set(),
                "aftersalers": set(),
                "group_names": set(),
                "start_dates": set(),
            })
            aggregate["group_names"].add(group_name)
            code_records = records_by_code.get(code, [])
            for project in code_records:
                for source_key, target_key in (
                    ("product_name", "project_names"),
                    ("customer_name", "customer_names"),
                    ("work_unit", "work_units"),
                    ("region", "regions"),
                ):
                    value = str(project.get(source_key) or "").strip()
                    if value:
                        aggregate[target_key].add(value)
                category = " / ".join(filter(None, (
                    str(project.get("category_l2") or "").strip(),
                    str(project.get("category_l3") or "").strip(),
                )))
                if category:
                    aggregate["product_categories"].add(category)
                aftersaler_name = str(
                    project.get("final_aftersaler")
                    or project.get("raw_aftersaler") or ""
                ).strip()
                if aftersaler_name:
                    aggregate["aftersalers"].add(aftersaler_name)
                start_date = parse_lims_start_date(project.get("start_time"))
                if start_date:
                    aggregate["start_dates"].add(start_date)
            if not code_records:
                aggregate["aftersalers"].update({
                    str(value).strip()
                    for value in dimension.get("aftersalers") or []
                    if str(value).strip()
                })

    projects: List[dict] = []
    for code, aggregate in projects_by_code.items():
        start_date = min(aggregate["start_dates"], default=None)
        year = extract_project_code_year(code)
        if year is not None:
            year_source = "project_code"
        elif start_date is not None:
            year = start_date.year
            year_source = "lims_start_time"
        else:
            year_source = "unknown"
        projects.append({
            "project_code": code,
            "year": year,
            "year_source": year_source,
            "start_time": start_date.isoformat() if start_date else "",
            "project_names": sorted(aggregate["project_names"]),
            "customer_names": sorted(aggregate["customer_names"]),
            "work_units": sorted(aggregate["work_units"]),
            "product_categories": sorted(aggregate["product_categories"]),
            "regions": sorted(aggregate["regions"]),
            "aftersalers": sorted(aggregate["aftersalers"]),
            "group_count": len(aggregate["group_names"]),
            "group_names": sorted(aggregate["group_names"]),
        })

    projects.sort(key=lambda item: item["project_code"])
    grouped: Dict[Optional[int], List[dict]] = defaultdict(list)
    for project in projects:
        grouped[project["year"]].append(project)

    total_projects = len(projects)
    items = []
    ordered_years = sorted(
        (year for year in grouped if year is not None), reverse=True,
    )
    if None in grouped:
        ordered_years.append(None)
    for year in ordered_years:
        year_projects = grouped[year]
        if year == as_of_year:
            label = f"今年（{year}）"
        elif year == as_of_year - 1:
            label = f"去年（{year}）"
        elif year is None:
            label = "未识别"
        else:
            label = f"{year}年"
        group_names = {
            group_name
            for project in year_projects
            for group_name in project["group_names"]
        }
        items.append({
            "year": year,
            "label": label,
            "project_count": len(year_projects),
            "percentage": _ratio(len(year_projects), total_projects),
            "group_count": len(group_names),
            "projects": year_projects,
        })

    recognized_projects = sum(1 for project in projects if project["year"] is not None)
    return {
        "current_year": as_of_year,
        "total_projects": total_projects,
        "recognized_projects": recognized_projects,
        "unknown_projects": total_projects - recognized_projects,
        "coverage_percentage": _ratio(recognized_projects, total_projects),
        "current_year_projects": sum(1 for project in projects if project["year"] == as_of_year),
        "previous_year_projects": sum(1 for project in projects if project["year"] == as_of_year - 1),
        "older_projects": sum(
            1 for project in projects
            if project["year"] is not None and project["year"] < as_of_year - 1
        ),
        "source_priority": ["project_code", "lims_start_time"],
        "items": items,
    }


def _ratio(value: int, total: int) -> float:
    return round(value / total * 100, 1) if total else 0.0


def _counter_items(counter: Counter, label: str = "name") -> List[dict]:
    total = sum(counter.values())
    return [
        {label: name, "count": count, "percentage": _ratio(count, total)}
        for name, count in counter.most_common()
    ]


def _dimension_matches(
    dimension: dict,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> bool:
    return bool(_matching_projects(dimension, region, category, key_account, aftersaler))


def _dimension_matches_or_degrades(
    dimension: dict,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> bool:
    if _dimension_matches(dimension, region, aftersaler, category, key_account):
        return True
    # If LIMS is temporarily unavailable, keep project-code groups visible when
    # no LIMS-dependent filter is requested. The response quality metadata still
    # marks these dimensions as unavailable, but the dashboard/detail does not
    # collapse to an empty result.
    return bool(
        dimension.get("dimension_source") == "lims_unavailable"
        and dimension.get("codes")
        and not any((region, aftersaler, category, key_account))
    )


def _matching_projects(
    dimension: dict,
    region: str = "",
    category: str = "",
    key_account: str = "",
    aftersaler: str = "",
) -> List[dict]:
    """Return projects in the dashboard product scope that match one record-level filter."""
    if category and category not in ALLOWED_PRODUCT_L2:
        return []
    return [
        project
        for project in dimension.get("projects", [])
        if project.get("category_l2") in ALLOWED_PRODUCT_L2
        and (not category or project.get("category_l2") == category)
        and (not region or project.get("region") == region)
        and (not key_account or project.get("key_account") == key_account)
        and (
            not aftersaler
            or (project.get("final_aftersaler") or project.get("raw_aftersaler")) == aftersaler
        )
    ]


def _scope_dimension(
    dimension: dict,
    region: str = "",
    category: str = "",
    key_account: str = "",
    aftersaler: str = "",
) -> dict:
    """Copy a dimension and remove projects outside the shared dashboard scope."""
    scoped = copy.deepcopy(dimension)
    projects = _matching_projects(dimension, region, category, key_account, aftersaler)
    if (
        not projects
        and dimension.get("dimension_source") == "lims_unavailable"
        and not any((region, category, key_account, aftersaler))
    ):
        return scoped
    scoped["projects"] = projects
    scoped["codes"] = sorted({
        str(project.get("project_code") or "").upper()
        for project in projects
        if project.get("project_code")
    })
    scoped["regions"] = sorted({
        project.get("region") for project in projects if project.get("region")
    })
    scoped["aftersalers"] = sorted({
        project.get("final_aftersaler") or project.get("raw_aftersaler")
        for project in projects
        if project.get("final_aftersaler") or project.get("raw_aftersaler")
    })
    scoped["raw_aftersalers"] = sorted({
        project.get("raw_aftersaler") for project in projects if project.get("raw_aftersaler")
    })
    return scoped


def _dashboard_filter_options(dimensions: Dict[str, dict]) -> dict:
    """Build stable filter choices before applying the current business filter."""
    base_dimensions = {
        group_name: _scope_dimension(dimension)
        for group_name, dimension in dimensions.items()
        if _dimension_matches_or_degrades(dimension)
    }
    return {
        "regions": sorted({
            value for dimension in base_dimensions.values()
            for value in dimension.get("regions", [])
        }),
        "aftersalers": sorted({
            value for dimension in base_dimensions.values()
            for value in dimension.get("aftersalers", [])
        }),
        "categories": sorted({
            project.get("category_l2")
            for dimension in base_dimensions.values()
            for project in dimension.get("projects", [])
            if project.get("category_l2") in ALLOWED_PRODUCT_L2
        }),
        "key_accounts": sorted({
            project.get("key_account")
            for dimension in base_dimensions.values()
            for project in dimension.get("projects", [])
            if project.get("key_account")
        }),
    }


def _normalize_person_identity(value: Any) -> str:
    """Normalize a display name for matching a LIMS owner to a chat sender."""
    import unicodedata

    return re.sub(r"[\s\u3000]+", "", unicodedata.normalize("NFKC", str(value or "")).casefold())


def _sender_matches_aftersaler(message: dict, aftersaler: str) -> bool:
    """Only attribute a message when it is clearly authored by that employee."""
    sender_role = str(message.get("sender_role") or chat_sender_role(message) or "")
    if sender_role not in {"员工", "售后", "销售"}:
        return False
    sender = _normalize_person_identity(message.get("sender_name") or chat_sender(message))
    owner = _normalize_person_identity(aftersaler)
    if not sender or not owner:
        return False
    if sender == owner:
        return True
    # Some enterprise address books append a phone number or a role in brackets
    # to the real name. Keep this narrow so similarly named people never share
    # one message.
    suffix = sender[len(owner):] if sender.startswith(owner) else ""
    return bool(
        suffix
        and (
            re.fullmatch(r"1\d{10}", suffix)
            or re.fullmatch(r"[（(][^()（）]{1,20}[)）]", suffix)
        )
    )


def _aftersaler_message_rows(aftersalers: Iterable[str], messages: Iterable[dict]) -> Dict[str, List[dict]]:
    """Apply the group/personal hybrid attribution required by the dashboard.

    A group with one final owner keeps the historical group-volume definition.
    A group with multiple final owners attributes only each owner's own messages.
    """
    people = sorted({str(person or "").strip() for person in aftersalers if str(person or "").strip()})
    rows = list(messages or [])
    if not people:
        return {}
    if len(people) == 1:
        return {people[0]: rows}
    attributed = {person: [] for person in people}
    for message in rows:
        matches = [person for person in people if _sender_matches_aftersaler(message, person)]
        if len(matches) == 1:
            attributed[matches[0]].append(message)
    return attributed


def _aftersaler_message_counts(
    aftersalers: Iterable[str], messages: Iterable[dict], group_message_count: int,
) -> Dict[str, int]:
    """Count the same qx_chat rows that are exposed by the drilldown.

    ``group_message_count`` is retained for compatibility with older callers but
    is deliberately ignored.  Using the analysis aggregate for single-owner
    groups and qx_chat for multi-owner groups made the dashboard impossible to
    reconcile with its own detail rows.
    """
    attributed = _aftersaler_message_rows(aftersalers, messages)
    return {person: len(rows) for person, rows in attributed.items()}



def _time_period_breakdown(groups, dimensions, chat_rows_by_group: Optional[Dict[str, List[dict]]] = None):
    """按最终售后统计群消息；多售后群仅统计各售后本人的发言。"""
    aftersaler_stats = {}
    chat_rows_by_group = chat_rows_by_group or {}
    for group_name, item in groups.items():
        dim = dimensions.get(group_name, item.get("dimension", {}))
        aftersalers = dim.get("aftersalers", []) or ["未关联售后"]
        messages = chat_rows_by_group.get(group_name, [])
        if not messages:
            continue
        attributed_rows = _aftersaler_message_rows(aftersalers, messages)
        for person, person_messages in attributed_rows.items():
            buckets = {"morning": 0, "afternoon": 0, "after_hours": 0, "weekend": 0}
            for message in person_messages:
                msg_time = chat_msgtime(message)
                if not msg_time:
                    continue
                minutes = msg_time.hour * 60 + msg_time.minute
                if msg_time.weekday() >= 5:
                    buckets["weekend"] += 1
                elif 8 * 60 + 30 <= minutes < 12 * 60:
                    buckets["morning"] += 1
                elif 12 * 60 <= minutes < 17 * 60 + 30:
                    buckets["afternoon"] += 1
                else:
                    buckets["after_hours"] += 1
            total_msgs = sum(buckets.values())
            if total_msgs == 0:
                continue
            s = aftersaler_stats.setdefault(person, {
                "aftersaler": person, "groups": set(),
                "morning": 0, "afternoon": 0, "after_hours": 0, "weekend": 0,
                "total": 0,
            })
            s["groups"].add(group_name)
            s["morning"] += buckets["morning"]
            s["afternoon"] += buckets["afternoon"]
            s["after_hours"] += buckets["after_hours"]
            s["weekend"] += buckets["weekend"]
            s["total"] += total_msgs
    items = []
    for person, s in aftersaler_stats.items():
        total = s["total"]
        items.append({
            "aftersaler": person,
            "group_count": len(s["groups"]),
            "morning": {"count": s["morning"], "percentage": round(s["morning"] / total * 100, 1) if total else 0},
            "afternoon": {"count": s["afternoon"], "percentage": round(s["afternoon"] / total * 100, 1) if total else 0},
            "after_hours": {"count": s["after_hours"], "percentage": round(s["after_hours"] / total * 100, 1) if total else 0},
            "weekend": {"count": s["weekend"], "percentage": round(s["weekend"] / total * 100, 1) if total else 0},
            "total": total,
        })
    items.sort(key=lambda x: -x["total"])
    all_groups = set()
    for s in aftersaler_stats.values():
        all_groups.update(s["groups"])
    return {
        "items": items,
        "total_aftersalers": len(items),
        "total_groups": len(all_groups),
        "workday_morning": "08:30-12:00",
        "workday_afternoon": "12:00-17:30",
        "after_hours": "工作日 17:30-次日08:30，不包含周末",
        "weekend": "周六、周日全天单独统计",
    }


def _filter_chat_rows_by_period(chat_rows: List[dict], start: date, end: date) -> List[dict]:
    filtered = []
    for row in chat_rows:
        msg_time = chat_msgtime(row)
        if msg_time and start <= msg_time.date() <= end:
            filtered.append(row)
    return filtered


def _chat_rows_by_group_name(group_rows: List[dict], chat_rows: List[dict]) -> Dict[str, List[dict]]:
    room_to_group = {
        str(row.get("chat_id") or "").strip(): str(row.get("name") or "").strip()
        for row in group_rows
        if str(row.get("chat_id") or "").strip() and str(row.get("name") or "").strip()
    }
    result: Dict[str, List[dict]] = defaultdict(list)
    for row in chat_rows:
        group_name = room_to_group.get(str(row.get("roomid") or "").strip())
        if group_name:
            result[group_name].append(row)
    return result


def _build_overview(
    start: date,
    end: date,
    period: str,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
    snapshot: Optional[dict] = None,
) -> dict:
    provided_snapshot = snapshot is not None
    if snapshot is None:
        with database("dashboard.overview") as conn:
            rows, raw_count = _latest_rows(conn, start, end)
            mapping_snapshot = aftersaler_mapping.get_mapping_snapshot(end, conn=conn)
            dimensions, quality = _load_dimensions(conn, rows, mapping_snapshot)
            filter_options = _dashboard_filter_options(dimensions)
            allowed_groups = {
                group_name
                for group_name, dim in dimensions.items()
                if _dimension_matches_or_degrades(dim, region, aftersaler, category, key_account)
            }
            rows = [row for row in rows if row["groupName"] in allowed_groups]
            dimensions = {
                name: _scope_dimension(value, region, category, key_account, aftersaler)
                for name, value in dimensions.items()
                if name in allowed_groups
            }
            quality = _quality_from_dimensions(dimensions, quality)
            scoped_group_names = sorted({row.get("groupName") for row in rows if row.get("groupName")})
            raw_count = _raw_analysis_count(conn, start, end, scoped_group_names)
            unanswered_rows = _latest_unanswered_rows(conn, start, end, scoped_group_names)
            group_rows = _query_group_rows(conn, scoped_group_names)
            audit_start = start - timedelta(days=2)
            audit_end = end + timedelta(days=1)
            audit_chat_rows = _query_chat_rows(conn, group_rows, audit_start, audit_end)
            chat_rows = _filter_chat_rows_by_period(audit_chat_rows, start, end)
            chat_rows_by_group = _raw_messages_by_group(group_rows, chat_rows)
            audit_raw_messages = _raw_messages_by_group(group_rows, audit_chat_rows)
            snapshot_raw_rows = list(rows)
            if raw_count > len(rows) and scoped_group_names:
                snapshot_raw_rows = _query_by_chunks(
                    conn,
                    "overview.analysis_raw_snapshot",
                    """SELECT * FROM qx_analysis_result
                       WHERE CREATEDTIME >= %s AND CREATEDTIME < %s
                         AND groupName IN ({placeholders})
                       ORDER BY CREATEDTIME""",
                    scoped_group_names,
                    prefix_params=[start.isoformat(), (end + timedelta(days=1)).isoformat()],
                )
            snapshot = {
                "latest_rows": rows,
                "unanswered_rows": unanswered_rows,
                "raw_rows": snapshot_raw_rows,
                "raw_count_before_filter": raw_count,
                "dimensions": dimensions,
                "quality": quality,
                "filter_options": filter_options,
                "group_names": scoped_group_names,
                "project_codes": sorted({
                    str(code).upper()
                    for dimension in dimensions.values()
                    for code in dimension.get("codes", [])
                    if code
                }),
                "group_rows": group_rows,
                "chat_rows": chat_rows,
                "chat_rows_by_group": chat_rows_by_group,
                "audit_chat_rows": audit_chat_rows,
                "audit_raw_messages": audit_raw_messages,
            }
    else:
        rows = snapshot.get("latest_rows", [])
        unanswered_rows = snapshot.get("unanswered_rows", [])
        dimensions = snapshot.get("dimensions", {})
        quality = snapshot.get("quality", {})
        filter_options = snapshot.get("filter_options", _dashboard_filter_options(dimensions))
        raw_count = int(snapshot.get("raw_count_before_filter") or len(snapshot.get("raw_rows", [])))
        group_rows = snapshot.get("group_rows", [])
        chat_rows = snapshot.get("chat_rows", [])
        chat_rows_by_group = snapshot.get("chat_rows_by_group") or _raw_messages_by_group(group_rows, chat_rows)
        audit_chat_rows = snapshot.get("audit_chat_rows", chat_rows)
        audit_raw_messages = snapshot.get("audit_raw_messages") or _raw_messages_by_group(group_rows, audit_chat_rows)

    groups = _group_aggregates(rows, dimensions, unanswered_rows)
    durations = _active_durations_from_lims(dimensions)
    missing_duration_groups = sorted(name for name in groups if name not in durations)

    _filter_region = region
    _filter_aftersaler = aftersaler
    _filter_category = category
    _filter_key_account = key_account
    total_groups = len(groups)
    total_messages = sum(item["messages"] for item in groups.values())
    unanswered_summary = _reconciled_unanswered_summary(unanswered_rows, audit_raw_messages)
    unanswered_group_names = {
        str(row.get("groupName") or "").strip()
        for row in unanswered_rows
        if str(row.get("groupName") or "").strip()
    }
    unanswered_total_groups = len(unanswered_group_names)
    missed_groups = len(unanswered_summary["groups"])
    daily: Dict[str, dict] = defaultdict(lambda: {"messages": 0, "groups": set(), "missed": 0})
    customer_good = customer_bad = employee_positive = employee_negative = 0
    words = Counter()
    regions = Counter()
    region_messages = Counter()
    region_group_count: Dict[str, int] = {}
    aftersalers = Counter()
    aftersaler_messages = Counter()
    aftersaler_projects: Dict[str, set] = defaultdict(set)
    tentative_aftersalers = Counter()
    categories = Counter()
    category_groups: Dict[str, set] = defaultdict(set)
    category_messages = Counter()
    category_projects: Dict[str, set] = defaultdict(set)
    product_tree = defaultdict(lambda: defaultdict(lambda: {"projects": set(), "groups": set(), "messages": 0}))
    region_sales = defaultdict(Counter)
    region_after = defaultdict(Counter)
    region_product = defaultdict(Counter)
    key_accounts: Dict[str, dict] = {}
    attention_projects: Dict[Tuple[str, str], dict] = {}
    attention_status_groups: Dict[str, set] = defaultdict(set)
    attention_status_messages = Counter()

    for row in rows:
        day = str(row.get("CREATEDTIME"))[:10]
        daily[day]["messages"] += int(row.get("messageToDayCount") or 0)
        daily[day]["groups"].add(row["groupName"])
    for day, group_names in unanswered_summary["daily_groups"].items():
        daily[day]["missed"] = len(group_names)

    for group_name, item in groups.items():
        customer_good += item["customer_good"]
        customer_bad += item["customer_bad"]
        employee_positive += item["employee_positive"]
        employee_negative += item["employee_negative"]
        words.update(item["high_freq"])
        dim = item["dimension"]
        project_codes_for_group = {
            project.get("project_code")
            for project in dim.get("projects", [])
            if project.get("project_code")
        }
        group_aftersalers = dim.get("aftersalers", [])
        raw_group_message_count = len(chat_rows_by_group.get(group_name, []))
        attributed_message_counts = _aftersaler_message_counts(
            group_aftersalers,
            chat_rows_by_group.get(group_name, []),
            item["messages"],
        )
        for person in dim.get("aftersalers", []):
            aftersalers[person] += 1
            aftersaler_messages[person] += attributed_message_counts.get(person, 0)
            aftersaler_projects[person].update(project_codes_for_group)
        for person in dim.get("tentative_aftersalers", []):
            tentative_aftersalers[person] += 1
        group_regions = set(dim.get("regions", [])) or {"未关联区域"}
        for region in group_regions:
            regions[region] += 1
            region_group_count[region] = region_group_count.get(region, 0) + 1
            region_messages[region] += item["messages"]
        seen_pairs = set()
        seen_message_categories = set()
        seen_message_leaves = set()
        for project in dim.get("projects", []):
            region = project.get("region") or "未关联区域"
            category = project.get("category_l2") or ""
            if category in ALLOWED_PRODUCT_L2:
                pair = (project.get("project_code"), region, category)
                if pair not in seen_pairs:
                    categories[category] += 1
                    category_projects[category].add(project.get("project_code"))
                    category_groups[category].add(group_name)
                    region_product[region][category] += 1
                    level3 = project.get("category_l3") or "未细分"
                    leaf = product_tree[category][level3]
                    leaf["projects"].add(project.get("project_code"))
                    leaf["groups"].add(group_name)
                    seen_pairs.add(pair)
                if category not in seen_message_categories:
                    category_messages[category] += raw_group_message_count
                    seen_message_categories.add(category)
                leaf_key = (category, project.get("category_l3") or "未细分")
                if leaf_key not in seen_message_leaves:
                    product_tree[leaf_key[0]][leaf_key[1]]["messages"] += raw_group_message_count
                    seen_message_leaves.add(leaf_key)
            attention_status = str(project.get("analysis_simple_remark") or "").strip()
            if attention_status in PROJECT_ATTENTION_STATUSES:
                if group_name not in attention_status_groups[attention_status]:
                    attention_status_groups[attention_status].add(group_name)
                    attention_status_messages[attention_status] += item["messages"]
                project_code = str(project.get("project_code") or "").upper()
                attention_key = (attention_status, project_code or f"{group_name}:{project.get('product_name') or ''}")
                attention = attention_projects.setdefault(attention_key, {
                    "status": attention_status,
                    "project_code": project_code,
                    "project_name": project.get("product_name") or "",
                    "category_l2": category,
                    "category_l3": project.get("category_l3") or "",
                    "customer_name": project.get("customer_name") or "",
                    "work_unit": project.get("work_unit") or "",
                    "key_account": project.get("key_account") or "",
                    "region": project.get("region") or "",
                    "sales_person": project.get("sales_person") or "",
                    "aftersaler": project.get("final_aftersaler") or project.get("raw_aftersaler") or "",
                    "groups": set(),
                    "message_count": 0,
                    "active_days": [],
                    "start_time": project.get("start_time") or "",
                    "end_time": project.get("end_time") or "",
                })
                if group_name not in attention["groups"]:
                    attention["groups"].add(group_name)
                    attention["message_count"] += item["messages"]
                if parse_active_day(project.get("active_day")) is not None:
                    attention["active_days"].append(parse_active_day(project.get("active_day")))
            sales = project.get("sales_person") or "未分配销售"
            region_sales[region][sales] += 1
            for after in dim.get("aftersalers", []) or ["未关联售后"]:
                region_after[region][after] += 1
            key = project.get("key_account")
            if key:
                account = key_accounts.setdefault(key, {
                    "key_account": key, "projects": set(), "customers": set(), "aftersalers": set(),
                    "groups": set(), "messages": 0, "work_units": set(), "category_l3": set(),
                    "high_freq": Counter(), "active_days": [],
                })
                account["projects"].add(project.get("project_code"))
                if project.get("customer_name"):
                    account["customers"].add(project["customer_name"])
                account["aftersalers"].update(dim.get("aftersalers", []))
                if group_name not in account["groups"]:
                    account["messages"] += item["messages"]
                    account["high_freq"].update(item["high_freq"])
                account["groups"].add(group_name)
                if project.get("work_unit"):
                    account["work_units"].add(project["work_unit"])
                if project.get("category_l3"):
                    account["category_l3"].add(project["category_l3"])
                if parse_active_day(project.get("active_day")) is not None:
                    account["active_days"].append(parse_active_day(project.get("active_day")))
    if _filter_aftersaler:
        aftersalers = Counter({_filter_aftersaler: aftersalers.get(_filter_aftersaler, 0)})
    if _filter_region:
        regions = Counter({_filter_region: regions.get(_filter_region, 0)})
        region_group_count = {_filter_region: region_group_count.get(_filter_region, 0)}
    if _filter_category:
        categories = Counter({_filter_category: categories.get(_filter_category, 0)})

    duration_defs = [
        ("≤7天", "极短期咨询", 0, 7), ("8-30天", "短期服务", 8, 30),
        ("1-3个月", "常规项目周期", 31, 90), ("3-6个月", "中长期项目", 91, 180),
        ("6-12个月", "长期服务", 181, 365), (">12个月", "超长期合作", 366, 10**9),
    ]
    duration_items = []
    for range_name, label, low, high in duration_defs:
        count = sum(1 for value in durations.values() if low <= value <= high)
        duration_items.append({"range": range_name, "label": label, "count": count, "percentage": _ratio(count, len(durations))})

    total_region_groups = sum(regions.values())
    region_items = []
    for region, count in regions.most_common():
        region_items.append({
            "region": region, "group_count": count,
            "message_count": region_messages[region],
            "percentage": _ratio(count, total_region_groups),
        })

    top5_coverage = _ratio(sum(x[1] for x in regions.most_common(5)), total_region_groups)
    account_items = []
    for value in key_accounts.values():
        customer_names = sorted(value.get("customers", set()))
        work_units = sorted(value.get("work_units", set()))
        account_items.append({
            "key_account": value["key_account"],
            "work_unit": work_units[0] if work_units else "",
            "work_units": work_units[:5],
            "customer_name": customer_names[0] if customer_names else "",
            "customer_names": customer_names[:5],
            "group_count": len(value.get("groups", set())),
            "project_count": len(value["projects"]),
            "message_count": value.get("messages", 0),
            "customer_count": len(value["customers"]),
            "aftersalers": sorted(value["aftersalers"]),
            "category_l3": sorted(value.get("category_l3", set()))[:8],
            "high_frequency_top5": [
                {"word": word, "count": count}
                for word, count in value.get("high_freq", Counter()).most_common(5)
            ],
            "active_day": max(value.get("active_days", []) or [0]),
        })
    account_items.sort(key=lambda value: (-value["group_count"], -value["project_count"], value["key_account"]))
    top_message_groups = _build_top_message_groups(groups, total_messages)
    project_year_distribution = _build_project_year_distribution(groups)
    attention_items = []
    status_order = {status: index for index, status in enumerate(PROJECT_ATTENTION_STATUSES)}
    for value in attention_projects.values():
        attention_items.append({
            **{key: item_value for key, item_value in value.items() if key not in ("groups", "active_days")},
            "group_count": len(value["groups"]),
            "group_names": sorted(value["groups"]),
            "active_day": max(value.get("active_days", []) or [0]),
        })
    attention_items.sort(key=lambda value: (
        status_order.get(value["status"], 99),
        -value["message_count"],
        value["project_code"],
    ))
    attention_summary = []
    total_attention_projects = len(attention_items)
    for status in PROJECT_ATTENTION_STATUSES:
        status_items = [value for value in attention_items if value["status"] == status]
        attention_summary.append({
            "status": status,
            "project_count": len(status_items),
            "group_count": len(attention_status_groups[status]),
            "message_count": attention_status_messages[status],
            "percentage": _ratio(len(status_items), total_attention_projects),
        })
    product_hierarchy = []
    for level2, level3_map in product_tree.items():
        children = []
        for level3, stats in level3_map.items():
            children.append({
                "name": level3,
                "project_count": len(stats["projects"]),
                "group_count": len(stats["groups"]),
                "message_count": stats["messages"],
                "count": len(stats["projects"]),
            })
        children.sort(key=lambda value: (-value["project_count"], value["name"]))
        product_hierarchy.append({
            "name": level2,
            "project_count": len(category_projects[level2]),
            "group_count": len(category_groups[level2]),
            "message_count": category_messages[level2],
            "count": len(category_projects[level2]),
            "children": children,
        })
    product_hierarchy.sort(key=lambda value: (-value["project_count"], value["name"]))

    matched_codes = quality["matched_project_codes"]
    project_codes = quality["project_codes"]
    matched_products = quality["matched_products"]
    product_projects = quality["product_projects"]
    aftersaler_groups = quality.get("groups_with_aftersaler", quality.get("groups_with_confirmed_aftersaler", 0))
    cross_sales = [{"region": region, "group_count": region_group_count.get(region, 0), "message_count": region_messages.get(region, 0), "items": _counter_items(counter)} for region, counter in region_sales.items()]
    cross_after = [{"region": region, "group_count": region_group_count.get(region, 0), "message_count": region_messages.get(region, 0), "items": _counter_items(counter)} for region, counter in region_after.items()]
    cross_product = [{"region": region, "group_count": region_group_count.get(region, 0), "message_count": region_messages.get(region, 0), "items": _counter_items(counter, "category")} for region, counter in region_product.items()]

    scope_key = _dashboard_snapshot_scope_key(
        start, end, period, region, aftersaler, category, key_account,
    )
    snapshot_id = (
        str(snapshot.get("_snapshot_id") or "")
        if provided_snapshot
        else _store_dashboard_snapshot(scope_key, snapshot)
    )
    return {
        "meta": {
            "period": period, "start_date": start.isoformat(), "end_date": end.isoformat(),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "snapshot_id": snapshot_id,
            "source_min_date": min((str(row["CREATEDTIME"])[:10] for row in rows), default=None),
            "source_max_date": max((str(row["CREATEDTIME"])[:10] for row in rows), default=None),
            "unanswered_source": UNANSWERED_RESULT_TABLE,
            "unanswered_source_min_date": min(
                (str(row["CREATEDTIME"])[:10] for row in unanswered_rows), default=None,
            ),
            "unanswered_source_max_date": max(
                (str(row["CREATEDTIME"])[:10] for row in unanswered_rows), default=None,
            ),
            "filters": {"region": region, "aftersaler": aftersaler, "category": category, "key_account": key_account},
            "allowed_product_categories": sorted(ALLOWED_PRODUCT_L2),
            "filter_options": filter_options,
        },
        "summary": {
            "total_groups": total_groups, "total_messages": total_messages,
            "project_groups": sum(1 for item in groups.values() if item["dimension"].get("codes")),
            "regions": len([name for name in regions if name != "未关联区域"]),
            "aftersaler_count": len(aftersalers),
            "confirmed_aftersalers": len(aftersalers), "product_categories": len(categories),
            "key_accounts": len(key_accounts),
            "short_active_ratio": _ratio(sum(1 for value in durations.values() if value <= 30), len(durations)),
        },
        "service_quality": {
            "unanswered": {
                "total_groups": unanswered_total_groups,
                "missed_groups": missed_groups,
                "answered_groups": max(0, unanswered_total_groups - missed_groups),
                "missed_rate": _ratio(missed_groups, unanswered_total_groups),
                "review_groups": len(unanswered_summary["review_groups"]),
                "definition": "分母和漏回判定均来自每日零点夜间分析表；漏回消息还需匹配原始客户消息，且截至统计截止未发现员工回复。",
                "source": UNANSWERED_RESULT_TABLE,
            },
            "sentiment": {"customer_good": customer_good, "customer_bad": customer_bad, "employee_positive": employee_positive, "employee_negative": employee_negative},
        },
        "communication": {
            "trend": [{"date": day, "messages": value["messages"], "groups": len(value["groups"]), "missed": value["missed"]} for day, value in sorted(daily.items())],
            "high_frequency": [{"word": word, "count": count} for word, count in words.most_common(20)],
            "active_duration": duration_items,
            "time_period_breakdown": _time_period_breakdown(groups, dimensions, chat_rows_by_group),
            "top_message_groups": top_message_groups,
        },
        "business": {
            "aftersaler_message_definition": "消息统一来自 qx_chat：单一最终售后群统计群内全部原始消息；多最终售后群仅统计各售后本人发送的原始消息。",
            "aftersalers": [
                {
                    "name": name,
                    "count": count,
                    "group_count": count,
                    "project_count": len(aftersaler_projects[name]),
                    "message_count": aftersaler_messages[name],
                    "percentage": _ratio(count, sum(aftersalers.values())),
                }
                for name, count in aftersalers.most_common()
            ],
            "tentative_aftersalers": _counter_items(tentative_aftersalers),
            "regions": region_items, "top5_coverage": top5_coverage,
            "product_categories": [
                {
                    "category": name,
                    "count": len(category_projects[name]),
                    "project_count": len(category_projects[name]),
                    "group_count": len(category_groups[name]),
                    "message_count": category_messages[name],
                    "percentage": _ratio(len(category_projects[name]), sum(len(v) for v in category_projects.values())),
                }
                for name, _ in categories.most_common()
            ],
            "product_hierarchy": product_hierarchy,
            "key_accounts": account_items,
            "project_year_distribution": project_year_distribution,
        },
        "project_attention": {
            "target_statuses": list(PROJECT_ATTENTION_STATUSES),
            "total_projects": total_attention_projects,
            "all_projects": product_projects,
            "summary": attention_summary,
            "items": attention_items,
        },
        "cross_analysis": {
            "region_sales": sorted(cross_sales, key=lambda x: -x.get("group_count", 0)),
            "region_after": sorted(cross_after, key=lambda x: -x.get("group_count", 0)),
            "region_product": sorted(cross_product, key=lambda x: -x.get("group_count", 0)),
        },
        "data_quality": {
            "raw_rows": raw_count, "deduplicated_rows": len(rows),
            "duplicate_rows_removed": max(0, raw_count - len(rows)),
            "project_codes": project_codes, "matched_project_codes": matched_codes,
            "unmatched_project_codes": quality.get("unmatched_project_codes", max(0, project_codes - matched_codes)),
            "groups_with_project_code": quality.get("groups_with_project_code", sum(1 for item in groups.values() if item["dimension"].get("codes"))),
            "groups_without_project_code": quality.get("groups_without_project_code", total_groups - sum(1 for item in groups.values() if item["dimension"].get("codes"))),
            "groups_without_lims_link": quality.get("groups_without_lims_link", 0),
            "project_match_rate": _ratio(matched_codes, project_codes),
            "product_records": product_projects, "matched_products": matched_products,
            "product_match_rate": _ratio(matched_products, product_projects),
            "groups_with_aftersaler": aftersaler_groups,
            "aftersaler_coverage_rate": _ratio(aftersaler_groups, total_groups),
            "groups_with_confirmed_aftersaler": aftersaler_groups,
            "aftersaler_confirmation_rate": _ratio(aftersaler_groups, total_groups),
            "groups_with_lims_link": quality.get("groups_with_lims_link", 0),
            "lims_link_rate": _ratio(quality.get("groups_with_lims_link", 0), total_groups),
            "groups_with_region": quality.get("groups_with_region", 0),
            "region_link_rate": _ratio(quality.get("groups_with_region", 0), total_groups),
            "groups_with_key_account": quality.get("groups_with_key_account", 0),
            "key_account_link_rate": _ratio(quality.get("groups_with_key_account", 0), total_groups),
            "groups_with_raw_aftersaler": quality.get("groups_with_raw_aftersaler", 0),
            "groups_with_lims_members": quality.get("groups_with_lims_members", 0),
            "active_duration_source": "lims_base_data_active_day",
            "active_duration_lims_groups": len(durations),
            "active_duration_missing_groups": len(missing_duration_groups),
            "lims_source": quality.get("lims_source", "unknown"),
            "lims_api_requests": quality.get("lims_api_requests", 0),
            "lims_api_records": quality.get("lims_api_records", 0),
            "lims_api_errors": quality.get("lims_api_errors", 0),
            "lims_api_cache_hits": quality.get("lims_api_cache_hits", 0),
            "lims_api_stale_hits": quality.get("lims_api_stale_hits", 0),
            "mapping_available": quality.get("mapping_available", False),
            "mapping_version_id": quality.get("mapping_version_id"),
            "mapping_effective_month": quality.get("mapping_effective_month"),
            "mapping_revision": quality.get("mapping_revision", 0),
            "mapping_reason": quality.get("mapping_reason", "unknown"),
            "mapping_matched_records": quality.get("mapping_matched_records", 0),
            "mapping_fallback_records": quality.get("mapping_fallback_records", 0),
            "mapping_conflict_records": quality.get("mapping_conflict_records", 0),
            "mapping_match_rate": _ratio(
                quality.get("mapping_matched_records", 0),
                quality.get("mapping_matched_records", 0)
                + quality.get("mapping_fallback_records", 0)
                + quality.get("mapping_conflict_records", 0),
            ),
            "note": "业务维度仅来源于 POST /unionLims/base_data/。售后人员优先按 productBigSortThree + orgName + afterSaler 对应表确定，未命中或配置不可用时回退 afterSaler；销售区域取 orgName，重点客户取 keyAccount，活跃周期取 activeDay/activellay。",
        },
    }


def get_overview(
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    force_refresh: bool = False,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> dict:
    start, end, normalized_period = resolve_period(period, start_date, end_date)
    key = f"{start}:{end}:{normalized_period}:{region}:{aftersaler}:{category}:{key_account}"
    now = time.time()
    with _cache_lock:
        cached = _cache.get(key)
        if not force_refresh and cached and now - cached[0] < CACHE_TTL_SECONDS:
            result = copy.deepcopy(cached[1])
            result["meta"]["cache"] = "hit"
            return result
        inflight = _overview_inflight.get(key)
        if inflight is None:
            inflight = threading.Event()
            _overview_inflight[key] = inflight
            is_builder = True
        else:
            is_builder = False
    if not is_builder:
        inflight.wait(timeout=35)
        with _cache_lock:
            completed = _cache.get(key)
            stale = _last_success.get(key)
        if completed:
            result = copy.deepcopy(completed[1])
            result["meta"]["cache"] = "coalesced"
            return result
        if stale:
            result = copy.deepcopy(stale)
            result["meta"]["stale"] = True
            result["meta"]["stale_reason"] = "refresh_incomplete"
            return result
        raise TimeoutError("dashboard overview refresh did not complete")
    try:
        result = _build_overview(
            start, end, normalized_period,
            region=region, aftersaler=aftersaler, category=category, key_account=key_account,
        )
        result["meta"]["cache"] = "miss"
        result["meta"]["stale"] = False
        with _cache_lock:
            _cache[key] = (now, copy.deepcopy(result))
            _last_success[key] = copy.deepcopy(result)
        return result
    except Exception as exc:
        logger.exception("dashboard.overview.error key={} error={}", key, exc)
        with _cache_lock:
            stale = _last_success.get(key)
        if stale:
            result = copy.deepcopy(stale)
            result["meta"]["stale"] = True
            result["meta"]["stale_reason"] = type(exc).__name__
            return result
        raise
    finally:
        with _cache_lock:
            event = _overview_inflight.pop(key, None)
            if event:
                event.set()


def _build_evidence(
    metric: str,
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    keyword: Optional[str] = None,
    search: Optional[str] = None,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
    page: int = 1,
    page_size: int = 20,
) -> dict:
    start, end, normalized_period = resolve_period(period, start_date, end_date)
    metric = metric.lower()
    with database("dashboard.evidence") as conn:
        analysis_rows, _ = _latest_rows(conn, start, end)
        mapping_snapshot = aftersaler_mapping.get_mapping_snapshot(end, conn=conn)
        dimensions, _ = _load_dimensions(conn, analysis_rows, mapping_snapshot)
        dimensions = {
            group_name: _scope_dimension(dim, region, category, key_account, aftersaler)
            for group_name, dim in dimensions.items()
            if _dimension_matches_or_degrades(dim, region, aftersaler, category, key_account)
        }
        allowed_groups = [
            row.get("groupName") for row in analysis_rows
            if row.get("groupName") in dimensions
        ]
        rows = (
            _latest_unanswered_rows(conn, start, end, sorted(set(allowed_groups)))
            if metric == "unanswered"
            else analysis_rows
        )
        group_rows = _query_group_rows(conn, sorted(set(allowed_groups)))
        audit_chat_rows = _query_chat_rows(
            conn, group_rows, start - timedelta(days=2), end + timedelta(days=1)
        )
        raw_messages = _raw_messages_by_group(group_rows, audit_chat_rows)
    items = []
    unanswered_items: Dict[str, dict] = {}
    verification_counts = Counter()
    seen_unanswered = set()
    for row in reversed(rows):
        group_name = row.get("groupName") or ""
        dim = dimensions.get(group_name, {})
        if group_name not in dimensions:
            continue
        content = ""
        matched = False
        if metric == "unanswered":
            matched = str(row.get("isMissedMessage")) == "1"
            content = row.get("missedMessageList") or ""
            missed_messages = _extract_missed_messages(content)
        elif metric == "customer_negative":
            mapping = parse_emotion_field(row.get("customerEmotionAnalysis"))
            matched = _emotion_total(mapping, ("差评", "负向", "不满")) > 0 or bool(row.get("customerNegativeEmotionInfo"))
            content = row.get("customerNegativeEmotionInfo") or ""
            missed_messages = []
        elif metric == "employee_negative":
            mapping = parse_emotion_field(row.get("saleEmotionAnalysis"))
            matched = _emotion_total(mapping, ("恶劣", "负向", "消极")) > 0 or bool(row.get("saleNegativeEmotionInfo"))
            content = row.get("saleNegativeEmotionInfo") or ""
            missed_messages = []
        elif metric == "highfreq":
            words = parse_high_freq(row.get("highFrequencyWords"))
            matched = bool(words) and (not keyword or any(keyword.lower() in value["word"].lower() for value in words))
            content = row.get("highFrequencyWords") or ""
            missed_messages = []
        else:
            raise ValueError("unsupported evidence metric")
        if not matched:
            continue
        if search and search.lower() not in (group_name + " " + str(content) + " " + str(row.get("member") or "")).lower():
            continue
        group_raw_messages = raw_messages.get(group_name, [])
        analysis_dt = parse_msg_datetime(row.get("CREATEDTIME"))
        if metric == "unanswered":
            evaluated = _evaluate_unanswered_messages(
                missed_messages, group_raw_messages, analysis_dt
            )
            for message in evaluated:
                verification_counts[message.get("verification_status") or "unverified"] += 1
            display_messages = []
            for message in evaluated:
                if message.get("verification_status") != "unanswered":
                    continue
                identity = (
                    group_name,
                    str(message.get("msgid") or "").strip()
                    or "|".join((
                        str(message.get("sender_userid") or message.get("sender_name") or "").strip(),
                        _normalize_chat_content(str(message.get("content") or "")),
                        str(message.get("msgtime") or "").strip(),
                    )),
                )
                if identity in seen_unanswered:
                    continue
                seen_unanswered.add(identity)
                display_messages.append(message)
            if not display_messages:
                continue
        else:
            display_messages = _evidence_messages(
                metric, content, keyword or "", missed_messages, group_raw_messages
            )
        msg_times = [item["msgtime"] for item in display_messages if item.get("msgtime")]
        evidence_item = {
            "id": row.get("id"), "group_name": group_name,
            "analysis_time": str(row.get("CREATEDTIME"))[:19].replace("T", " "),
            "analysis_date": str(row.get("CREATEDTIME"))[:19].replace("T", " "),
            "msg_times": msg_times,
            "messages": display_messages,
            "members": parse_members(row.get("member")),
            "content": content, "core_summary": row.get("coreInfoSummary") or "",
            "project_codes": dim.get("codes", []), "projects": dim.get("projects", []),
            "aftersalers": dim.get("aftersalers", []),
        }
        if metric == "unanswered":
            existing = unanswered_items.get(group_name)
            if existing:
                existing["messages"].extend(display_messages)
                existing["msg_times"].extend(msg_times)
                if evidence_item["analysis_time"] > existing["analysis_time"]:
                    existing["analysis_time"] = evidence_item["analysis_time"]
                    existing["analysis_date"] = evidence_item["analysis_date"]
            else:
                unanswered_items[group_name] = evidence_item
        else:
            items.append(evidence_item)
    if metric == "unanswered":
        items = list(unanswered_items.values())
    total = len(items)
    start_index = (page - 1) * page_size
    return {
        "metric": metric, "period": normalized_period,
        "start_date": start.isoformat(), "end_date": end.isoformat(),
        "total": total, "page": page, "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
        "definition": ({
            "unit": "一条明细是一条能匹配到 qx_chat 原始记录、由客户发送且截至统计截止仍未发现员工后续回复的消息。",
            "message_time": "消息发送时间来自 qx_chat.msgtime，是客户原消息的发送时间。",
            "analysis_time": f"分析时间来自 {UNANSWERED_RESULT_TABLE}.CREATEDTIME，是每日零点夜间分析的批次时间，不是消息发送时间。",
            "source": f"漏回判定统一来自 {UNANSWERED_RESULT_TABLE}，原始消息仅用于证据匹配和后续回复核验。",
            "sender": "发送人姓名仅取原始姓名字段；from/roomid 只用于身份核验，不作为姓名展示。",
            "deduplication": "同一 msgid，或同一发送人、正文和原始发送时间，只展示一次。",
        } if metric == "unanswered" else {}),
        "verification": dict(verification_counts) if metric == "unanswered" else {},
        "items": items[start_index:start_index + page_size],
    }


def get_evidence(
    metric: str,
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    keyword: Optional[str] = None,
    search: Optional[str] = None,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
    page: int = 1,
    page_size: int = 20,
) -> dict:
    key = json.dumps(
        [metric, period, start_date, end_date, keyword, search, region, aftersaler,
         category, key_account, page, page_size],
        ensure_ascii=False,
    )
    now = time.time()
    with _cache_lock:
        cached = _evidence_cache.get(key)
        if cached and now - cached[0] < EVIDENCE_CACHE_TTL_SECONDS:
            return copy.deepcopy(cached[1])
        inflight = _evidence_inflight.get(key)
        if inflight is None:
            inflight = threading.Event()
            _evidence_inflight[key] = inflight
            is_builder = True
        else:
            is_builder = False
    if not is_builder:
        inflight.wait(timeout=35)
        with _cache_lock:
            completed = _evidence_cache.get(key)
        if completed:
            return copy.deepcopy(completed[1])
        raise TimeoutError("dashboard evidence refresh did not complete")
    try:
        result = _build_evidence(
            metric=metric, period=period, start_date=start_date, end_date=end_date,
            keyword=keyword, search=search, region=region, aftersaler=aftersaler,
            category=category, key_account=key_account, page=page, page_size=page_size,
        )
        with _cache_lock:
            _evidence_cache[key] = (time.time(), copy.deepcopy(result))
        return result
    finally:
        with _cache_lock:
            event = _evidence_inflight.pop(key, None)
            if event:
                event.set()


# Compatibility helpers for existing callers.
def get_summary(date_str: Optional[str] = None) -> dict:
    data = get_overview("custom", date_str, date_str) if date_str else get_overview("year")
    return {**data["summary"], "date_range": f"{data['meta']['start_date']} ~ {data['meta']['end_date']}", "missed_groups": data["service_quality"]["unanswered"]["missed_groups"]}


def get_full_summary() -> dict:
    return get_summary()


def get_after_saler_distribution() -> dict:
    data = get_overview("year")
    return {"items": data["business"]["aftersalers"], "total_groups": data["summary"]["total_groups"]}


def get_product_category_hierarchy() -> dict:
    data = get_overview("year")
    return {
        "categories": data["business"]["product_hierarchy"],
        "total_projects": data["data_quality"]["product_records"],
    }


def get_key_account_hierarchy() -> dict:
    data = get_overview("year")
    return {"hierarchy": data["business"]["key_accounts"], "total_key_accounts": data["summary"]["key_accounts"]}


def get_org_distribution() -> dict:
    data = get_overview("year")
    return {"items": data["business"]["regions"], "total_regions": data["summary"]["regions"], "top5_coverage": data["business"]["top5_coverage"]}


def get_sentiment_analysis_summary() -> dict:
    return get_overview("year")["service_quality"]["sentiment"]


def get_high_freq_summary(limit: int = 20) -> dict:
    values = get_overview("year")["communication"]["high_frequency"][:limit]
    return {"top_words": values, "total_unique_words": len(values)}





def _parse_missed_text(text: str) -> List[dict]:
    """解析纯文本格式的漏回消息：每行 "<sender>:<content>"。

    兼容：
    - sender 含空格、姓名+手机号等（如 "张敏  13103738626"）
    - 同一行包含多条消息（用中/英逗号分隔）
    - "<sender>:「<quoted message>」"（含全角引号的内容，引号内部的冒号
      不会再被切分成新消息）
    - 文本开头的纯自由文本（无 sender:content 格式）作为独立一条

    注：中间夹在两条 sender:content 之间的纯文本片段（测试数据中
    `"南枝:... \n 1呢,南枝:..."` 的 "1呢"）会归入上一条的 content，
    以保持现有 `test_extract_missed_messages_parses_pure_text_with_multiple_messages`
    的断言。
    """
    messages: List[dict] = []
    if not text:
        return messages

    positions = _find_sender_colon_matches(text)

    if not positions:
        stripped = text.strip()
        if stripped:
            messages.append({
                "msgid": "",
                "sender_name": "",
                "sender_userid": "",
                "sender_role": "未知",
                "roomid": "",
                "content": stripped,
                "msgtime": "",
                "time_source": "unavailable",
            })
        return messages

    # 1) 解析开头的自由文本（无 sender 的部分）
    first_start, _, _ = positions[0]
    if first_start > 0:
        prefix = text[:first_start].rstrip(",，").strip()
        if prefix:
            messages.append({
                "msgid": "",
                "sender_name": "",
                "sender_userid": "",
                "sender_role": "未知",
                "roomid": "",
                "content": prefix,
                "msgtime": "",
                "time_source": "unavailable",
            })

    # 2) 解析 sender:content 对
    #    content = 冒号之后、到下一条 sender:content 起点之间的文本
    for i, (start, end, sender) in enumerate(positions):
        if i + 1 < len(positions):
            next_start = positions[i + 1][0]
            content = text[end:next_start].rstrip(",，").strip()
        else:
            content = text[end:].strip()
        if not content:
            continue
        # 兜底：sender 太长（>60字符）很可能是误识别
        if len(sender) > 60:
            content = f"{sender}:{content}"
            sender = ""
        messages.append({
            "msgid": "",
            "sender_name": sender,
            "sender_userid": "",
            "sender_role": "未知",
            "roomid": "",
            "content": content,
            "msgtime": "",
            "time_source": "unavailable",
        })

    return messages


def _find_sender_colon_matches(text: str) -> List[Tuple[int, int, str]]:
    """在纯文本中查找所有 sender:content 的边界。

    返回 `[(start, end_of_colon, sender), ...]`：
    - `start`：sender 的起始下标（含）
    - `end_of_colon`：冒号之后的下标（content 起点）
    - `sender`：strip 后的发送人文本

    规则：
    - 仅在「」之外的顶层查找（引号内不再切分）
    - sender 不能包含冒号、换行、逗号、「、」
    - 非贪婪：从冒号往前逐字回溯，遇到分隔符即停
    """
    results: List[Tuple[int, int, str]] = []
    bracket_depth = 0
    for i, ch in enumerate(text):
        if ch in "「『\u201c":
            bracket_depth += 1
            continue
        if ch in "」』\u201d":
            bracket_depth = max(0, bracket_depth - 1)
            continue
        if bracket_depth != 0 or ch not in "：:":
            continue
        # 找到顶层冒号，从 i-1 向前回溯找 sender
        j = i - 1
        chars: List[str] = []
        sender_start = i
        while j >= 0:
            cj = text[j]
            if cj in "：:\n,，\"「」『』\u201c\u201d":
                break
            chars.insert(0, cj)
            j -= 1
            if len(chars) >= 80:
                break
        sender = "".join(chars).strip()
        if not sender:
            continue
        sender_start = j + 1
        results.append((sender_start, i + 1, sender))
    return results


def _extract_missed_messages(missed_list_json: Any) -> List[dict]:
    """Normalize missedMessageList into displayable original messages.
    支持三种格式（按优先级尝试）：
      1) JSON 数组 [{msgid, msgtime, content, ...}, ...]
      2) 纯文本 "<sender>:<content>" 多行 / 同行多条
      3) 兜底：整段当作一条 content
    """
    if not missed_list_json:
        return []
    text = missed_list_json.strip() if isinstance(missed_list_json, str) else str(missed_list_json).strip()
    if not text:
        return []

    # 1) 尝试 JSON 解析
    try:
        import json
        items = json.loads(text) if isinstance(missed_list_json, str) else text
        if isinstance(items, list):
            messages = []
            for item in items:
                if not isinstance(item, dict):
                    continue
                msgtime = (
                    item.get("msgtime") or item.get("msgTime") or item.get("msg_time")
                    or item.get("time") or item.get("timestamp")
                    or item.get("createTime") or item.get("createtime")
                    or ""
                )
                sender_userid = first_nonempty(
                    item.get("sender_userid"), item.get("from_userid"),
                    item.get("fromUserId"), item.get("from"),
                )
                roomid = first_nonempty(item.get("roomid"), item.get("roomId"))
                sender_name = safe_sender_name(
                    first_nonempty(
                        item.get("sender_name"), item.get("sender"),
                        item.get("truename"), item.get("fromName"),
                    ),
                    sender_userid,
                    roomid,
                )
                messages.append({
                    "msgid": item.get("msgid") or item.get("id") or "",
                    "sender_name": sender_name,
                    "sender_userid": sender_userid,
                    "sender_role": item.get("sender_role") or item.get("role") or "未知",
                    "roomid": roomid,
                    "content": item.get("content") or item.get("text") or item.get("message") or "",
                    "msgtime": str(msgtime) if msgtime else "",
                    "time_source": "analysis_payload" if msgtime else "unavailable",
                })
            if messages:
                return messages
    except Exception:
        logger.debug("dashboard.evidence.missed_messages_parse_failed", exc_info=True)

    # 2) 纯文本格式：每行 <sender>:<content>
    text_messages = _parse_missed_text(text)
    if text_messages:
        return text_messages

    # 3) 兜底：整段当作一条 content
    return [{
        "msgid": "",
        "sender_name": "",
        "sender_userid": "",
        "sender_role": "未知",
        "roomid": "",
        "content": text,
        "msgtime": "",
        "time_source": "unavailable",
    }]


def _extract_msg_times(missed_list_json: Any) -> list:
    """Extract original msgtime values from missedMessageList."""
    return [item["msgtime"] for item in _extract_missed_messages(missed_list_json) if item.get("msgtime")]


def _display_chat_message(row: dict) -> dict:
    msg_time = chat_msgtime(row)
    return {
        "msgid": chat_msgid(row),
        "sender_name": chat_sender(row),
        "sender_userid": chat_sender_userid(row),
        "sender_role": chat_sender_role(row),
        "roomid": str(row_get(row, "roomid", default="") or ""),
        "content": chat_text(row),
        "msgtime": msg_time.strftime("%Y-%m-%d %H:%M:%S") if msg_time else "",
        "time_source": "original_message" if msg_time else "unavailable",
    }


def _raw_messages_by_group(group_rows: List[dict], chat_rows: List[dict]) -> Dict[str, List[dict]]:
    grouped = _chat_rows_by_group_name(group_rows, chat_rows)
    members_by_room = {
        str(row.get("chat_id") or "").strip(): (
            row.get("member_list_json")
            or parse_json_object(row.get("completeData")).get("member_list")
            or []
        )
        for row in group_rows
        if str(row.get("chat_id") or "").strip()
    }
    result = {}
    for group_name, rows in grouped.items():
        messages = []
        for row in rows:
            enriched = dict(row)
            roomid = str(row.get("roomid") or "").strip()
            if not enriched.get("members") and members_by_room.get(roomid):
                enriched["members"] = members_by_room[roomid]
            messages.append(_display_chat_message(enriched))
        messages.sort(key=lambda item: item.get("msgtime") or "")
        result[group_name] = messages
    return result


def _normalize_chat_content(value: str) -> str:
    """把任意空白（换行/制表/多空格）压缩为单个空格，便于跨源匹配。"""
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def _match_raw_message(
    message: dict,
    raw_messages: List[dict],
    used_indices: Optional[set] = None,
) -> dict:
    """Match one analysis candidate to one original chat row.

    Matching by sender alone is deliberately forbidden: it previously copied the
    timestamp of an unrelated message and made room/user ids look like senders.
    """
    used_indices = used_indices if used_indices is not None else set()
    msgid = str(message.get("msgid") or "").strip()
    content = str(message.get("content") or "").strip()
    sender = str(message.get("sender_name") or "").strip()
    norm_content = _normalize_chat_content(content)
    candidates = []
    for index, raw in enumerate(raw_messages):
        if index in used_indices:
            continue
        if msgid and msgid == str(raw.get("msgid") or "").strip():
            candidates.append((1000, index, raw, "msgid"))
            continue
        norm_raw = _normalize_chat_content(str(raw.get("content") or ""))
        if not norm_content or not norm_raw:
            continue
        raw_sender = str(raw.get("sender_name") or "").strip()
        sender_bonus = 20 if sender and sender != "未知发送人" and sender == raw_sender else 0
        if norm_content == norm_raw:
            candidates.append((500 + sender_bonus, index, raw, "exact_content"))
            continue
        shorter, longer = sorted((norm_content, norm_raw), key=len)
        if len(shorter) >= 8 and shorter in longer and len(shorter) / len(longer) >= 0.75:
            candidates.append((300 + sender_bonus + int(100 * len(shorter) / len(longer)), index, raw, "content_fragment"))
    if not candidates:
        return {**message, "match_status": "unverified"}
    _, index, raw, match_method = max(candidates, key=lambda item: (item[0], -item[1]))
    used_indices.add(index)
    return {
        **message,
        **{key: value for key, value in raw.items() if value not in (None, "")},
        "match_status": "matched",
        "match_method": match_method,
        "raw_index": index,
    }


def _deduplicate_missed_messages(messages: List[dict]) -> List[dict]:
    result = []
    seen = set()
    for message in messages:
        key = str(message.get("msgid") or "").strip()
        if not key:
            key = "|".join((
                str(message.get("sender_name") or "").strip(),
                _normalize_chat_content(str(message.get("content") or "")),
                str(message.get("msgtime") or "").strip(),
            ))
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(message)
    return result


def _is_actionable_missed_content(value: str) -> bool:
    raw_text = str(value or "")
    if "引用/回复消息" in raw_text or raw_text.lstrip().startswith(("「", "\"")):
        parts = re.split(r"\n\s*(?:-\s*){5,}\n", raw_text)
        if len(parts) > 1:
            raw_text = parts[-1]
    text = _normalize_chat_content(raw_text).strip(" ,，。!！~～")
    if not text:
        return False
    without_mentions = re.sub(r"@[\w\u4e00-\u9fff ._-]+", "", text).strip(" ,，。!！")
    if not without_mentions:
        return False
    if re.fullmatch(r"[+\d\s()-]{6,}", without_mentions):
        return False
    if re.fullmatch(
        r"(?:(?:好的?|好滴|收到|谢谢|感谢|辛苦(?:了)?|麻烦了|知道了|ok|已加|我?加你了|[嗯哦哈]+|1)(?:[，,、\s]+)?)+",
        without_mentions,
        re.I,
    ):
        return False
    if re.fullmatch(r"(?:\[[^\]]+\]|[\W_])+", without_mentions):
        return False
    # Historical rows only become *confirmed* unanswered evidence when the text
    # contains an explicit question, request, error, complaint, or progress cue.
    # Declarative context is retained in the source table but is not promoted to
    # a dashboard alert without stronger evidence.
    return bool(re.search(
        r"[?？]|(?:吗|么|呢|是吧|对吧|怎么|如何|为什么|能否|能不能|可不可以|请问|麻烦|帮忙|帮我|协助|需要|希望|想问|想要|哪里|在哪|是否|有没有|什么时候|何时|进度|报错|错误|失败|异常|不对|打不开|无法|还没|未收到|催|尽快|链接|下载|制作|补送|沟通一下)",
        without_mentions,
        re.I,
    ))


def _evaluate_unanswered_messages(
    missed_messages: List[dict],
    raw_messages: List[dict],
    analysis_time: Optional[datetime],
) -> List[dict]:
    """Reconcile stored model output against the original timeline."""
    used_indices: set = set()
    matched = [
        _match_raw_message(message, raw_messages, used_indices)
        for message in _deduplicate_missed_messages(missed_messages)
    ]
    evaluated = []
    for message in matched:
        result = dict(message)
        result.pop("raw_index", None)
        if not _is_actionable_missed_content(str(result.get("content") or "")):
            result["verification_status"] = "no_action_needed"
            evaluated.append(result)
            continue
        if result.get("match_status") != "matched":
            result["verification_status"] = "unverified"
            evaluated.append(result)
            continue
        if result.get("sender_role") != "客户":
            result["verification_status"] = "invalid_sender"
            evaluated.append(result)
            continue
        sent_at = parse_msg_datetime(result.get("msgtime"))
        if not sent_at:
            result["verification_status"] = "unverified"
            evaluated.append(result)
            continue
        later_replies = []
        for raw in raw_messages:
            reply_at = parse_msg_datetime(raw.get("msgtime"))
            if raw.get("sender_role") in {"员工", "售后", "销售"} and reply_at and reply_at > sent_at:
                later_replies.append((reply_at, raw))
        if later_replies:
            reply_at, reply = min(later_replies, key=lambda item: item[0])
            result.update({
                "verification_status": (
                    "answered_before_analysis"
                    if analysis_time and reply_at <= analysis_time
                    else "answered_later"
                ),
                "reply_time": reply_at.strftime("%Y-%m-%d %H:%M:%S"),
                "reply_sender_name": reply.get("sender_name") or "未知员工",
            })
        else:
            result["verification_status"] = "unanswered"
        if analysis_time:
            result["analysis_time"] = analysis_time.strftime("%Y-%m-%d %H:%M:%S")
            result["waiting_minutes"] = max(0, int((analysis_time - sent_at).total_seconds() // 60))
        evaluated.append(result)
    return evaluated


def _reconciled_unanswered_summary(rows: List[dict], raw_messages_by_group: Dict[str, List[dict]]) -> dict:
    groups = set()
    review_groups = set()
    daily_groups: Dict[str, set] = defaultdict(set)
    seen = set()
    status_counts = Counter()
    for row in rows:
        if str(row.get("isMissedMessage")) != "1":
            continue
        group_name = str(row.get("groupName") or "").strip()
        if not group_name:
            continue
        analysis_time = parse_msg_datetime(row.get("CREATEDTIME"))
        evaluated = _evaluate_unanswered_messages(
            _extract_missed_messages(row.get("missedMessageList")),
            raw_messages_by_group.get(group_name, []),
            analysis_time,
        )
        for message in evaluated:
            identity = (
                group_name,
                str(message.get("msgid") or "").strip()
                or "|".join((
                    str(message.get("sender_userid") or message.get("sender_name") or "").strip(),
                    _normalize_chat_content(str(message.get("content") or "")),
                    str(message.get("msgtime") or "").strip(),
                )),
            )
            if identity in seen:
                continue
            seen.add(identity)
            status = str(message.get("verification_status") or "unverified")
            status_counts[status] += 1
            if status == "unanswered":
                groups.add(group_name)
                day = (analysis_time.strftime("%Y-%m-%d") if analysis_time else "")
                if day:
                    daily_groups[day].add(group_name)
            elif status == "unverified":
                review_groups.add(group_name)
    return {
        "groups": groups,
        "review_groups": review_groups,
        "daily_groups": daily_groups,
        "status_counts": dict(status_counts),
    }


def _evidence_messages(metric: str, content: str, keyword: str, missed_messages: List[dict], raw_messages: List[dict]) -> List[dict]:
    if metric == "unanswered":
        used_indices: set = set()
        return [
            _match_raw_message(message, raw_messages, used_indices)
            for message in _deduplicate_missed_messages(missed_messages)
        ]
    if metric == "highfreq" and keyword:
        lowered = keyword.lower()
        return [
            message for message in raw_messages
            if lowered in str(message.get("content") or "").lower()
        ][:10]
    text = str(content or "")
    if metric in ("customer_negative", "employee_negative") and text:
        return [
            message for message in raw_messages
            if message.get("content") and str(message["content"]) in text
        ][:10]
    return []


def _project_code_diagnostics(
    group_names: List[str],
    dimensions: Dict[str, dict],
    lims_records_by_code: Dict[str, List[dict]],
) -> dict:
    group_to_codes = {
        group_name: [
            str(code).strip().upper()
            for code in dimensions.get(group_name, {}).get("codes", [])
            if str(code or "").strip()
        ]
        for group_name in group_names
    }
    matched_codes = {
        str(code).strip().upper()
        for code, records in lims_records_by_code.items()
        if str(code or "").strip() and records
    }
    code_to_groups: Dict[str, List[str]] = defaultdict(list)
    for group_name, codes in group_to_codes.items():
        for code in codes:
            code_to_groups[code].append(group_name)

    requested_codes = sorted(code_to_groups)
    unmatched_codes = [code for code in requested_codes if code not in matched_codes]
    unmatched_set = set(unmatched_codes)
    groups_without_project_code = sorted(
        group_name for group_name, codes in group_to_codes.items() if not codes
    )
    groups_with_unmatched_project_code = sorted(
        group_name
        for group_name, codes in group_to_codes.items()
        if any(code in unmatched_set for code in codes)
    )
    groups_without_lims_link = sorted(
        group_name
        for group_name, codes in group_to_codes.items()
        if codes and not any(code in matched_codes for code in codes)
    )

    return {
        "group_to_codes": group_to_codes,
        "code_to_groups": {code: sorted(groups) for code, groups in code_to_groups.items()},
        "requested_codes": requested_codes,
        "matched_codes": sorted(matched_codes),
        "unmatched_codes": unmatched_codes,
        "groups_without_project_code": groups_without_project_code,
        "groups_with_unmatched_project_code": groups_with_unmatched_project_code,
        "groups_without_lims_link": groups_without_lims_link,
    }


def get_verification_stats(
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> dict:
    """Return scoped verification data for the current dashboard filters."""
    start, end, normalized_period = resolve_period(period, start_date, end_date)

    with database("dashboard.verify") as conn:
        scope = _current_dashboard_scope(conn, start, end, region, aftersaler, category, key_account)
        project_codes = scope["project_codes"]
        latest_rows = scope["latest_rows"]
        raw_rows = scope["raw_rows"]
        unanswered_rows = scope["unanswered_rows"]
        raw_unanswered_rows = scope["raw_unanswered_rows"]
        dimensions = scope["dimensions"]
        group_rows = _query_group_rows(conn, scope["group_names"])
        chat_rows = _query_chat_rows(conn, group_rows, start, end)

    lims_records_by_code, lims_stats = fetch_lims_base_data(project_codes)

    group_names = scope["group_names"]
    nightly_group_names = {
        row.get("groupName") for row in unanswered_rows if row.get("groupName")
    }
    dashboard_missed_groups = {
        row.get("groupName")
        for row in unanswered_rows
        if row.get("groupName") and str(row.get("isMissedMessage")) == "1"
    }
    qx_group_names = {str(row.get("name") or "") for row in group_rows if row.get("name")}
    qx_group_ids = {str(row.get("chat_id") or "") for row in group_rows if row.get("chat_id")}
    qx_chat_room_ids = {str(row.get("roomid") or "") for row in chat_rows if row.get("roomid")}
    qx_chat_with_msgtime = sum(1 for row in chat_rows if row.get("msgtime"))

    diagnostics = _project_code_diagnostics(group_names, dimensions, lims_records_by_code)
    lims_codes = set(lims_records_by_code)
    lims_region_codes = set()
    lims_after_codes = set()
    lims_key_codes = set()
    lims_active_codes = set()
    lims_remark_codes = set()
    for code, records in lims_records_by_code.items():
        for record in records:
            if str(_field_value_case_insensitive(record, "orgName") or "").strip():
                lims_region_codes.add(code)
            if str(_field_value_case_insensitive(record, "afterSaler") or "").strip():
                lims_after_codes.add(code)
            if _valid_key_account(_field_value_case_insensitive(record, "keyAccount")):
                lims_key_codes.add(code)
            if extract_lims_active_day(record) is not None:
                lims_active_codes.add(code)
            if str(_field_value_case_insensitive(record, "analysisSimpleRemark") or "").strip():
                lims_remark_codes.add(code)

    lims_active_groups = set()
    attention_codes_by_status: Dict[str, set] = defaultdict(set)
    for group_name, dim in dimensions.items():
        if any(parse_active_day(project.get("active_day")) is not None for project in dim.get("projects", [])):
            lims_active_groups.add(group_name)
        for project in dim.get("projects", []):
            status = str(project.get("analysis_simple_remark") or "").strip()
            if status in PROJECT_ATTENTION_STATUSES and project.get("project_code"):
                attention_codes_by_status[status].add(str(project["project_code"]).upper())

    project_total = len(project_codes)
    lims_unreturned_note = (
        "LIMS接口不可用，本项表示接口未返回，不能直接判定项目号错误。"
        if not lims_stats.get("available")
        else "项目号请求后 LIMS base_data 未返回记录，需核对项目号或 LIMS 数据。"
    )
    missing_group_rows = [
        [group_name, "群名中未匹配 LC-* 项目号"]
        for group_name in diagnostics["groups_without_project_code"]
    ] or [["(无)", "当前范围内所有群名都提取到了项目号"]]
    unmatched_code_rows = [
        [
            code,
            len(diagnostics["code_to_groups"].get(code, [])),
            "；".join(diagnostics["code_to_groups"].get(code, [])),
            lims_unreturned_note,
        ]
        for code in diagnostics["unmatched_codes"]
    ] or [["(无)", 0, "", "当前范围内提取到的项目号均有 LIMS 返回记录"]]
    unlinked_group_rows = [
        [
            group_name,
            "、".join(diagnostics["group_to_codes"].get(group_name, [])),
            "、".join(
                code for code in diagnostics["group_to_codes"].get(group_name, [])
                if code in set(diagnostics["unmatched_codes"])
            ),
            lims_unreturned_note,
        ]
        for group_name in diagnostics["groups_without_lims_link"]
    ] or [["(无)", "", "", "当前范围内带项目号的群均至少命中一个 LIMS 项目"]]
    sections = [
        {
            "title": "当前筛选范围",
            "columns": ["验证项", "范围值", "命中值", "说明"],
            "rows": [
                ["群聊数", len(group_names), len(group_names), "当前周期和筛选条件命中的群聊"],
                ["项目号数", project_total, project_total, "从命中群名提取 LC-* 项目号后去重"],
                ["分析原始记录数", len(raw_rows), len(raw_rows), "qx_analysis_result 当前范围原始记录"],
                ["看板去重记录数", len(latest_rows), len(latest_rows), "按 groupName + DATE(CREATEDTIME) 取最新记录"],
                ["夜间漏回原始记录数", len(raw_unanswered_rows), len(raw_unanswered_rows), f"{UNANSWERED_RESULT_TABLE} 当前范围原始记录"],
                ["夜间漏回去重记录数", len(unanswered_rows), len(unanswered_rows), "按 groupName + DATE(CREATEDTIME) 取夜间最新记录"],
                ["夜间分析覆盖群数", len(group_names), len(nightly_group_names), "漏回模块分母仅包含夜间表实际覆盖的群"],
                ["漏回群数", len(nightly_group_names), len(dashboard_missed_groups), f"漏回判定来自 {UNANSWERED_RESULT_TABLE}；漏回率 {_ratio(len(dashboard_missed_groups), len(nightly_group_names))}%"],
            ],
        },
        {
            "title": "项目号与 LIMS 关联诊断",
            "columns": ["验证项", "当前范围", "异常/未命中", "说明"],
            "rows": [
                ["群名项目号提取", len(group_names), len(diagnostics["groups_without_project_code"]), "异常值为无法从群名提取 LC-* 项目号的群"],
                ["LIMS未返回项目号", project_total, len(diagnostics["unmatched_codes"]), lims_unreturned_note],
                ["完全无LIMS关联群", len(group_names), len(diagnostics["groups_without_lims_link"]), "群名能提取项目号，但这些项目号均未拿到 LIMS 返回记录"],
                ["含未返回项目号群", len(group_names), len(diagnostics["groups_with_unmatched_project_code"]), "群名中至少有一个项目号未被 LIMS 返回"],
            ],
        },
        {
            "title": "未提取项目号的群名",
            "columns": ["群名", "说明"],
            "rows": missing_group_rows,
        },
        {
            "title": "LIMS未返回项目号明细",
            "columns": ["项目号", "涉及群数", "涉及群名", "说明"],
            "rows": unmatched_code_rows,
        },
        {
            "title": "完全无LIMS关联的群名",
            "columns": ["群名", "提取项目号", "未返回项目号", "说明"],
            "rows": unlinked_group_rows,
        },
        {
            "title": "数据库群聊数据源覆盖",
            "columns": ["验证项", "当前范围", "命中值", "说明"],
            "rows": [
                ["qx_analysis_result 原始记录", len(group_names), len(raw_rows), "当前周期和筛选条件内的分析结果原始记录"],
                [UNANSWERED_RESULT_TABLE + " 原始记录", len(group_names), len(raw_unanswered_rows), "当前周期和筛选条件内的每日零点漏回分析结果"],
                ["qx_group 群信息", len(group_names), len(qx_group_names), "按 qx_analysis_result.groupName = qx_group.name 匹配"],
                ["qx_chat 原始消息", len(qx_group_ids), len(qx_chat_room_ids), "按 qx_group.chat_id = qx_chat.roomid 匹配"],
                ["qx_chat 消息时间", len(chat_rows), qx_chat_with_msgtime, "原始消息 msgtime 非空数量"],
            ],
        },
        {
            "title": "LIMS base_data 接口覆盖",
            "columns": ["验证项", "当前范围项目号", "接口命中项目号", "说明"],
            "rows": [
                ["接口返回项目号", project_total, len(lims_codes), f"请求 {lims_stats.get('requests', 0)} 次，返回记录 {lims_stats.get('records', 0)} 条，错误 {lims_stats.get('errors', 0)} 次"],
                ["区域(orgName)", project_total, len(lims_region_codes), f"覆盖率 {_ratio(len(lims_region_codes), project_total)}%"],
                ["售后(afterSaler)", project_total, len(lims_after_codes), f"覆盖率 {_ratio(len(lims_after_codes), project_total)}%"],
                ["重点客户(keyAccount)", project_total, len(lims_key_codes), f"覆盖率 {_ratio(len(lims_key_codes), project_total)}%"],
                ["活跃周期(activeDay/activellay)", project_total, len(lims_active_codes), f"覆盖群聊 {len(lims_active_groups)} 个"],
                ["项目状态(analysisSimpleRemark)", project_total, len(lims_remark_codes), f"非空覆盖率 {_ratio(len(lims_remark_codes), project_total)}%"],
                ["问题项目", project_total, len(attention_codes_by_status["问题项目"]), "仅统计当前公共筛选作用域内的限定三类产品"],
                ["暂不交付", project_total, len(attention_codes_by_status["暂不交付"]), "仅统计当前公共筛选作用域内的限定三类产品"],
            ],
        },
    ]

    return {
        "period": {"start": start.isoformat(), "end": end.isoformat(), "normalized": normalized_period},
        "scope": {
            "groups": len(group_names),
            "project_codes": project_total,
            "groups_without_project_code": len(diagnostics["groups_without_project_code"]),
            "unmatched_project_codes": len(diagnostics["unmatched_codes"]),
            "groups_without_lims_link": len(diagnostics["groups_without_lims_link"]),
            "raw_analysis_rows": len(raw_rows),
            "latest_analysis_rows": len(latest_rows),
            "raw_unanswered_rows": len(raw_unanswered_rows),
            "latest_unanswered_rows": len(unanswered_rows),
            "unanswered_groups": len(nightly_group_names),
        },
        "sections": sections,
        "note": f"数据源限定为数据库 qx_analysis_result/{UNANSWERED_RESULT_TABLE}/qx_chat/qx_group 与 LIMS base_data 接口；漏回模块统一读取夜间分析表。",
    }


def get_aftersaler_mapping_preview(
    version_id: int,
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> dict:
    """Preview a selected mapping version against LIMS records in the dashboard range."""
    start, end, normalized_period = resolve_period(period, start_date, end_date)
    with database("dashboard.mapping_preview") as conn:
        rows, _ = _latest_rows(conn, start, end)
    project_codes = sorted({
        code for row in rows for code in extract_project_codes(row.get("groupName", ""))
    })
    records_by_code, lims_stats = fetch_lims_base_data(project_codes)
    snapshot = aftersaler_mapping.get_mapping_snapshot_by_version(version_id)
    scoped_records = []
    for records in records_by_code.values():
        for record in records:
            normalized = normalize_lims_api_record(record, record.get("projectCode") or "", snapshot)
            if normalized.get("category_l2") not in ALLOWED_PRODUCT_L2:
                continue
            if category and normalized.get("category_l2") != category:
                continue
            if region and normalized.get("region") != region:
                continue
            if key_account and normalized.get("key_account") != key_account:
                continue
            if aftersaler and normalized.get("final_aftersaler") != aftersaler:
                continue
            scoped_records.append(record)
    result = aftersaler_mapping.preview_records(scoped_records, snapshot)
    result.update({
        "period": {"start": start.isoformat(), "end": end.isoformat(), "normalized": normalized_period},
        "project_codes": len(project_codes),
        "lims_api_records": lims_stats.get("records", 0),
        "lims_api_errors": lims_stats.get("errors", 0),
    })
    return result


def get_unanswered_summary() -> dict:
    return get_overview("year")["service_quality"]["unanswered"]


def _query_by_chunks(
    conn,
    operation: str,
    sql_template: str,
    values: List[Any],
    prefix_params: Optional[List[Any]] = None,
    suffix_params: Optional[List[Any]] = None,
    chunk_size: int = 100,
) -> List[dict]:
    if not values:
        return []
    rows: List[dict] = []
    prefix_params = prefix_params or []
    suffix_params = suffix_params or []
    for chunk in _chunked(values, chunk_size):
        placeholders = ",".join(["%s"] * len(chunk))
        rows.extend(_query(conn, operation, sql_template.format(placeholders=placeholders), prefix_params + chunk + suffix_params))
    return rows


def _current_dashboard_scope(
    conn,
    start: date,
    end: date,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> dict:
    latest_rows, raw_count = _latest_rows(conn, start, end)
    mapping_snapshot = aftersaler_mapping.get_mapping_snapshot(end, conn=conn)
    dimensions, quality = _load_dimensions(conn, latest_rows, mapping_snapshot)
    filter_options = _dashboard_filter_options(dimensions)
    allowed_groups = {
        group_name
        for group_name, dim in dimensions.items()
        if _dimension_matches_or_degrades(dim, region, aftersaler, category, key_account)
    }
    latest_rows = [row for row in latest_rows if row.get("groupName") in allowed_groups]
    dimensions = {
        name: _scope_dimension(value, region, category, key_account, aftersaler)
        for name, value in dimensions.items()
        if name in allowed_groups
    }
    quality = _quality_from_dimensions(dimensions, quality)

    group_names = sorted({row.get("groupName") for row in latest_rows if row.get("groupName")})
    raw_count = _raw_analysis_count(conn, start, end, group_names)
    unanswered_rows = _latest_unanswered_rows(conn, start, end, group_names)
    raw_unanswered_rows = _raw_unanswered_rows(conn, start, end, group_names)
    project_codes = sorted({
        code.upper()
        for dim in dimensions.values()
        for code in dim.get("codes", [])
        if code
    })
    raw_rows = _query_by_chunks(
        conn,
        "scope.analysis_raw",
        """SELECT * FROM qx_analysis_result
           WHERE CREATEDTIME >= %s AND CREATEDTIME < %s
             AND groupName IN ({placeholders})
           ORDER BY CREATEDTIME""",
        group_names,
        prefix_params=[start.isoformat(), (end + timedelta(days=1)).isoformat()],
    ) if group_names else []
    return {
        "latest_rows": latest_rows,
        "unanswered_rows": unanswered_rows,
        "raw_rows": raw_rows,
        "raw_unanswered_rows": raw_unanswered_rows,
        "raw_count_before_filter": raw_count,
        "dimensions": dimensions,
        "quality": quality,
        "filter_options": filter_options,
        "group_names": group_names,
        "project_codes": project_codes,
    }


def _query_group_rows(conn, group_names: List[str]) -> List[dict]:
    return _query_by_chunks(
        conn,
        "raw.qx_group",
        "SELECT * FROM qx_group WHERE name IN ({placeholders})",
        group_names,
    )


def _query_chat_rows(
    conn,
    group_rows: List[dict],
    start: Optional[date] = None,
    end: Optional[date] = None,
) -> List[dict]:
    room_ids = sorted({
        str(row.get("chat_id") or "").strip()
        for row in group_rows
        if str(row.get("chat_id") or "").strip()
    })
    if not room_ids:
        return []
    rows_by_key: Dict[str, dict] = {}
    sql = "SELECT * FROM qx_chat WHERE roomid IN ({placeholders})"
    suffix_params: List[Any] = []
    if start is not None and end is not None:
        sql += " AND msgtime >= %s AND msgtime < %s"
        suffix_params = [start.isoformat(), (end + timedelta(days=1)).isoformat()]
    sql += " ORDER BY msgtime"
    for row in _query_by_chunks(
        conn,
        "raw.qx_chat",
        sql,
        room_ids,
        suffix_params=suffix_params,
    ):
        key = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
        rows_by_key[key] = row
    return list(rows_by_key.values())


def _valid_key_account(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return bool(text and text not in ("0", "false", "no", "null", "none", "否", "无", "?"))


def _query_qx_raw_scope(
    conn,
    start: date,
    end: date,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> dict:
    scope = _current_dashboard_scope(conn, start, end, region, aftersaler, category, key_account)
    group_rows = _query_group_rows(conn, scope["group_names"])
    audit_chat_rows = _query_chat_rows(
        conn, group_rows, start - timedelta(days=2), end + timedelta(days=1)
    )
    chat_rows = _filter_chat_rows_by_period(audit_chat_rows, start, end)
    return {
        **scope,
        "group_rows": group_rows,
        "chat_rows": chat_rows,
        "chat_rows_by_group": _raw_messages_by_group(group_rows, chat_rows),
        "audit_chat_rows": audit_chat_rows,
        "audit_raw_messages": _raw_messages_by_group(group_rows, audit_chat_rows),
    }


def _project_code_set(rows: List[dict], *columns: str) -> set:
    values = set()
    for row in rows:
        for column in columns:
            value = row_get(row, column, default=None)
            if value:
                values.add(str(value).upper())
    return values


_DRILLDOWN_COLUMN_LABELS = {
    "name": "名称", "region": "销售区域", "aftersaler": "售后人员",
    "category": "产品分类", "key_account": "重点客户", "work_unit": "客户单位",
    "customer_name": "客户名称", "group_name": "群聊名称", "group_count": "群聊数",
    "project_count": "项目数", "message_count": "消息数", "project_code": "项目号",
    "project_codes": "项目号", "project_name": "项目/产品", "product_name": "项目/产品",
    "category_l2": "产品二级分类", "category_l3": "产品三级分类",
    "sales_person": "销售员", "aftersalers": "售后人员", "regions": "销售区域",
    "key_accounts": "重点客户", "active_day": "活跃天数", "active_days": "活跃天数",
    "analysis_time": "分析时间", "analysis_date": "分析日期", "analysis_id": "分析记录ID",
    "msgtime": "消息时间", "sender_name": "发送人", "sender_role": "发送者角色",
    "content": "消息内容", "status": "项目状态", "start_time": "开始时间",
    "end_time": "结束时间", "year": "项目年份", "year_source": "年份来源",
    "source": "数据来源", "mapping_source": "售后确定来源", "raw_aftersaler": "LIMS 原售后",
    "mapping_conflict": "映射冲突",
    "final_aftersaler": "最终售后", "days": "统计天数", "missed_days": "待回复分析天数",
    "high_frequency_top5": "高频主题 Top5", "metric_value": "指标值",
}


def _mask_drilldown_text(value: Any) -> str:
    """Mask contact details before a message reaches preview or Excel."""
    text = str(value or "")
    text = re.sub(r"(?<!\d)(1\d{2})\d{4}(\d{4})(?!\d)", r"\1****\2", text)
    text = re.sub(
        r"(?i)\b(wxid_[a-z0-9_-]{4})[a-z0-9_-]+\b",
        lambda match: match.group(1) + "***",
        text,
    )
    return text


def _joined(values: Iterable[Any]) -> str:
    return "、".join(sorted({str(value).strip() for value in values if str(value or "").strip()}))


def _dimension_project_values(dimension: dict, key: str) -> set[str]:
    return {
        str(project.get(key) or "").strip()
        for project in dimension.get("projects", [])
        if str(project.get(key) or "").strip()
    }


def _drilldown_group_rows(scope: dict, groups: Dict[str, dict]) -> List[dict]:
    result = []
    for group_name in scope.get("group_names", []):
        dimension = scope.get("dimensions", {}).get(group_name, {})
        aggregate = groups.get(group_name, {})
        projects = dimension.get("projects", [])
        analysis_times = [
            str(row.get("CREATEDTIME") or "")[:19].replace("T", " ")
            for row in aggregate.get("rows", [])
            if row.get("CREATEDTIME")
        ]
        result.append({
            "group_name": group_name,
            "project_codes": _joined(dimension.get("codes", [])),
            "project_count": len({p.get("project_code") for p in projects if p.get("project_code")}),
            "regions": _joined(dimension.get("regions", [])),
            "aftersalers": _joined(dimension.get("aftersalers", [])),
            "category_l2": _joined(p.get("category_l2") for p in projects),
            "key_accounts": _joined(p.get("key_account") for p in projects),
            "message_count": int(aggregate.get("messages") or 0),
            "days": len(aggregate.get("dates", set())),
            "missed_days": int(aggregate.get("missed_days") or 0),
            "analysis_time": max(analysis_times, default=""),
            "source": f"qx_analysis_result + {UNANSWERED_RESULT_TABLE} + qx_group + LIMS",
        })
    return result


def _drilldown_project_rows(scope: dict, groups: Dict[str, dict]) -> List[dict]:
    projects: Dict[str, dict] = {}
    for group_name, dimension in scope.get("dimensions", {}).items():
        aggregate = groups.get(group_name, {})
        for index, project in enumerate(dimension.get("projects", [])):
            code = str(project.get("project_code") or "").strip().upper()
            identity = code or f"{group_name}:{index}:{project.get('product_name') or ''}"
            item = projects.setdefault(identity, {
                "project_code": code,
                "project_name": project.get("product_name") or "",
                "customer_name": project.get("customer_name") or "",
                "work_unit": project.get("work_unit") or "",
                "category_l2": project.get("category_l2") or "",
                "category_l3": project.get("category_l3") or "",
                "region": project.get("region") or "",
                "sales_person": project.get("sales_person") or "",
                "raw_aftersaler": project.get("raw_aftersaler") or "",
                "final_aftersaler": project.get("final_aftersaler") or project.get("raw_aftersaler") or "",
                "mapping_source": project.get("aftersaler_source") or "",
                "mapping_conflict": bool(project.get("mapping_conflict")),
                "key_account": project.get("key_account") or "",
                "status": project.get("analysis_simple_remark") or "",
                "active_day": project.get("active_day"),
                "start_time": project.get("start_time") or "",
                "end_time": project.get("end_time") or "",
                "group_names": set(),
                "message_count": 0,
                "source": "LIMS base_data API",
            })
            if group_name not in item["group_names"]:
                item["group_names"].add(group_name)
                item["message_count"] += int(aggregate.get("messages") or 0)
    result = []
    for item in projects.values():
        row = dict(item)
        row["group_count"] = len(item["group_names"])
        row["group_name"] = _joined(item["group_names"])
        row.pop("group_names", None)
        result.append(row)
    return sorted(result, key=lambda row: (row.get("project_code") or "", row.get("project_name") or ""))


def _drilldown_message_rows(scope: dict) -> List[dict]:
    grouped = _raw_messages_by_group(scope.get("group_rows", []), scope.get("chat_rows", []))
    dimensions = scope.get("dimensions", {})
    result = []
    for group_name, messages in grouped.items():
        dimension = dimensions.get(group_name, {})
        projects = dimension.get("projects", [])
        for message in messages:
            result.append({
                "msgtime": message.get("msgtime") or "",
                "group_name": group_name,
                "sender_name": _mask_drilldown_text(message.get("sender_name") or "未知发送人"),
                "sender_role": message.get("sender_role") or "未知",
                "content": _mask_drilldown_text(message.get("content") or ""),
                "project_codes": _joined(dimension.get("codes", [])),
                "regions": _joined(dimension.get("regions", [])),
                "aftersalers": _joined(dimension.get("aftersalers", [])),
                "category_l2": _joined(p.get("category_l2") for p in projects),
                "key_accounts": _joined(p.get("key_account") for p in projects),
                "source": "qx_chat",
            })
    return sorted(result, key=lambda row: (row.get("msgtime") or "", row.get("group_name") or ""), reverse=True)


def _drilldown_analysis_rows(scope: dict) -> List[dict]:
    dimensions = scope.get("dimensions", {})
    result = []
    for row in scope.get("raw_rows", []):
        group_name = row.get("groupName") or ""
        dimension = dimensions.get(group_name, {})
        result.append({
            "analysis_id": row.get("id"),
            "analysis_time": str(row.get("CREATEDTIME") or "")[:19].replace("T", " "),
            "group_name": group_name,
            "message_count": int(row.get("messageToDayCount") or 0),
            "project_codes": _joined(dimension.get("codes", [])),
            "regions": _joined(dimension.get("regions", [])),
            "aftersalers": _joined(dimension.get("aftersalers", [])),
            "customer_emotion": row.get("customerEmotionAnalysis") or "",
            "employee_emotion": row.get("saleEmotionAnalysis") or "",
            "high_frequency_top5": row.get("highFrequencyWords") or "",
            "source": "qx_analysis_result",
            "_raw": row,
        })
    return result


def _duration_bucket(value: Optional[float]) -> str:
    if value is None:
        return "未返回"
    if value <= 7:
        return "≤7天"
    if value <= 30:
        return "8-30天"
    if value <= 90:
        return "1-3个月"
    if value <= 180:
        return "3-6个月"
    if value <= 365:
        return "6-12个月"
    return ">12个月"


def _message_in_period_bucket(row: dict, bucket: str) -> bool:
    msg_time = parse_msg_datetime(row.get("msgtime"))
    if not msg_time:
        return False
    minutes = msg_time.hour * 60 + msg_time.minute
    if bucket == "weekend":
        return msg_time.weekday() >= 5
    if msg_time.weekday() >= 5:
        return False
    if bucket == "morning":
        return 8 * 60 + 30 <= minutes < 12 * 60
    if bucket == "afternoon":
        return 12 * 60 <= minutes < 17 * 60 + 30
    if bucket == "after_hours":
        return minutes < 8 * 60 + 30 or minutes >= 17 * 60 + 30
    return bucket == "total"


def _evidence_drilldown_rows(data: dict) -> List[dict]:
    rows = []
    for item in data.get("items", []):
        base = {
            "group_name": item.get("group_name") or "",
            "analysis_time": item.get("analysis_time") or item.get("analysis_date") or "",
            "project_codes": _joined(item.get("project_codes", [])),
            "aftersalers": _joined(item.get("aftersalers", [])),
            "source": f"{UNANSWERED_RESULT_TABLE} + qx_chat",
        }
        messages = item.get("messages") or []
        if messages:
            for message in messages:
                rows.append({
                    "msgtime": message.get("msgtime") or "",
                    **base,
                    "sender_name": _mask_drilldown_text(message.get("sender_name") or "未知发送人"),
                    "sender_role": message.get("sender_role") or "未知",
                    "content": _mask_drilldown_text(message.get("content") or ""),
                })
        else:
            rows.append({**base, "content": _mask_drilldown_text(item.get("content") or "")})
    return rows


def _group_matches_drilldown(
    target: str, dimension: dict, dimension_value: str, secondary_value: str,
) -> bool:
    projects = dimension.get("projects", [])
    if target == "business.region":
        return dimension_value in dimension.get("regions", [])
    if target == "business.aftersaler":
        return dimension_value in dimension.get("aftersalers", [])
    if target == "business.product":
        return any(project.get("category_l2") == dimension_value for project in projects)
    if target == "business.key_account":
        return any(project.get("key_account") == dimension_value for project in projects)
    if target == "business.project_attention":
        return any(
            project.get("analysis_simple_remark") == dimension_value
            and (not secondary_value or project.get("project_code") == secondary_value)
            for project in projects
        )
    if target.startswith("cross."):
        for project in projects:
            if project.get("region") != dimension_value:
                continue
            if not secondary_value:
                return True
            if target == "cross.region_sales" and project.get("sales_person") == secondary_value:
                return True
            if target == "cross.region_product" and project.get("category_l2") == secondary_value:
                return True
        if target == "cross.region_after":
            return dimension_value in dimension.get("regions", []) and secondary_value in dimension.get("aftersalers", [])
        return False
    return True


def _expected_drilldown_value(
    overview: dict, target: str, measure: str, dimension_value: str, secondary_value: str,
) -> Optional[int]:
    summary_map = {
        "summary.total_groups": "total_groups", "summary.total_messages": "total_messages",
        "summary.project_groups": "project_groups", "summary.regions": "regions",
        "summary.aftersalers": "aftersaler_count", "summary.product_categories": "product_categories",
        "summary.key_accounts": "key_accounts",
    }
    if target in summary_map:
        return int(overview.get("summary", {}).get(summary_map[target]) or 0)
    if target == "communication.daily":
        key = {"message_count": "messages", "group_count": "groups", "missed_count": "missed"}[measure]
        item = next((row for row in overview["communication"]["trend"] if row.get("date") == dimension_value), {})
        return int(item.get(key) or 0)
    if target == "communication.period":
        item = next((row for row in overview["communication"]["time_period_breakdown"].get("items", []) if row.get("aftersaler") == dimension_value), {})
        value = item.get(measure, 0)
        return int(value.get("count") if isinstance(value, dict) else value or 0)
    if target == "communication.top_group":
        item = next((row for row in overview["communication"]["top_message_groups"].get("items", []) if row.get("group_name") == dimension_value), {})
        return int(item.get("message_count") or 0)
    if target == "communication.active_duration":
        item = next((row for row in overview["communication"]["active_duration"] if row.get("range") == dimension_value), {})
        return int(item.get("count") or 0)
    if target == "communication.highfreq":
        item = next((row for row in overview["communication"]["high_frequency"] if row.get("word") == dimension_value), {})
        return int(item.get("count") or 0)
    business_key = {
        "business.region": ("regions", "region"), "business.aftersaler": ("aftersalers", "name"),
        "business.product": ("product_categories", "category"),
        "business.key_account": ("key_accounts", "key_account"),
    }.get(target)
    if business_key:
        collection, identity = business_key
        item = next((row for row in overview["business"].get(collection, []) if str(row.get(identity) or "") == dimension_value), {})
        return int(item.get(measure) or item.get("count") or 0)
    if target == "business.project_year":
        item = next((row for row in overview["business"].get("project_year_distribution", {}).get("items", []) if str(row.get("year") if row.get("year") is not None else "unknown") == dimension_value), {})
        return int(item.get("project_count") or 0)
    if target == "business.project_attention":
        items = overview.get("project_attention", {}).get("items", [])
        if secondary_value:
            item = next((row for row in items if row.get("project_code") == secondary_value), {})
            return int(item.get(measure) or (1 if measure == "project_count" and item else 0))
        summary = next((row for row in overview.get("project_attention", {}).get("summary", []) if row.get("status") == dimension_value), {})
        return int(summary.get(measure) or 0)
    if target.startswith("cross."):
        key = target.split(".", 1)[1]
        item = next((row for row in overview.get("cross_analysis", {}).get(key, []) if row.get("region") == dimension_value), {})
        if secondary_value and measure == "group_count":
            subkey = "category" if target == "cross.region_product" else "name"
            detail = next((row for row in item.get("items", []) if row.get(subkey) == secondary_value), {})
            return int(detail.get("count") or 0)
        return int(item.get(measure) or 0)
    service_map = {
        "service.unanswered": ("unanswered", "missed_groups"),
        "service.answered": ("unanswered", "answered_groups"),
        "service.customer_positive": ("sentiment", "customer_good"),
        "service.customer_negative": ("sentiment", "customer_bad"),
        "service.employee_positive": ("sentiment", "employee_positive"),
        "service.employee_negative": ("sentiment", "employee_negative"),
    }
    if target in service_map:
        section, key = service_map[target]
        return int(overview.get("service_quality", {}).get(section, {}).get(key) or 0)
    if target.startswith("quality."):
        return int(overview.get("data_quality", {}).get(target.split(".", 1)[1]) or 0)
    return None


def _drilldown_columns(rows: List[dict]) -> List[dict]:
    preferred = [
        "msgtime", "analysis_time", "group_name", "project_code", "project_codes",
        "project_name", "customer_name", "work_unit", "region", "regions", "sales_person",
        "aftersaler", "aftersalers", "raw_aftersaler", "final_aftersaler", "mapping_source",
        "category_l2", "category_l3", "key_account", "key_accounts", "status", "sender_name",
        "sender_role", "content", "project_count", "group_count", "message_count", "active_day",
        "days", "missed_days", "start_time", "end_time", "year", "year_source", "source",
    ]
    available = {key for row in rows for key in row if not key.startswith("_")}
    ordered = [key for key in preferred if key in available]
    ordered.extend(sorted(available - set(ordered)))
    return [{"key": key, "label": _DRILLDOWN_COLUMN_LABELS.get(key, key)} for key in ordered]


def _drilldown_definition(target: str, level: str, measure: str) -> str:
    units = {
        "group": "一条明细代表一个去重后的群聊。",
        "message": "一条明细代表一条 qx_chat 原始消息，正文中的手机号和用户标识已脱敏。",
        "project": "一条明细代表一个唯一项目号；无项目号记录按群聊和产品区分。",
        "analysis": "一条明细代表一条 qx_analysis_result 分析记录。",
        "dimension": "一条明细代表一个去重后的业务维度值。",
        "project_code": "一条明细代表一个从群名提取但未被 LIMS 返回的项目号。",
    }
    note = units.get(level, "一条明细代表当前指标的一条核对记录。")
    if measure == "message_count":
        if target in {"business.aftersaler", "business.product"}:
            note += " 看板值与明细核对值均由本次请求中的同一批 qx_chat 原始消息计算。"
        else:
            note += " 看板消息数来自分析汇总，明细核对值来自原始消息，差异会原样显示。"
    if target == "communication.highfreq":
        note += " 看板值是词频累计，明细核对值是涉及的证据记录数，两者统计单位不同。"
    return note


def _build_drilldown(
    target: str,
    measure: str,
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    dimension_value: str = "",
    secondary_value: str = "",
    search: str = "",
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
    snapshot_id: str = "",
) -> dict:
    spec = DRILLDOWN_TARGETS.get(target)
    if not spec:
        raise ValueError("unsupported drilldown target")
    if measure not in spec["measures"]:
        raise ValueError("unsupported drilldown measure for target")
    start, end, normalized_period = resolve_period(period, start_date, end_date)
    scope_key = _dashboard_snapshot_scope_key(
        start, end, normalized_period, region, aftersaler, category, key_account,
    )
    scope, snapshot_lookup = _get_dashboard_snapshot(snapshot_id, scope_key)
    snapshot_status = "hit" if scope is not None else ("rebuilt" if snapshot_id else "not_requested")
    if scope is None:
        if snapshot_id:
            logger.warning(
                "dashboard.snapshot.rebuild target={} token_prefix={} period={} start={} end={}",
                target,
                snapshot_id[:8],
                normalized_period,
                start,
                end,
            )
        with database("dashboard.drilldown") as conn:
            scope = _query_qx_raw_scope(conn, start, end, region, aftersaler, category, key_account)
        if snapshot_id and snapshot_lookup in {"missing", "expired"}:
            _store_dashboard_snapshot(scope_key, scope, token=snapshot_id)
    overview = _build_overview(
        start=start, end=end, period=normalized_period,
        region=region, aftersaler=aftersaler, category=category, key_account=key_account,
        snapshot=scope,
    )

    evidence_metric = {
        "service.unanswered": "unanswered", "service.customer_negative": "customer_negative",
        "service.employee_negative": "employee_negative", "communication.highfreq": "highfreq",
    }.get(target)
    if target == "communication.daily" and measure == "missed_count":
        evidence_metric = "unanswered"
    if evidence_metric:
        evidence_start = dimension_value if target == "communication.daily" else start.isoformat()
        evidence_end = dimension_value if target == "communication.daily" else end.isoformat()
        evidence = _build_evidence(
            metric=evidence_metric, period="custom", start_date=evidence_start, end_date=evidence_end,
            keyword=dimension_value if target == "communication.highfreq" else None,
            region=region, aftersaler=aftersaler, category=category, key_account=key_account,
            page=1, page_size=DASHBOARD_DRILLDOWN_EXPORT_MAX_ROWS + 1,
        )
        rows = _evidence_drilldown_rows(evidence)
        level = "message"
    else:
        groups = _group_aggregates(
            scope.get("latest_rows", []),
            scope.get("dimensions", {}),
            scope.get("unanswered_rows", []),
        )
        group_rows = _drilldown_group_rows(scope, groups)
        project_rows = _drilldown_project_rows(scope, groups)
        message_rows = _drilldown_message_rows(scope)
        analysis_rows = _drilldown_analysis_rows(scope)
        selected_groups = set(scope.get("group_names", []))

        if target in {"business.region", "business.aftersaler", "business.product", "business.key_account", "business.project_attention"} or target.startswith("cross."):
            selected_groups = {
                name for name, dim in scope.get("dimensions", {}).items()
                if _group_matches_drilldown(target, dim, dimension_value, secondary_value)
            }
        if target == "summary.project_groups":
            selected_groups = {
                name for name, dim in scope.get("dimensions", {}).items() if dim.get("codes")
            }
        if target == "service.answered":
            unanswered = _build_evidence(
                metric="unanswered", period=period, start_date=start_date, end_date=end_date,
                region=region, aftersaler=aftersaler, category=category, key_account=key_account,
                page=1, page_size=DASHBOARD_DRILLDOWN_EXPORT_MAX_ROWS + 1,
            )
            missed_group_names = {item.get("group_name") for item in unanswered.get("items", [])}
            analyzed_group_names = {
                row.get("groupName")
                for row in scope.get("unanswered_rows", [])
                if row.get("groupName")
            }
            selected_groups = analyzed_group_names - missed_group_names
        if target == "quality.groups_without_project_code":
            selected_groups = {
                name for name, dim in scope.get("dimensions", {}).items() if not dim.get("codes")
            }
        if target == "quality.groups_without_lims_link":
            selected_groups = {
                name for name, dim in scope.get("dimensions", {}).items() if not dim.get("projects")
            }
        if target == "communication.top_group":
            selected_groups = {dimension_value}
        if target == "communication.daily":
            selected_groups = {
                row.get("groupName") for row in scope.get("latest_rows", [])
                if str(row.get("CREATEDTIME") or "")[:10] == dimension_value
            }
        if target == "communication.active_duration":
            durations = _active_durations_from_lims(scope.get("dimensions", {}))
            selected_groups = {name for name, value in durations.items() if _duration_bucket(value) == dimension_value}

        if target.startswith("summary.") and spec["level"] == "dimension":
            dimension_sources = {
                "summary.regions": (overview["business"].get("regions", []), "region"),
                "summary.aftersalers": (overview["business"].get("aftersalers", []), "name"),
                "summary.product_categories": (overview["business"].get("product_categories", []), "category"),
                "summary.key_accounts": (overview["business"].get("key_accounts", []), "key_account"),
            }
            source, identity = dimension_sources[target]
            rows = [{"name": row.get(identity), **row, "source": "看板业务维度汇总"} for row in source]
            level = "dimension"
        elif target == "business.project_year":
            year_item = next((item for item in overview["business"].get("project_year_distribution", {}).get("items", []) if str(item.get("year") if item.get("year") is not None else "unknown") == dimension_value), {})
            rows = []
            for project in year_item.get("projects", []):
                rows.append({
                    "year": project.get("year"), "project_code": project.get("project_code"),
                    "year_source": project.get("year_source"), "start_time": project.get("start_time"),
                    "project_name": _joined(project.get("project_names", [])),
                    "customer_name": _joined(project.get("customer_names", [])),
                    "work_unit": _joined(project.get("work_units", [])),
                    "category_l2": _joined(project.get("product_categories", [])),
                    "region": _joined(project.get("regions", [])),
                    "aftersalers": _joined(project.get("aftersalers", [])),
                    "group_count": project.get("group_count"),
                    "group_name": _joined(project.get("group_names", [])), "source": "LIMS + 项目号年份",
                })
            level = "project"
        elif target == "quality.unmatched_project_codes":
            rows = []
            for group_name, dim in scope.get("dimensions", {}).items():
                matched = {str(p.get("project_code") or "").upper() for p in dim.get("projects", [])}
                for code in dim.get("codes", []):
                    if str(code).upper() not in matched:
                        rows.append({"project_code": str(code).upper(), "group_name": group_name, "source": "群名提取 / LIMS 未返回"})
            unique = {}
            for row in rows:
                unique.setdefault(row["project_code"], row)
            rows = list(unique.values())
            level = "project_code"
        elif target in {"quality.mapping_fallback_records", "quality.mapping_conflict_records"}:
            if target.endswith("fallback_records"):
                rows = [
                    row for row in project_rows
                    if row.get("mapping_source") == "lims_fallback" and not row.get("mapping_conflict")
                ]
            else:
                rows = [row for row in project_rows if row.get("mapping_conflict")]
            level = "project"
        elif target == "quality.duplicate_rows_removed":
            kept_ids = {str(row.get("id")) for row in scope.get("latest_rows", []) if row.get("id") is not None}
            rows = [row for row in analysis_rows if str(row.get("analysis_id")) not in kept_ids]
            level = "analysis"
        elif target in {"service.customer_positive", "service.employee_positive"}:
            rows = []
            latest_ids = {str(item.get("id")) for item in scope.get("latest_rows", []) if item.get("id") is not None}
            for row in analysis_rows:
                if latest_ids and str(row.get("analysis_id")) not in latest_ids:
                    continue
                raw = row.get("_raw", {})
                mapping = parse_emotion_field(raw.get("customerEmotionAnalysis") if target == "service.customer_positive" else raw.get("saleEmotionAnalysis"))
                labels = ("好评", "正向", "满意") if target == "service.customer_positive" else ("积极", "正向")
                value = _emotion_total(mapping, labels)
                if value > 0:
                    rows.append({**row, "metric_value": value})
            level = "analysis"
        elif target == "communication.period":
            selected_groups = {
                name for name, dim in scope.get("dimensions", {}).items()
                if dimension_value in dim.get("aftersalers", [])
            }
            rows = []
            messages_by_group = defaultdict(list)
            for row in message_rows:
                messages_by_group[row.get("group_name")].append(row)
            for group_name in selected_groups:
                owners = scope.get("dimensions", {}).get(group_name, {}).get("aftersalers", [])
                candidates = messages_by_group.get(group_name, [])
                if len(owners) > 1:
                    candidates = [row for row in candidates if _sender_matches_aftersaler(row, dimension_value)]
                rows.extend(row for row in candidates if _message_in_period_bucket(row, measure))
            level = "message"
        else:
            if measure == "project_count":
                rows = [row for row in project_rows if any(name in selected_groups for name in str(row.get("group_name") or "").split("、"))]
                if target == "business.region":
                    rows = [row for row in rows if row.get("region") == dimension_value]
                elif target == "business.aftersaler":
                    rows = [row for row in rows if row.get("final_aftersaler") == dimension_value]
                elif target == "business.product":
                    rows = [row for row in rows if row.get("category_l2") == dimension_value]
                elif target == "business.key_account":
                    rows = [row for row in rows if row.get("key_account") == dimension_value]
                elif target == "business.project_attention":
                    rows = [
                        row for row in rows
                        if row.get("status") == dimension_value
                        and (not secondary_value or row.get("project_code") == secondary_value)
                    ]
                level = "project"
            elif measure == "message_count":
                rows = [row for row in message_rows if row.get("group_name") in selected_groups]
                level = "message"
                if target == "communication.daily":
                    rows = [row for row in rows if str(row.get("msgtime") or "")[:10] == dimension_value]
                if target == "business.aftersaler":
                    filtered = []
                    for row in rows:
                        owners = scope.get("dimensions", {}).get(row.get("group_name"), {}).get("aftersalers", [])
                        if len(owners) <= 1 or _sender_matches_aftersaler(row, dimension_value):
                            filtered.append(row)
                    rows = filtered
            else:
                rows = [row for row in group_rows if row.get("group_name") in selected_groups]
                level = "group"

    for row in rows:
        row.pop("_raw", None)
    if search:
        term = search.casefold().strip()
        rows = [
            row for row in rows
            if term in json.dumps(row, ensure_ascii=False, default=str).casefold()
        ]
    expected = _expected_drilldown_value(overview, target, measure, dimension_value, secondary_value)
    if measure in {"group_count", "missed_count"}:
        detail_value = len({row.get("group_name") for row in rows if row.get("group_name")})
    elif measure == "project_count":
        identities = {
            row.get("project_code") or f"{row.get('group_name')}:{row.get('project_name')}"
            for row in rows
        }
        detail_value = len(identities)
    elif target in {"service.customer_positive", "service.employee_positive"}:
        detail_value = sum(int(row.get("metric_value") or 0) for row in rows)
    else:
        detail_value = len(rows)
    non_comparable_targets = {
        "communication.highfreq", "service.customer_negative", "service.employee_negative",
        "quality.mapping_fallback_records", "quality.mapping_conflict_records",
    }
    comparable = target not in non_comparable_targets and not bool(search.strip())
    consistent = (expected == detail_value) if expected is not None and comparable else None
    if search.strip():
        reconciliation_note = "已应用抽屉内搜索，当前为看板全集的子集，不直接判断一致性。"
    elif comparable:
        reconciliation_note = "统计单位一致，可直接核对。"
    else:
        reconciliation_note = "看板汇总值与当前证据明细统计单位不同，仅供追溯。"
    return {
        "target": target, "measure": measure, "title": spec["title"], "level": level,
        "snapshot_status": snapshot_status,
        "definition": _drilldown_definition(target, level, measure),
        "period": {"name": normalized_period, "start": start.isoformat(), "end": end.isoformat()},
        "filters": {
            "region": region, "aftersaler": aftersaler, "category": category,
            "key_account": key_account, "dimension_value": dimension_value,
            "secondary_value": secondary_value, "search": search,
        },
        "reconciliation": {
            "dashboard_value": expected, "detail_value": detail_value,
            "comparable": comparable, "consistent": consistent,
            "note": reconciliation_note,
        },
        "columns": _drilldown_columns(rows), "rows": rows,
    }


def get_drilldown(
    target: str,
    measure: str,
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    dimension_value: str = "",
    secondary_value: str = "",
    search: str = "",
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
    snapshot_id: str = "",
    page: int = 1,
    page_size: int = 20,
) -> dict:
    if page < 1:
        raise ValueError("page must be greater than or equal to 1")
    if page_size < 1 or page_size > 100:
        raise ValueError("page_size must be between 1 and 100")
    data = _build_drilldown(
        target=target, measure=measure, period=period, start_date=start_date, end_date=end_date,
        dimension_value=dimension_value, secondary_value=secondary_value, search=search,
        region=region, aftersaler=aftersaler, category=category, key_account=key_account,
        snapshot_id=snapshot_id,
    )
    rows = data.pop("rows")
    total = len(rows)
    offset = (page - 1) * page_size
    return {
        **data, "total": total, "page": page, "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
        "items": rows[offset:offset + page_size],
    }


def get_drilldown_export_excel(**kwargs) -> bytes:
    """Export exactly the same scoped rows used by the drilldown preview."""
    from io import BytesIO
    from openpyxl import Workbook

    data = _build_drilldown(**kwargs)
    rows = data.pop("rows")
    if len(rows) > DASHBOARD_DRILLDOWN_EXPORT_MAX_ROWS:
        raise ValueError(
            f"明细共 {len(rows)} 行，超过单次导出上限 {DASHBOARD_DRILLDOWN_EXPORT_MAX_ROWS} 行，请缩小筛选范围"
        )
    workbook = Workbook()
    workbook.remove(workbook.active)
    recon = data["reconciliation"]
    filters = data["filters"]
    _write_excel_sheet(workbook, "导出说明", [
        {"项目": "明细标题", "内容": data["title"]},
        {"项目": "统计口径", "内容": data["definition"]},
        {"项目": "开始日期", "内容": data["period"]["start"]},
        {"项目": "结束日期", "内容": data["period"]["end"]},
        {"项目": "销售区域", "内容": filters.get("region") or "全部"},
        {"项目": "售后人员", "内容": filters.get("aftersaler") or "全部"},
        {"项目": "产品分类", "内容": filters.get("category") or "全部"},
        {"项目": "重点客户", "内容": filters.get("key_account") or "全部"},
        {"项目": "点击维度", "内容": filters.get("dimension_value") or "-"},
        {"项目": "次级维度", "内容": filters.get("secondary_value") or "-"},
        {"项目": "搜索条件", "内容": filters.get("search") or "-"},
        {"项目": "看板值", "内容": recon.get("dashboard_value")},
        {"项目": "导出明细数", "内容": len(rows)},
        {"项目": "核对结果", "内容": "一致" if recon.get("consistent") is True else ("不一致" if recon.get("consistent") is False else "统计单位不同")},
        {"项目": "生成时间", "内容": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ], ["项目", "内容"])
    columns = [item["key"] for item in data["columns"]]
    labels = {item["key"]: item["label"] for item in data["columns"]}
    export_rows = [{labels[key]: row.get(key) for key in columns} for row in rows]
    _write_excel_sheet(workbook, "明细数据", export_rows, [labels[key] for key in columns])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _csv_cell(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return "" if value is None else value


def _write_raw_csv_section(writer, title: str, rows: List[dict]) -> None:
    writer.writerow([])
    writer.writerow([f"===== {title} ====="])
    if not rows:
        writer.writerow(["(无数据)"])
        return
    columns = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_csv_cell(row.get(column)) for column in columns])


def get_export_csv(
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> str:
    """Export scoped raw database/API data as UTF-8-SIG CSV for Excel."""
    import csv, io
    start, end, normalized_period = resolve_period(period, start_date, end_date)
    with database("dashboard.export") as conn:
        scope = _query_qx_raw_scope(conn, start, end, region, aftersaler, category, key_account)
        project_codes = scope["project_codes"]
        mapping_snapshot = aftersaler_mapping.get_mapping_snapshot(end, conn=conn)

    lims_records_by_code, lims_stats = fetch_lims_base_data(project_codes)
    lims_rows = []
    for requested_code, records in sorted(lims_records_by_code.items()):
        for record in records:
            row = {"_requested_project_code": requested_code}
            row.update(record)
            normalized = normalize_lims_api_record(record, requested_code, mapping_snapshot)
            row["LIMS原始售后"] = normalized.get("raw_aftersaler") or ""
            row["最终售后"] = normalized.get("final_aftersaler") or ""
            row["售后确定来源"] = normalized.get("aftersaler_source") or "lims_fallback"
            lims_rows.append(row)

    output = io.StringIO()
    output.write("\ufeff")  # BOM for Excel
    w = csv.writer(output)
    w.writerow(["===== 导出信息 ====="])
    w.writerow(["导出类型", "当前筛选范围原始数据"])
    w.writerow(["周期", normalized_period, "开始日期", start.isoformat(), "结束日期", end.isoformat()])
    w.writerow(["筛选条件", f"区域={region or '(全部)'}", f"售后={aftersaler or '(全部)'}",
                 f"产品={category or '(全部)'}", f"重点客户={key_account or '(全部)'}"])
    w.writerow(["命中群聊数", len(scope["group_names"]), "命中项目号数", len(project_codes)])
    w.writerow(["LIMS API请求", lims_stats.get("requests", 0), "返回记录", lims_stats.get("records", 0), "错误", lims_stats.get("errors", 0)])

    _write_raw_csv_section(w, "qx_analysis_result 原始记录", scope["raw_rows"])
    _write_raw_csv_section(
        w,
        f"{UNANSWERED_RESULT_TABLE} 夜间漏回原始记录",
        scope.get("raw_unanswered_rows", []),
    )
    _write_raw_csv_section(w, "qx_group 原始记录", scope["group_rows"])
    _write_raw_csv_section(w, "qx_chat 原始记录", scope["chat_rows"])
    _write_raw_csv_section(w, "LIMS base_data API 原始返回", lims_rows)

    return output.getvalue()


def _excel_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _write_excel_sheet(workbook, title: str, rows: List[dict], columns: Optional[List[str]] = None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(title=title[:31])
    if columns is None:
        columns = []
        seen = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    columns.append(key)
    if not columns:
        sheet.append(["无数据"])
        return sheet

    sheet.append(columns)
    for row in rows:
        sheet.append([_excel_cell(row.get(column)) for column in columns])

    header_fill = PatternFill("solid", fgColor="176B5B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, 1):
        values = [str(column)] + [str(_excel_cell(row.get(column)) or "") for row in rows[:200]]
        width = min(42, max(10, max(len(value) for value in values) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    return sheet


def _cross_rows(items: List[dict]) -> List[dict]:
    return [
        {
            "销售区域": row.get("region"),
            "关联群数": row.get("group_count"),
            "消息数": row.get("message_count"),
            "主要分布": row.get("items", []),
        }
        for row in items
    ]


def _top_message_group_export_rows(data: dict) -> List[dict]:
    return [
        {
            "排名": group.get("rank"),
            "群聊名称": group.get("group_name"),
            "消息数": group.get("message_count"),
            "全部消息占比(%)": group.get("percentage_of_all"),
            "活跃天数": group.get("active_days"),
            "高频问题Top5": "；".join(
                f"{item.get('word')}({item.get('count')})"
                for item in group.get("high_frequency_top5") or []
            ),
            "项目号": "、".join(group.get("project_codes") or []),
            "客户单位": "、".join(group.get("customer_units") or []),
            "客户名称": "、".join(group.get("customer_names") or []),
            "产品分类": "、".join(group.get("product_categories") or []),
            "售后员": "、".join(group.get("aftersalers") or []),
        }
        for group in data.get("items") or []
    ]


def _project_year_distribution_export_rows(data: dict) -> List[dict]:
    return [
        {
            "年份": item.get("year") if item.get("year") is not None else "未识别",
            "年份标签": item.get("label"),
            "项目数": item.get("project_count"),
            "项目占比(%)": item.get("percentage"),
            "关联群聊数": item.get("group_count"),
        }
        for item in data.get("items") or []
    ]


def _project_year_detail_export_rows(data: dict) -> List[dict]:
    rows = []
    for item in data.get("items") or []:
        for project in item.get("projects") or []:
            rows.append({
                "年份": project.get("year") if project.get("year") is not None else "未识别",
                "年份标签": item.get("label"),
                "项目号": project.get("project_code"),
                "年份来源": project.get("year_source"),
                "LIMS开始时间": project.get("start_time"),
                "项目或产品": "、".join(project.get("project_names") or []),
                "客户名称": "、".join(project.get("customer_names") or []),
                "客户单位": "、".join(project.get("work_units") or []),
                "产品分类": "、".join(project.get("product_categories") or []),
                "销售区域": "、".join(project.get("regions") or []),
                "售后人员": "、".join(project.get("aftersalers") or []),
                "关联群聊数": project.get("group_count"),
                "群聊名称": "、".join(project.get("group_names") or []),
            })
    return rows


def get_export_excel(
    period: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    region: str = "",
    aftersaler: str = "",
    category: str = "",
    key_account: str = "",
) -> bytes:
    """Export filtered dashboard modules and source data into separate XLSX sheets."""
    from io import BytesIO
    from openpyxl import Workbook

    start, end, normalized_period = resolve_period(period, start_date, end_date)
    overview = get_overview(
        period=period,
        start_date=start_date,
        end_date=end_date,
        region=region,
        aftersaler=aftersaler,
        category=category,
        key_account=key_account,
    )
    with database("dashboard.export") as conn:
        scope = _query_qx_raw_scope(conn, start, end, region, aftersaler, category, key_account)
        project_codes = scope["project_codes"]
        mapping_snapshot = aftersaler_mapping.get_mapping_snapshot(end, conn=conn)

    lims_records_by_code, lims_stats = fetch_lims_base_data(project_codes)
    lims_rows = []
    for requested_code, records in sorted(lims_records_by_code.items()):
        for record in records:
            normalized = normalize_lims_api_record(record, requested_code, mapping_snapshot)
            if normalized.get("category_l2") not in ALLOWED_PRODUCT_L2:
                continue
            if category and normalized.get("category_l2") != category:
                continue
            if region and normalized.get("region") != region:
                continue
            if key_account and normalized.get("key_account") != key_account:
                continue
            if aftersaler and normalized.get("final_aftersaler") != aftersaler:
                continue
            row = {"_requested_project_code": requested_code}
            row.update(record)
            row["LIMS原始售后"] = normalized.get("raw_aftersaler") or ""
            row["最终售后"] = normalized.get("final_aftersaler") or ""
            row["售后确定来源"] = normalized.get("aftersaler_source") or "lims_fallback"
            lims_rows.append(row)

    workbook = Workbook()
    workbook.remove(workbook.active)
    allowed_text = "、".join(sorted(ALLOWED_PRODUCT_L2))
    _write_excel_sheet(workbook, "导出说明", [
        {"项目": "统计周期", "内容": normalized_period},
        {"项目": "开始日期", "内容": start.isoformat()},
        {"项目": "结束日期", "内容": end.isoformat()},
        {"项目": "销售区域", "内容": region or "全部区域"},
        {"项目": "售后员", "内容": aftersaler or "全部售后"},
        {"项目": "产品类别", "内容": category or f"全部产品（仅{allowed_text}）"},
        {"项目": "重点客户", "内容": key_account or "全部客户"},
        {"项目": "命中群聊数", "内容": len(scope["group_names"])},
        {"项目": "命中项目号数", "内容": len(project_codes)},
        {"项目": "LIMS API请求", "内容": lims_stats.get("requests", 0)},
        {"项目": "LIMS返回记录", "内容": lims_stats.get("records", 0)},
        {"项目": "LIMS错误", "内容": lims_stats.get("errors", 0)},
    ], ["项目", "内容"])

    summary_labels = {
        "total_groups": "服务群数", "total_messages": "沟通消息数",
        "project_groups": "项目群数", "regions": "销售区域数",
        "aftersaler_count": "售后人员数", "product_categories": "产品分类数",
        "key_accounts": "重点客户数", "short_active_ratio": "短周期群占比",
    }
    _write_excel_sheet(workbook, "经营摘要", [
        {"指标": summary_labels.get(key, key), "值": value}
        for key, value in overview["summary"].items()
        if key in summary_labels
    ], ["指标", "值"])
    _write_excel_sheet(workbook, "消息趋势", overview["communication"]["trend"], ["date", "messages", "groups", "missed"])
    _write_excel_sheet(workbook, "高频关注主题", overview["communication"]["high_frequency"], ["word", "count"])
    _write_excel_sheet(workbook, "群活跃周期", overview["communication"]["active_duration"], ["range", "label", "count", "percentage"])
    _write_excel_sheet(workbook, "销售区域覆盖", overview["business"]["regions"])
    _write_excel_sheet(workbook, "售后人员分布", overview["business"]["aftersalers"])
    _write_excel_sheet(workbook, "产品分类", overview["business"]["product_categories"])
    _write_excel_sheet(workbook, "重点客户", overview["business"]["key_accounts"])
    project_year_data = overview["business"].get("project_year_distribution") or {}
    _write_excel_sheet(
        workbook, "项目年份分布",
        _project_year_distribution_export_rows(project_year_data),
    )
    _write_excel_sheet(
        workbook, "项目年份明细",
        _project_year_detail_export_rows(project_year_data),
    )
    _write_excel_sheet(
        workbook,
        "消息Top5群聊",
        _top_message_group_export_rows(
            overview["communication"].get("top_message_groups") or {}
        ),
    )
    _write_excel_sheet(workbook, "项目状态关注", overview["project_attention"]["items"])
    _write_excel_sheet(workbook, "交叉分析-销售", _cross_rows(overview["cross_analysis"]["region_sales"]))
    _write_excel_sheet(workbook, "交叉分析-售后", _cross_rows(overview["cross_analysis"]["region_after"]))
    _write_excel_sheet(workbook, "交叉分析-产品", _cross_rows(overview["cross_analysis"]["region_product"]))

    service_quality = overview["service_quality"]
    _write_excel_sheet(workbook, "服务质量", [
        {"模块": "消息漏回", "指标": key, "值": value}
        for key, value in service_quality["unanswered"].items()
    ] + [
        {"模块": "情感与服务态度", "指标": key, "值": value}
        for key, value in service_quality["sentiment"].items()
    ], ["模块", "指标", "值"])
    _write_excel_sheet(workbook, "数据质量", [
        {"指标": key, "值": value} for key, value in overview["data_quality"].items()
    ], ["指标", "值"])
    _write_excel_sheet(workbook, "analysis原始数据", scope["raw_rows"])
    _write_excel_sheet(workbook, "夜间漏回原始数据", scope.get("raw_unanswered_rows", []))
    _write_excel_sheet(workbook, "group原始数据", scope["group_rows"])
    _write_excel_sheet(workbook, "chat原始数据", scope["chat_rows"])
    _write_excel_sheet(workbook, "LIMS原始数据", lims_rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
